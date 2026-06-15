"""扫描分排期Excel文件，生成 货号→分排期归属 映射表（sub_schedule_map.json）
用法：python scan_schedules.py
输出：data/sub_schedule_map.json
"""
import os, sys, json, re, logging

logging.basicConfig(level=logging.INFO, format='%(message)s')

# 扫描源目录（本地副本）
SCAN_DIR = r'C:\Users\Administrator\Desktop\ZURU排期扫描'

# ITEM#列表头关键词（用于自动检测列位置，不含SKU因为SKU列是PO行号）
ITEM_KEYWORDS = ['ITEM#', 'ITEM＃', '货号']

# 跳过的sheet名关键词
SKIP_SHEET_KW = ['MA', '取消', '旧', '总排期', '汇总', 'Sheet']


def _should_skip_sheet(name):
    """判断sheet是否应跳过"""
    n = name.strip()
    n_upper = n.upper()
    # 含"MA"的跳过（精确：末尾MA、MA+空格、MA+中文等）
    if 'MA' in n_upper:
        return True
    for kw in SKIP_SHEET_KW:
        if kw == 'MA':
            continue
        if kw in n:
            return True
    # 纯Sheet1/Sheet2等默认sheet跳过
    if re.match(r'^Sheet\d*$', n, re.I):
        return True
    return False


def _detect_item_col(ws, max_row=5, max_col=40):
    """自动检测ITEM#列位置（1-based），扫描前几行表头"""
    for row in ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=False):
        for cell in row:
            v = str(cell.value or '').strip()
            if not v:
                continue
            v_upper = v.upper().replace(' ', '').replace('\n', '')
            # 精确匹配ITEM#或货号
            if v_upper in ('ITEM#', 'ITEM＃', '货号'):
                return cell.column
            # 模糊匹配：含ITEM和#
            if 'ITEM' in v_upper and ('#' in v_upper or '＃' in v_upper):
                return cell.column
    return None


def _is_valid_item(s):
    """判断是否为有效货号（排除表头、汇总行等）"""
    s = str(s).strip()
    if not s or len(s) < 3 or len(s) > 50:
        return False
    s_upper = s.upper()
    # 排除汇总/表头关键词
    skip_kw = ('TOTAL', 'SUBTOTAL', '合计', '小计', 'ITEM', '货号',
               'N/A', 'TBD', 'NONE', '---', '合並', '合并')
    for kw in skip_kw:
        if kw in s_upper:
            return False
    # 至少包含一个数字
    if not re.search(r'\d', s):
        return False
    return True


def _normalize_item(s):
    """标准化货号：去空格，统一大写"""
    s = str(s).strip()
    # 去掉规格码 -S001 等，只保留基础货号用于归属匹配
    m = re.match(r'^(.+?)(-S\d+.*)$', s, re.I)
    if m:
        return m.group(1).upper()
    return s.upper()


def scan():
    try:
        import openpyxl
    except ImportError:
        logging.error("需要安装openpyxl: pip install openpyxl")
        sys.exit(1)

    if not os.path.isdir(SCAN_DIR):
        logging.error(f"扫描目录不存在: {SCAN_DIR}")
        sys.exit(1)

    # 收集所有xlsx文件
    files = [f for f in os.listdir(SCAN_DIR) if f.endswith('.xlsx') and not f.startswith('~$')]
    logging.info(f"找到 {len(files)} 个排期文件")

    # 结果：{货号: [{file, sheet}]}
    result = {}
    stats = {'files': 0, 'sheets': 0, 'items': 0, 'skipped_sheets': []}

    for fname in sorted(files):
        fpath = os.path.join(SCAN_DIR, fname)
        logging.info(f"\n扫描: {fname}")
        stats['files'] += 1

        try:
            wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
        except Exception as e:
            logging.warning(f"  打开失败: {e}")
            continue

        for sname in wb.sheetnames:
            if _should_skip_sheet(sname):
                stats['skipped_sheets'].append(f"{fname} / {sname}")
                continue

            try:
                ws = wb[sname]
            except Exception:
                continue

            # 自动检测ITEM#列
            item_col = _detect_item_col(ws)
            if item_col is None:
                logging.info(f"  [{sname}] 未检测到ITEM#列，跳过")
                continue

            logging.info(f"  [{sname}] ITEM#列={item_col}")
            stats['sheets'] += 1

            # 收集该sheet的所有货号
            sheet_items = set()
            try:
                for row in ws.iter_rows(min_row=2, max_col=item_col, values_only=False):
                    cells = list(row)
                    if len(cells) < item_col:
                        continue
                    cell_val = cells[item_col - 1].value
                    if cell_val is None:
                        continue
                    raw = str(cell_val).strip()
                    if _is_valid_item(raw):
                        normalized = _normalize_item(raw)
                        if normalized:
                            sheet_items.add(normalized)
            except (ValueError, Exception) as e:
                # WPS文件可能触发openpyxl异常
                logging.warning(f"  [{sname}] 读取中断: {e}，已收集 {len(sheet_items)} 个货号")

            # 写入结果
            for item in sheet_items:
                entry = {'file': fname, 'sheet': sname}
                if item not in result:
                    result[item] = [entry]
                else:
                    # 去重（同文件同sheet不重复添加）
                    if not any(e['file'] == fname and e['sheet'] == sname for e in result[item]):
                        result[item].append(entry)
                stats['items'] += 1

        wb.close()

    # 保存
    out_dir = os.path.join(os.path.dirname(__file__), 'data')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'sub_schedule_map.json')

    with open(out_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    logging.info(f"\n{'='*50}")
    logging.info(f"扫描完成:")
    logging.info(f"  文件数: {stats['files']}")
    logging.info(f"  有效sheet数: {stats['sheets']}")
    logging.info(f"  唯一货号数: {len(result)}")
    logging.info(f"  跳过sheet数: {len(stats['skipped_sheets'])}")
    logging.info(f"  保存到: {out_path}")

    # 显示多归属货号（出现在2个以上文件中）
    multi = {k: v for k, v in result.items() if len(v) > 1}
    if multi:
        logging.info(f"\n{len(multi)} 个货号出现在多个排期中:")
        for item, locs in sorted(multi.items())[:20]:
            loc_str = '; '.join(f"{l['file']}[{l['sheet']}]" for l in locs)
            logging.info(f"  {item}: {loc_str}")
        if len(multi) > 20:
            logging.info(f"  ...还有 {len(multi)-20} 个")

    return out_path, result


if __name__ == '__main__':
    scan()
