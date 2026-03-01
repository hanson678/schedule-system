# -*- coding: utf-8 -*-
"""Excel PO解析器 — 读取WPS转换后的ZURU PO Excel文件
返回与PDFParser.parse()完全相同的dict格式，实现无缝替换"""
import os, re
import openpyxl


def _normalize_date(s):
    """将各种日期格式统一为YYYY-MM-DD"""
    if not s:
        return ''
    s = str(s).strip().replace('/', '-')
    # YYYY-MM-DD
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # DD-MM-YYYY or MM-DD-YYYY
    m = re.match(r'(\d{1,2})-(\d{1,2})-(\d{4})', s)
    if m:
        a, b, year = int(m.group(1)), int(m.group(2)), m.group(3)
        if a > 12:
            return f"{year}-{b:02d}-{a:02d}"
        elif b > 12:
            return f"{year}-{a:02d}-{b:02d}"
        else:
            return f"{year}-{a:02d}-{b:02d}"
    # datetime对象
    from datetime import datetime
    if isinstance(s, datetime):
        return s.strftime('%Y-%m-%d')
    return str(s)


# 国家英文→中文映射
_COUNTRY_MAP = {
    'usa': '美国', 'us': '美国', 'united states': '美国',
    'france': '法国', 'fr': '法国',
    'germany': '德国', 'de': '德国',
    'uk': '英国', 'gb': '英国', 'united kingdom': '英国',
    'australia': '澳大利亚', 'au': '澳大利亚',
    'canada': '加拿大', 'ca': '加拿大',
    'japan': '日本', 'jp': '日本',
    'netherlands': '荷兰', 'nl': '荷兰',
    'spain': '西班牙', 'italy': '意大利',
    'slovakia': '斯洛伐克', 'czech republic': '捷克',
    'poland': '波兰', 'new zealand': '新西兰',
    'south korea': '韩国', 'korea': '韩国',
    'mexico': '墨西哥', 'brazil': '巴西',
    'india': '印度', 'south africa': '南非',
    'china': '中国', 'hong kong': '香港', 'taiwan': '台湾',
    'singapore': '新加坡', 'malaysia': '马来西亚',
    'thailand': '泰国', 'indonesia': '印度尼西亚',
    'russia': '俄罗斯', 'russian fed': '俄罗斯', 'russian federation': '俄罗斯',
    'turkey': '土耳其', 'uae': '阿联酋',
    'sweden': '瑞典', 'norway': '挪威', 'denmark': '丹麦',
    'finland': '芬兰', 'belgium': '比利时', 'austria': '奥地利',
    'switzerland': '瑞士', 'portugal': '葡萄牙',
    'greece': '希腊', 'ireland': '爱尔兰',
    'chile': '智利', 'argentina': '阿根廷',
    'romania': '罗马尼亚', 'hungary': '匈牙利',
    'croatia': '克罗地亚', 'israel': '以色列',
}


def _country_cn(c):
    if not c:
        return ''
    cl = c.strip().lower().split(',')[0].strip().rstrip('.')
    result = _COUNTRY_MAP.get(cl, c.split(',')[0].strip())
    # 中文简称统一（"俄罗斯联邦"→"俄罗斯"等）
    _CN_SIMPLIFY = {'俄罗斯联邦': '俄罗斯', '大韩民国': '韩国', '阿拉伯联合酋长国': '阿联酋'}
    return _CN_SIMPLIFY.get(result, result)


def _to_float(v):
    """安全转float：处理逗号分隔、空格、字符串等"""
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).replace(',', '').strip())
    except (ValueError, TypeError):
        return 0.0


def _to_int(v):
    """安全转int：处理逗号分隔、空格、字符串等"""
    return int(_to_float(v))


class ExcelPOParser:
    """解析WPS转换后的ZURU PO Excel文件"""

    def parse(self, excel_path):
        wb = openpyxl.load_workbook(excel_path, data_only=True, read_only=True)
        ws = wb.active

        # 1. 收集所有单元格文本（用于正则匹配头部信息）
        all_text = self._collect_text(ws)

        # 2. 解析头部信息
        header = self._parse_header(all_text)

        # 3. 解析商品行
        lines = self._parse_lines(ws)

        # 4. 解析备注/包装信息
        reqs = self._parse_requirements(all_text)

        # 5. 检测取消单
        is_cancel = self._detect_cancel(all_text)

        wb.close()
        return {**header, 'lines': lines, **reqs,
                'is_cancel': is_cancel, 'raw_text': all_text[:8000]}

    def _collect_text(self, ws):
        """将所有单元格值拼接为文本（模拟PDF提取文本）"""
        rows_text = []
        for row in ws.iter_rows(max_col=30, max_row=200):
            cells = []
            for c in row:
                v = c.value
                if v is not None:
                    cells.append(str(v).strip())
            if cells:
                rows_text.append('  '.join(cells))
        return '\n'.join(rows_text)

    def _parse_header(self, text):
        """用正则从文本中提取头部字段（与PDFParser._header逻辑一致）"""
        def f(p, default=''):
            m = re.search(p, text, re.IGNORECASE)
            return m.group(1).strip() if m else default

        po = (f(r'PO#[:\s]*(4500\d{6})') or
              f(r'Purchase\s+Order[:\s#]*(\d{10})') or
              f(r'PO\s+Number[:\s]*(4\d{9})') or
              f(r'PO[:\s]*#?\s*(4500\d{6})') or
              f(r'Order\s+No\.?[:\s]*(4500\d{6})'))

        dest_raw = (f(r'Destination\s+Country[:\s]*(.+?)(?:\s{2,}|\n)') or
                    f(r'Ship\s+To\s+Country[:\s]*(.+?)(?:\s{2,}|\n)') or
                    f(r'Destination[:\s]*(.+?)(?:\s{2,}|\n)'))

        ship_date = (f(r'Shipment\s+Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                     f(r'Shipment\s+Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})') or
                     f(r'Ship\s+Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                     f(r'Ship\s+Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})') or
                     f(r'Delivery\s+Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                     f(r'Delivery\s+Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})'))

        customer = (f(r'Customer\s+Name[:\s]*(.+?)(?:\s{2,}|Payment|Loading|Shipment|\n)') or
                    f(r'Sold\s+To[:\s]*(.+?)(?:\s{2,}|Payment|Loading|Shipment|\n)') or
                    f(r'Bill\s+To[:\s]*(.+?)(?:\s{2,}|Payment|Loading|Shipment|\n)'))

        po_date_raw = (f(r'(?<![a-zA-Z])Date[:\s]*(\d{4}[-/]\d{1,2}[-/]\d{1,2})') or
                       f(r'(?<![a-zA-Z])Date[:\s]*(\d{1,2}[-/]\d{1,2}[-/]\d{4})'))

        cpo = f(r'Customer\s+PO#?[:\s]*(.+?)(?:\s{2,}|Loading|Shipment|Payment|\n)')
        if cpo and re.match(r'(?:Loading|Shipment|Payment|Ship|Destination|Supplier)', cpo, re.I):
            cpo = ''

        from_person = f(r'From[:\s]*(.+?)(?:\s{2,}|\n)') or f(r'Contact[:\s]*(.+?)(?:\s{2,}|\n)')

        return {
            'po_number': po,
            'po_date': _normalize_date(po_date_raw),
            'customer': customer,
            'customer_po_header': cpo,
            'from_person': from_person,
            'ship_date': _normalize_date(ship_date),
            'ship_type': f(r'Shipment\s+Type[:\s]*(.+?)(?:\s{2,}|\n)'),
            'sales_order': f(r'Sales\s+Order#?[:\s]*(\d+)'),
            'destination': dest_raw,
            'destination_cn': _country_cn(dest_raw),
            'loading_port': f(r'Loading\s+Port[:\s]*(.+?)(?:\s{2,}|\n)'),
        }

    def _parse_lines(self, ws):
        """从Excel中找到商品行表格并解析"""
        # 1. 找表头行：搜索含"SKU"或"Line"的行
        header_row = None
        col_map = {}  # key: field_name, value: col_index (0-based)

        for row in ws.iter_rows(max_col=30, max_row=100):
            vals = []
            for c in row:
                v = str(c.value).strip().upper() if c.value else ''
                vals.append(v)

            # 检测表头行标志
            has_sku = any('SKU' in v for v in vals)
            has_line = any(v == 'LINE' or v == 'LINE NO' for v in vals)
            has_qty = any('QTY' in v for v in vals)

            if has_sku and (has_qty or has_line):
                header_row = row[0].row
                # 建立列映射
                for i, v in enumerate(vals):
                    vu = v.upper()
                    if vu == 'LINE' or vu == 'LINE NO':
                        col_map['line_no'] = i
                    elif 'SKU' in vu and 'SPEC' not in vu:
                        col_map['sku'] = i
                    elif 'SKU' in vu and 'SPEC' in vu:
                        col_map['sku_spec'] = i
                    elif vu == 'NAME' or '品名' in vu or 'PRODUCT' in vu:
                        col_map['name'] = i
                    elif 'BOM' in vu:
                        col_map['bom'] = i
                    elif 'BARCODE' in vu or 'UPC' in vu or 'EAN' in vu:
                        col_map['barcode'] = i
                    elif 'DELIVERY' in vu or ('DATE' in vu and 'PO' not in vu):
                        col_map['delivery'] = i
                    elif 'PRICE' in vu and 'TOTAL' not in vu and 'STATUS' not in vu:
                        col_map['price'] = i
                    elif 'QTY' in vu and 'PCS' in vu:
                        col_map['qty'] = i
                    elif 'QTY' in vu:
                        col_map.setdefault('qty', i)
                    elif vu.startswith('TOTAL') and 'USD' in vu:
                        col_map['total_usd'] = i
                    elif vu.startswith('TOTAL') and ('CTN' in vu or 'CARTON' in vu):
                        col_map['total_ctns'] = i
                    elif vu.startswith('TOTAL') and 'CBM' in vu:
                        col_map['total_cbm'] = i
                    elif 'CUSTOMER' in vu and 'PO' in vu:
                        col_map['customer_po'] = i
                    elif vu == 'OUTER' or 'OUTER QTY' in vu:
                        col_map['outer_qty'] = i
                break

        if not header_row or 'sku' not in col_map:
            return []

        # 2. 从表头下一行开始读取数据
        lines = []
        for row in ws.iter_rows(min_row=header_row + 1, max_col=30, max_row=300):
            vals = [c.value for c in row]
            # 空行或总计行 → 停止
            non_empty = [v for v in vals[:20] if v is not None and str(v).strip()]
            if not non_empty:
                continue
            first_text = ' '.join(str(v) for v in non_empty).upper()
            if 'TOTAL' in first_text and ('USD' in first_text or 'CTN' in first_text or 'CBM' in first_text):
                break

            # 读取SKU — 必须有SKU才算有效行
            sku_val = vals[col_map['sku']] if col_map.get('sku') is not None else None
            if not sku_val or not str(sku_val).strip():
                continue

            sku = str(sku_val).strip()

            # 解析各字段
            def get(key, default=''):
                idx = col_map.get(key)
                if idx is None:
                    return default
                v = vals[idx]
                if v is None:
                    return default
                return v

            line_no = str(get('line_no', '')).strip()
            sku_spec = str(get('sku_spec', sku)).strip()
            name = str(get('name', '')).strip()
            barcode_raw = get('barcode', '')
            barcode = str(int(barcode_raw)) if isinstance(barcode_raw, float) else str(barcode_raw).strip()

            delivery_raw = get('delivery', '')
            from datetime import datetime
            if isinstance(delivery_raw, datetime):
                delivery = delivery_raw.strftime('%Y-%m-%d')
            else:
                delivery = _normalize_date(str(delivery_raw))

            price = _to_float(get('price', 0))
            qty = _to_int(get('qty', 0))
            outer_qty = _to_int(get('outer_qty', 0))
            total_usd = _to_float(get('total_usd', 0))

            total_ctns = _to_int(get('total_ctns', 0))

            cpo = str(get('customer_po', '')).strip()
            # 清理customer_po：排除CBM等数字误取
            if cpo and re.match(r'^\d+\.\d+$', cpo):
                cpo = ''

            lines.append({
                'line_no': line_no,
                'sku': sku,
                'item_code': sku_spec if sku_spec != sku else sku,
                'sku_spec': sku_spec if sku_spec != sku else sku,
                'name': name,
                'barcode': barcode,
                'delivery': delivery,
                'price': price,
                'qty': qty,
                'outer_qty': outer_qty,
                'total_usd': total_usd,
                'total_ctns': total_ctns,
                'customer_po': cpo,
            })

        return lines

    def _parse_requirements(self, text):
        """提取备注、包装信息、修订记录"""
        def ext(p, lim=2000):
            m = re.search(p, text, re.DOTALL | re.I)
            return m.group(1).strip()[:lim] if m else ''

        tracking = ''
        m = re.search(r'日期码格式[：:]\s*(.*?)\s*日期[：:]\s*(.*?)(?:\n)', text)
        if m:
            tracking = f'日期码格式：{m.group(1).strip()} 日期：{m.group(2).strip()}'

        packaging = ext(r'Packaging\s+Info[：:\s]*(.*?)(?=Remark[：:\s]|$)')
        remark = ext(r'Remark[：:\s]*(.*?)(?=Order Modifiable|$)', 3000)

        revision = ''
        rev_m = re.search(r'Order\s+Modifiable\s+Records\s*(.*?)(?=Special|Additional|Confirmed|$)',
                          text, re.DOTALL | re.I)
        if rev_m:
            entries = []
            for line in rev_m.group(1).strip().split('\n'):
                line = line.strip()
                if not line or ('Revision' in line and '#' in line):
                    continue
                rm = re.match(r'(\d+)\s+(\d{2}-\d{2}-\d{4})\s+(.*)', line)
                if rm:
                    entries.append(f"Rev.{rm.group(1)} ({_normalize_date(rm.group(2))}): {rm.group(3).strip()}")
            if entries:
                revision = '; '.join(entries)

        return {'tracking_code': tracking, 'packaging_info': packaging,
                'remark': remark, 'revision': revision}

    def _detect_cancel(self, text):
        """检测取消单"""
        clean = re.sub(r'(?:Remark|Packaging\s+Info|备注)[：:\s].*', '', text,
                       flags=re.DOTALL | re.I)
        return '取消' in clean or '取 消' in clean
