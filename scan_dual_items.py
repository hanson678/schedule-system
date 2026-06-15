"""
扫描Z盘排期文件，找出目标货号在不同sheet中的实际ITEM#列值。
"""
import json
import openpyxl
import re
import sys
import io
from collections import defaultdict

# 解决Windows终端编码问题
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

MAP_PATH = r"C:\Users\Administrator\Desktop\排期系统-总排期版\data\item_schedule_map.json"
TARGET_NUMS = ["77785", "77758", "77721", "92123", "92124", "77896", "92149"]

def main():
    with open(MAP_PATH, "r", encoding="utf-8") as f:
        mapping = json.load(f)

    # 第1步：按目标货号找出所有匹配的key及对应文件/sheet
    # key中包含目标数字即匹配
    targets_info = defaultdict(list)  # {目标货号: [{key, file, sheet, path}, ...]}

    for sku_key, entries in mapping.items():
        for tgt in TARGET_NUMS:
            if tgt in sku_key:
                for entry in entries:
                    targets_info[tgt].append({
                        "key": sku_key,
                        "file": entry.get("file", ""),
                        "sheet": entry.get("sheet", ""),
                        "path": entry.get("path", ""),
                    })

    # 打印匹配的key概览
    print("=" * 80)
    print("第1步：item_schedule_map.json 中匹配的key")
    print("=" * 80)
    for tgt in TARGET_NUMS:
        entries = targets_info[tgt]
        if not entries:
            print(f"\n【{tgt}】无匹配key")
            continue
        # 去重key
        unique_keys = sorted(set(e["key"] for e in entries))
        print(f"\n【{tgt}】共 {len(unique_keys)} 个匹配key：")
        for k in unique_keys:
            print(f"  - {k}")

    # 第2步：收集需要打开的文件和sheet
    # {path: set(sheet_names)}
    file_sheets = defaultdict(set)
    for tgt, entries in targets_info.items():
        for e in entries:
            if e["path"] and e["sheet"]:
                # 跳过取消单sheet
                if "取消" in e["sheet"]:
                    continue
                file_sheets[e["path"]].add(e["sheet"])

    # 第3步：打开文件，搜索ITEM#列
    print("\n" + "=" * 80)
    print("第2步：在排期文件中搜索实际ITEM#列值")
    print("=" * 80)

    results = defaultdict(lambda: defaultdict(set))  # {目标货号: {sheet名: set(实际值)}}

    for filepath, sheets in sorted(file_sheets.items()):
        filename = filepath.split("\\")[-1].replace('\xa0', ' ')
        print(f"\n--- 打开文件: {filename} ---")
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
        except Exception as e:
            print(f"  [错误] 无法打开: {e}")
            continue

        for sheet_name in sheets:
            if sheet_name not in wb.sheetnames:
                # 尝试模糊匹配（去空格）
                found = False
                for sn in wb.sheetnames:
                    if sn.strip() == sheet_name.strip():
                        sheet_name = sn
                        found = True
                        break
                if not found:
                    print(f"  [跳过] sheet '{sheet_name}' 不存在")
                    continue

            ws = wb[sheet_name]
            print(f"  Sheet: {sheet_name}")

            # 先找ITEM#列位置（搜前10行的表头）
            item_col = None
            try:
                for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=15, values_only=False), 1):
                    for cell in row:
                        val = str(cell.value).strip() if cell.value is not None else ""
                        val_upper = val.upper()
                        if "ITEM" in val_upper and ("#" in val_upper or "NO" in val_upper or "号" in val_upper):
                            item_col = cell.column
                            print(f"    ITEM#列 = 第{item_col}列 (表头: '{val}', 行{row_idx})")
                            break
                        elif val_upper in ("ITEM#", "ITEM #", "ITEM", "货号"):
                            item_col = cell.column
                            print(f"    ITEM#列 = 第{item_col}列 (表头: '{val}', 行{row_idx})")
                            break
                    if item_col:
                        break
            except Exception as e:
                print(f"    [错误] 搜索表头时出错: {e}")
                continue

            if not item_col:
                # 兜底：扫描所有列找包含目标货号的列
                print(f"    未找到ITEM#表头，尝试暴力搜索...")
                # 遍历前50行所有列找含目标货号的
                try:
                    for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
                        for cell in row:
                            val = str(cell.value).strip() if cell.value is not None else ""
                            for tgt in TARGET_NUMS:
                                if tgt in val and len(val) < 50:  # 避免匹配到超长文本
                                    item_col = cell.column
                                    print(f"    推测ITEM#列 = 第{item_col}列 (样本: '{val}')")
                                    break
                            if item_col:
                                break
                        if item_col:
                            break
                except Exception:
                    pass

            if not item_col:
                print(f"    [跳过] 无法确定ITEM#列")
                continue

            # 搜索所有行
            found_count = 0
            try:
                for row in ws.iter_rows(min_row=1, values_only=False):
                    cell = row[item_col - 1] if len(row) >= item_col else None
                    if cell is None or cell.value is None:
                        continue
                    val = str(cell.value).strip()
                    for tgt in TARGET_NUMS:
                        if tgt in val:
                            results[tgt][sheet_name].add(val)
                            found_count += 1
            except Exception as e:
                print(f"    [错误] 遍历行时出错: {e}")

            print(f"    找到 {found_count} 条匹配记录")

        wb.close()

    # 第4步：汇总输出
    print("\n" + "=" * 80)
    print("汇总：每个货号在各sheet中的ITEM#实际值")
    print("=" * 80)

    for tgt in TARGET_NUMS:
        sheet_vals = results[tgt]
        if not sheet_vals:
            print(f"\n【{tgt}】未找到任何记录")
            continue
        print(f"\n【{tgt}】")
        for sheet_name in sorted(sheet_vals.keys()):
            vals = sorted(sheet_vals[sheet_name])
            print(f"  Sheet [{sheet_name}]:")
            for v in vals:
                print(f"    → {v}")

if __name__ == "__main__":
    main()
