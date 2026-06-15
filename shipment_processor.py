# -*- coding: utf-8 -*-
"""
出货标记核心处理程序

用法: python shipment_processor.py <接单表.xlsx> <出货文件夹>

规则：
1. 匹配：合同号 + 简货号（开头连续数字）
2. 可标记：备注列为空 + 数量>0
3. 顺序匹配可用行，标记出货日期到备注列
4. XML手术式写入，100%保留原文件格式
5. 输出：接单表_更新.xlsx
"""

import os
import re
import shutil
import logging
import zipfile
from io import BytesIO
from datetime import datetime

from lxml import etree

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

logger = logging.getLogger(__name__)

# ============================================================
# 常量
# ============================================================

# 卡板单后缀
SET_SUFFIX_RE = re.compile(r'^(.+?)(SLB|SLD|SLT|SK)$', re.IGNORECASE)

# 提取开头连续数字作为简货号
_LEADING_DIGITS_RE = re.compile(r'^(\d+)')

# 出货日期提取（如 "3月9日" 或 "3月9出"）
_DATE_RE = re.compile(r'(\d{1,2}月\d{1,2})[日出]')

# xlsx XML命名空间
_XLSX_NS = 'http://schemas.openxmlformats.org/spreadsheetml/2006/main'
_REL_NS = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

# 混装货号映射
MIXED_MAP = {
    '7153': ['7149', '7150'],
    '7154': ['7151', '7152'],
    '25257': ['25251', '25252', '25253'],
}

# 产品组辅助行关键词
AUX_KEYWORDS = ('收缩指商', 'PDQ')

def _today_serial():
    """返回今天的Excel日期序列号"""
    return (datetime.now() - datetime(1899, 12, 31)).days


# 需要清除的隐藏/特殊字符
_HIDDEN_CHARS = re.compile(
    r'[\x00-\x08\x0b\x0c\x0e-\x1f\x7f'
    r'\u00a0'
    r'\u200b\u200c\u200d\u200e\u200f'
    r'\u2028\u2029'
    r'\u202a-\u202e'
    r'\ufeff'
    r'\u3000'
    r']'
)


# ============================================================
# 工具函数
# ============================================================

def _normalize(val):
    """单元格值 → 干净字符串（用于匹配键）"""
    if val is None:
        return ''
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    if isinstance(val, int):
        return str(val)
    s = str(val)
    s = _HIDDEN_CHARS.sub('', s)
    s = s.strip()
    s = s.upper()
    return s


def _extract_simple_no(item_no):
    """提取开头连续数字作为简货号。15750A→15750, 92148B→92148"""
    m = _LEADING_DIGITS_RE.match(item_no)
    return m.group(1) if m else item_no


def _to_simple_key(item_no):
    """出货货号→简货号：保留SLB/SLD/SLT/SK后缀，取开头连续数字+后缀。
    15750A→15750, 77711SLB→77711SLB, 15779SLD→15779SLD"""
    m = SET_SUFFIX_RE.match(item_no)
    if m:
        base = _extract_simple_no(m.group(1))
        return base + m.group(2).upper()
    return _extract_simple_no(item_no)


def _is_set_item(item_no):
    """判断货号是否为卡板单（含SLB/SLD/SLT/SK后缀）"""
    return bool(SET_SUFFIX_RE.match(item_no))


def _strip_set_suffix(item_no):
    """去掉SLB/SLD/SLT/SK后缀"""
    m = SET_SUFFIX_RE.match(item_no)
    return m.group(1) if m else item_no


def _is_auxiliary_row(ws, row_idx, col_item, max_col):
    """判断是否为产品组辅助行（收缩指商、PDQ）"""
    item_val = str(ws.cell(row=row_idx, column=col_item).value or '').strip()
    for kw in AUX_KEYWORDS:
        if kw in item_val:
            return True
    for c in range(1, max_col + 1):
        val = ws.cell(row=row_idx, column=c).value
        if val is None:
            continue
        s = str(val).strip()
        for kw in AUX_KEYWORDS:
            if kw == s:
                return True
    return False


def _collect_available(ws, candidates, col_qty, col_beizhu, col_contract,
                       col_item, max_col, is_slb=False, log_func=None):
    """从候选行中筛选可标记行，返回 [(row, qty, needs_contract_fill), ...]
    普通货号：只标合同列不为空的行
    SLB货号：合同列为空但货号含SLB后缀也可标记（需补填合同号）
    """
    available = []
    for r in candidates:
        contract_val = ws.cell(row=r, column=col_contract).value
        has_contract = contract_val is not None and str(contract_val).strip()
        needs_fill = False

        if not has_contract:
            if is_slb:
                # SLB货号：检查该行货号是否也含SLB后缀
                full_item = _normalize(ws.cell(row=r, column=col_item).value)
                if _is_set_item(full_item):
                    needs_fill = True  # 需要补填合同号
                else:
                    if log_func:
                        log_func(f"    跳过行{r}: 非SLB子行不标记")
                    continue
            else:
                if log_func:
                    log_func(f"    跳过行{r}: 合同列为空（子行不标记）")
                continue
        if col_beizhu:
            bz = ws.cell(row=r, column=col_beizhu).value
            if bz is not None and str(bz).strip():
                bz_str = str(bz).strip()
                # 允许覆盖"额外费用"
                if bz_str != '额外费用':
                    if log_func:
                        log_func(f"    排除行{r}: 备注不为空='{bz_str[:30]}'")
                    continue
        try:
            qty = int(float(ws.cell(row=r, column=col_qty).value))
        except (ValueError, TypeError):
            if log_func:
                log_func(f"    排除行{r}: 数量无效={ws.cell(row=r, column=col_qty).value}")
            continue
        if qty > 0:
            available.append((r, qty, needs_fill))
        else:
            if log_func:
                log_func(f"    排除行{r}: 数量={qty}≤0")
    return available


def _expand_mixed(shipments, dates):
    """展开混装货号。7153 qty=100 → 7149 qty+=100, 7150 qty+=100"""
    expanded = {}
    expanded_dates = {}
    for (contract, item_no), qty in shipments.items():
        if item_no in MIXED_MAP:
            components = MIXED_MAP[item_no]
            for comp in components:
                key = (contract, comp)
                expanded[key] = expanded.get(key, 0) + qty
                if (contract, item_no) in dates:
                    expanded_dates[key] = dates[(contract, item_no)]
            logger.info(f"  混装展开: {item_no} → "
                        f"{', '.join(components)}, 每个={qty}")
        else:
            key = (contract, item_no)
            expanded[key] = expanded.get(key, 0) + qty
            if key in dates:
                expanded_dates[key] = dates[key]
    return expanded, expanded_dates


# ============================================================
# XML手术式写入（100%保留格式）
# ============================================================

def _col_letter(col):
    """1-based列号 → Excel列字母。1→A, 14→N, 27→AA"""
    result = ''
    while col > 0:
        col, rem = divmod(col - 1, 26)
        result = chr(65 + rem) + result
    return result


def _find_sheet_xml_path(zf, sheet_title):
    """在xlsx ZIP中查找指定sheet名称对应的XML路径"""
    ns_wb = _XLSX_NS
    ns_r = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'

    try:
        wb_xml = zf.read('xl/workbook.xml')
        wb_root = etree.fromstring(wb_xml)

        rid = None
        for sheet_elem in wb_root.iter(f'{{{ns_wb}}}sheet'):
            if sheet_elem.get('name') == sheet_title:
                rid = sheet_elem.get(f'{{{ns_r}}}id')
                break

        if not rid:
            # 找不到匹配名称，取第一个sheet
            for sheet_elem in wb_root.iter(f'{{{ns_wb}}}sheet'):
                rid = sheet_elem.get(f'{{{ns_r}}}id')
                break

        if rid:
            rels_xml = zf.read('xl/_rels/workbook.xml.rels')
            rels_root = etree.fromstring(rels_xml)
            for rel in rels_root.iter(f'{{{_REL_NS}}}Relationship'):
                if rel.get('Id') == rid:
                    target = rel.get('Target')
                    if not target.startswith('/'):
                        return f'xl/{target}'
                    return target.lstrip('/')
    except Exception as e:
        logger.warning(f"查找sheet路径失败: {e}")

    return 'xl/worksheets/sheet1.xml'


def _write_cell_value(cell_elem, value, ns):
    """将单元格设置为数字或inline string，保留样式属性"""
    for child in list(cell_elem):
        cell_elem.remove(child)
    # 清理可能的公式相关属性
    for attr in ('cm',):
        if attr in cell_elem.attrib:
            del cell_elem.attrib[attr]

    if isinstance(value, (int, float)):
        # 数字值：<v>46094</v>
        if 't' in cell_elem.attrib:
            del cell_elem.attrib['t']
        v_elem = etree.SubElement(cell_elem, f'{{{ns}}}v')
        v_elem.text = str(value)
    else:
        # 文本值：inline string
        cell_elem.set('t', 'inlineStr')
        is_elem = etree.SubElement(cell_elem, f'{{{ns}}}is')
        t_elem = etree.SubElement(is_elem, f'{{{ns}}}t')
        t_elem.text = str(value)


def _surgical_xlsx_write(src_path, dst_path, sheet_title, cell_edits,
                         blue_rows=None):
    """
    复制src到dst，然后在dst中手术式修改指定单元格。

    cell_edits: [(row, col, value), ...] — 1-based行列号
    blue_rows: set of row numbers — 这些行整行蓝色填充
    """
    shutil.copy2(src_path, dst_path)

    if not cell_edits:
        return

    ns = _XLSX_NS

    # 按行分组编辑
    edits_map = {}
    for row_num, col_num, value in cell_edits:
        ref = f'{_col_letter(col_num)}{row_num}'
        edits_map.setdefault(row_num, {})[ref] = value

    temp_path = dst_path + '.tmp'

    with zipfile.ZipFile(dst_path, 'r') as zin:
        sheet_path = _find_sheet_xml_path(zin, sheet_title)
        sheet_xml = zin.read(sheet_path)

        # 解析XML（保留空白和格式）
        parser = etree.XMLParser(remove_blank_text=False)
        root = etree.fromstring(sheet_xml, parser)

        sheet_data = root.find(f'{{{ns}}}sheetData')
        if sheet_data is None:
            logger.warning("sheetData未找到，跳过写入")
            return

        # --- 在styles.xml中添加蓝色填充 ---
        styles_xml = zin.read('xl/styles.xml')
        styles_root = etree.fromstring(styles_xml, parser)

        # 添加蓝色填充 #00B0F0
        fills_elem = styles_root.find(f'{{{ns}}}fills')
        fill_wrapper = etree.SubElement(fills_elem, f'{{{ns}}}fill')
        pf = etree.SubElement(fill_wrapper, f'{{{ns}}}patternFill')
        pf.set('patternType', 'solid')
        fg = etree.SubElement(pf, f'{{{ns}}}fgColor')
        fg.set('rgb', 'FF00B0F0')
        bg = etree.SubElement(pf, f'{{{ns}}}bgColor')
        bg.set('indexed', '64')
        blue_fill_id = str(len(fills_elem) - 1)
        fills_elem.set('count', str(len(fills_elem)))

        # 为每种原始样式创建蓝色填充克隆
        cell_xfs = styles_root.find(f'{{{ns}}}cellXfs')
        blue_clone_map = {}  # 原样式idx → 蓝色克隆idx

        def _get_blue_clone(orig_s):
            """获取原样式的蓝色填充版本，缓存避免重复创建"""
            if orig_s in blue_clone_map:
                return blue_clone_map[orig_s]
            orig_xf = cell_xfs[int(orig_s)]
            new_xf = etree.SubElement(cell_xfs, f'{{{ns}}}xf')
            # 复制原样式的所有属性
            for attr, val in orig_xf.attrib.items():
                new_xf.set(attr, val)
            # 改fillId为蓝色
            new_xf.set('fillId', blue_fill_id)
            new_xf.set('applyFill', '1')
            new_idx = str(len(cell_xfs) - 1)
            blue_clone_map[orig_s] = new_idx
            return new_idx

        # 默认蓝色样式（无原始样式的单元格用）
        default_blue = _get_blue_clone('0')

        # --- 修改sheet数据 ---
        blue_set = blue_rows or set()
        target_rows = set(edits_map.keys()) | blue_set

        for row_elem in sheet_data.findall(f'{{{ns}}}row'):
            rn = int(row_elem.get('r'))
            if rn not in target_rows:
                continue

            pending = dict(edits_map.get(rn, {}))
            need_blue = rn in blue_set

            for cell_elem in row_elem.findall(f'{{{ns}}}c'):
                cell_ref = cell_elem.get('r')
                # 蓝色填充整行（保留原字体/数字格式）
                if need_blue:
                    orig_s = cell_elem.get('s', '0')
                    cell_elem.set('s', _get_blue_clone(orig_s))
                # 写入数据
                if cell_ref in pending:
                    _write_cell_value(cell_elem, pending.pop(cell_ref), ns)

            # 新建不存在的单元格
            for cell_ref, value in pending.items():
                cell_elem = etree.SubElement(row_elem, f'{{{ns}}}c')
                cell_elem.set('r', cell_ref)
                if need_blue:
                    cell_elem.set('s', default_blue)
                _write_cell_value(cell_elem, value, ns)

        cell_xfs.set('count', str(len(cell_xfs)))
        modified_styles = etree.tostring(
            styles_root, xml_declaration=True, encoding='UTF-8',
            standalone=True)

        # 序列化
        modified_bytes = etree.tostring(
            root, xml_declaration=True, encoding='UTF-8', standalone=True)

        # 写入新ZIP（含修改后的styles.xml）
        with zipfile.ZipFile(temp_path, 'w', zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                if item.filename == sheet_path:
                    zout.writestr(item, modified_bytes)
                elif item.filename == 'xl/styles.xml':
                    zout.writestr(item, modified_styles)
                else:
                    zout.writestr(item, zin.read(item.filename))

    os.replace(temp_path, dst_path)


# ============================================================
# 库存错误报告
# ============================================================

def _generate_error_report(errors, output_dir):
    """生成库存错误报告Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "库存错误报告"

    header_font = Font(name='Microsoft YaHei', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(patternType='solid', fgColor='C0392B')
    header_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    headers = ['合同号', '出货货号', '简货号', '出货数量', '可用库存', '缺口数量', '原因']
    col_widths = [18, 18, 14, 14, 14, 14, 32]

    for c, (title, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=c, value=title)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        if c <= 26:
            ws.column_dimensions[chr(64 + c)].width = width

    data_font = Font(name='Microsoft YaHei', size=10)
    red_font = Font(name='Microsoft YaHei', size=10, color='C0392B', bold=True)
    center_align = Alignment(horizontal='center')

    for i, err in enumerate(errors, 2):
        gap = err['ship_qty'] - err['available_qty']
        row_data = [
            err['contract'], err['item_no'], err.get('simple_no', ''),
            err['ship_qty'], err['available_qty'], gap, err['reason'],
        ]
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(row=i, column=c, value=val)
            cell.font = red_font if c == 6 else data_font
            cell.border = thin_border
            if c >= 4:
                cell.alignment = center_align

    report_path = os.path.join(output_dir, "库存错误报告.xlsx")
    wb.save(report_path)
    wb.close()
    return report_path


# ============================================================
# 读取接单表
# ============================================================

def detect_columns(ws, keywords):
    """动态检测表头列。额外检测'简货号'列。"""
    for r in range(1, min(11, ws.max_row + 1)):
        col_map = {}
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is None:
                continue
            s = str(val).strip().replace('\n', '')
            if '简货号' in s and '简货号' not in col_map:
                col_map['简货号'] = c
                continue
            for kw in keywords:
                if kw in s and kw not in col_map:
                    col_map[kw] = c
                    break
        if all(k in col_map for k in ['合同', '货号', '数量']):
            return r, col_map
    return None, None


# ============================================================
# 读取出货资料
# ============================================================

def _read_sheet_data(filepath):
    """读取Excel首个Sheet全部数据，支持 .xls 和 .xlsx"""
    ext = os.path.splitext(filepath)[1].lower()
    if ext == '.xls':
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheet_by_index(0)
        return [[ws.cell_value(i, j) for j in range(ws.ncols)]
                for i in range(ws.nrows)]
    if ext == '.xlsx':
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        data = [list(row) for row in ws.iter_rows(values_only=True)]
        wb.close()
        return data
    return []


def _extract_file_date(rows):
    """从出货文件前5行提取出货日期（如 "3月9日" 或 "3月9出"），统一返回 "X月X日" """
    for i in range(min(5, len(rows))):
        for val in rows[i]:
            if not isinstance(val, str):
                continue
            m = _DATE_RE.search(val)
            if m:
                return m.group(1) + '日'
    return None


def read_shipment_folder(folder):
    """
    读取出货文件夹内所有Excel，筛选备注=华登/兴信。
    返回: (shipments, dates)
        shipments: {(合同, 货号): 总出货量}
        dates: {(合同, 货号): 出货日期字符串}
    """
    shipments = {}
    dates = {}

    if not os.path.isdir(folder):
        logger.error(f"文件夹不存在: {folder}")
        return shipments, dates

    for fname in os.listdir(folder):
        if fname.startswith('~$'):
            continue
        if not fname.lower().endswith(('.xls', '.xlsx')):
            continue

        fpath = os.path.join(folder, fname)
        try:
            rows = _read_sheet_data(fpath)
        except Exception as e:
            logger.error(f"读取失败 {fname}: {e}")
            continue

        if not rows:
            continue

        file_date = _extract_file_date(rows)
        if file_date:
            logger.info(f"  {fname}: 出货日期={file_date}")

        target = ['备注', '合同', '货号', '数量']
        header_idx = None
        col_map = {}
        for i in range(min(10, len(rows))):
            vals = [str(v).strip().replace('\n', '') if v else ''
                    for v in rows[i]]
            found = {}
            for kw in target:
                for j, v in enumerate(vals):
                    if kw in v and kw not in found:
                        found[kw] = j
                        break
            if all(k in found for k in target):
                header_idx = i
                col_map = found
                break

        if header_idx is None:
            logger.warning(f"  {fname}: 未找到表头，跳过")
            continue

        def _get(row, idx):
            return row[idx] if idx < len(row) else None

        count = 0
        for i in range(header_idx + 1, len(rows)):
            row = rows[i]
            factory = str(_get(row, col_map['备注']) or '').strip()
            if factory not in ('华登', '兴信'):
                continue

            contract = _normalize(_get(row, col_map['合同']))
            item_no = _normalize(_get(row, col_map['货号']))
            if not contract or not item_no:
                continue

            try:
                qty = int(float(_get(row, col_map['数量'])))
            except (ValueError, TypeError):
                continue
            if qty <= 0:
                continue

            key = (contract, item_no)
            shipments[key] = shipments.get(key, 0) + qty
            if file_date:
                dates[key] = file_date
            count += 1

        logger.info(f"  {fname}: {count} 条有效出货记录")

    logger.info(f"出货汇总: {len(shipments)} 组合同+货号")
    return shipments, dates


# ============================================================
# 核心处理：匹配出货 → 标记备注
# ============================================================

def process(order_path, shipment_folder, output_path=None,
            mark_date='', log_callback=None):
    """
    出货标记主函数。

    参数:
        order_path: 接单表Excel路径
        shipment_folder: 出货资料文件夹路径
        output_path: 输出路径（默认 接单表_更新.xlsx）
        mark_date: 统一标记日期（如 "3月14日"），为空则用出货文件日期
        log_callback: 日志回调函数（用于GUI实时显示）

    返回: (output_path, stats) 成功 或 (None, error_msg) 失败
    """
    def log(msg):
        logger.info(msg)
        if log_callback:
            log_callback(msg)

    log("=" * 50)
    log("步骤1: 读取出货资料")
    log("=" * 50)

    shipments, dates = read_shipment_folder(shipment_folder)
    if not shipments:
        log("没有有效的出货数据")
        return None, "没有有效的出货数据"

    shipments, dates = _expand_mixed(shipments, dates)

    # 统一日期：如果指定了mark_date，所有条目用同一个日期
    if mark_date:
        log(f"  统一备注日期: {mark_date}")

    for (c, i), qty in shipments.items():
        simple = _to_simple_key(i)
        tag = " [卡板单]" if _is_set_item(i) else ""
        log(f"  合同={c}, 货号={i}→简={simple}{tag}, 数量={qty}")

    # ---- 读取接单表（只读模式，不修改原文件）----
    log("")
    log("=" * 50)
    log("步骤2: 读取接单表")
    log("=" * 50)

    try:
        wb = load_workbook(order_path, data_only=True)
    except Exception as e:
        log(f"无法打开接单表: {e}")
        return None, f"无法打开接单表: {e}"

    ws = wb.active
    sheet_title = ws.title
    header_row, col_map = detect_columns(
        ws, ['合同', '货号', '数量', '备注'])

    if header_row is None:
        log("接单表未找到有效表头（至少需要: 合同、货号、数量）")
        wb.close()
        return None, "接单表未找到有效表头"

    col_contract = col_map['合同']
    col_item = col_map['货号']
    col_qty = col_map['数量']
    col_beizhu = col_map.get('备注')
    col_simple = col_map.get('简货号', col_item)
    max_col = ws.max_column

    has_simple_col = '简货号' in col_map
    log(f"接单表: {ws.max_row} 行 × {max_col} 列, Sheet='{sheet_title}'")
    if has_simple_col:
        log(f"  ✓ 检测到简货号列: 第{col_simple}列")
    else:
        log(f"  ✗ 未检测到简货号列，将从货号列自动提取简货号")
    log(f"  表头行={header_row}, 合同列={col_contract}, "
        f"简货号列={col_simple}, 货号列={col_item}, "
        f"数量列={col_qty}, 备注列={col_beizhu}")

    # 建立行索引: (合同, 简货号) → [row_idx, ...]
    row_index = {}
    aux_skipped = 0
    last_contract = ''
    for r in range(header_row + 1, ws.max_row + 1):
        rc = _normalize(ws.cell(row=r, column=col_contract).value)
        if rc:
            last_contract = rc
        effective_contract = rc or last_contract

        ri_raw = _normalize(ws.cell(row=r, column=col_simple).value)
        if not effective_contract or not ri_raw:
            continue

        if _is_auxiliary_row(ws, r, col_item, max_col):
            aux_skipped += 1
            continue

        full_item = _normalize(ws.cell(row=r, column=col_item).value)

        # SLB/SLD/SLT/SK行也参与索引，key带后缀以区分
        ri = _extract_simple_no(ri_raw)
        m = SET_SUFFIX_RE.match(full_item)
        if m:
            ri = ri + m.group(2).upper()
        row_index.setdefault((effective_contract, ri), []).append(r)

    log(f"  行索引: {len(row_index)} 个合同+简货号组合")
    if aux_skipped:
        log(f"  跳过辅助行(收缩指商/PDQ): {aux_skipped} 行")

    # ---- 匹配出货 → 收集标记操作 ----
    log("")
    log("=" * 50)
    log("步骤3: 匹配出货数据")
    log("=" * 50)

    stats = {
        'processed': 0,
        'not_found': 0,
        'rows_marked': 0,
    }

    actions = []  # [(row_idx, date_str, needs_fill, contract), ...]
    inventory_errors = []

    # 找到接单日期列（A列=1）
    col_order_date = 1

    # 收集每个合同的首行接单日期（用于补填SLB行）
    contract_order_dates = {}
    for r in range(header_row + 1, ws.max_row + 1):
        rc = _normalize(ws.cell(row=r, column=col_contract).value)
        if rc and rc not in contract_order_dates:
            od = ws.cell(row=r, column=col_order_date).value
            if od is not None:
                contract_order_dates[rc] = od

    for (contract, item_no), ship_qty in shipments.items():
        simple_key = _to_simple_key(item_no)
        ship_date = mark_date or dates.get((contract, item_no), '')
        slb = _is_set_item(item_no)

        candidates = row_index.get((contract, simple_key), [])
        available = _collect_available(
            ws, candidates, col_qty, col_beizhu, col_contract,
            col_item, max_col, is_slb=slb)

        log(f"  合同={contract}, 出货货号={item_no}, 简货号={simple_key}, "
            f"出货={ship_qty}, 出货日期={ship_date}, "
            f"{'SLB ' if slb else ''}"
            f"候选行={len(candidates)}, 可用行={len(available)}")

        if candidates and not available:
            _collect_available(
                ws, candidates, col_qty, col_beizhu, col_contract,
                col_item, max_col, is_slb=slb, log_func=log)

        if not available:
            if candidates:
                # 有候选行但都是子行（合同列为空）→ 正常跳过，不计入错误
                stats['sub_skipped'] = stats.get('sub_skipped', 0) + 1
                log(f"    - 子行跳过: 合同={contract}, "
                    f"货号={item_no} (候选{len(candidates)}行均为子行)")
            else:
                stats['not_found'] += 1
                inventory_errors.append({
                    'contract': contract, 'item_no': item_no,
                    'simple_no': simple_key, 'ship_qty': ship_qty,
                    'available_qty': 0,
                    'reason': '接单表未找到该合同+货号',
                })
                log(f"    ✗ 未找到匹配: 合同={contract}, "
                    f"出货货号={item_no}, 简货号={simple_key}")
            continue

        # 标记第一个可用行
        row_idx, row_qty, needs_fill = available[0]
        actions.append((row_idx, ship_date, needs_fill, contract))
        stats['processed'] += 1
        fill_tag = " (需补填合同)" if needs_fill else ""
        log(f"    ✓ 行{row_idx}: 标记出货日期{fill_tag}")

    # 收集需要蓝色填充的行（标记行 + 其下方子行）
    blue_rows = set()
    for row_idx, date_str, needs_fill, contract in actions:
        if not date_str:
            continue
        blue_rows.add(row_idx)
        # 往下找子行（合同列为空、属于同合同组的行）
        for r in range(row_idx + 1, ws.max_row + 1):
            rc = _normalize(ws.cell(row=r, column=col_contract).value)
            if rc:
                break  # 遇到新合同号行，停止
            blue_rows.add(r)

    # 关闭只读工作簿
    wb.close()

    # ---- XML手术式写入 ----
    log("")
    log("=" * 50)
    log("步骤4: 写入出货日期到备注列")
    log("=" * 50)

    if output_path is None:
        dir_name = os.path.dirname(order_path)
        output_path = os.path.join(dir_name, "接单表_更新.xlsx")

    if not col_beizhu:
        log("  ✗ 未检测到备注列，无法写入出货日期")
        shutil.copy2(order_path, output_path)
    elif not actions:
        log("  没有需要标记的行")
        shutil.copy2(order_path, output_path)
    else:
        cell_edits = []
        for row_idx, date_str, needs_fill, contract in actions:
            if not date_str:
                log(f"  跳过行{row_idx}: 出货文件无日期")
                continue
            # 写备注日期（文本格式）
            cell_edits.append((row_idx, col_beizhu, date_str))
            stats['rows_marked'] += 1
            log(f"  行{row_idx}: 备注 ← '{date_str}'")
            # SLB行需补填合同号和接单日期
            if needs_fill:
                cell_edits.append((row_idx, col_contract, contract))
                od = contract_order_dates.get(contract)
                if od is not None:
                    cell_edits.append((row_idx, col_order_date, od))
                log(f"  行{row_idx}: 补填合同={contract}, 接单日期={od}")

        try:
            _surgical_xlsx_write(order_path, output_path,
                                 sheet_title, cell_edits, blue_rows)
            log(f"  ✓ XML手术式写入完成，格式100%保留")
        except Exception as e:
            log(f"  XML写入失败: {e}，回退到openpyxl保存")
            # 兜底：用openpyxl写（可能有格式损失）
            _fallback_openpyxl_write(
                order_path, output_path, cell_edits, log)

    log(f"保存成功: {output_path}")

    # ---- 库存错误报告 ----
    report_path = None
    if inventory_errors:
        try:
            report_dir = os.path.dirname(order_path)
            report_path = _generate_error_report(inventory_errors, report_dir)
            stats['error_report'] = report_path
            log("")
            log(f"库存错误报告: {report_path}")
            log(f"  共 {len(inventory_errors)} 条异常，请人工检查")
        except Exception as e:
            log(f"生成错误报告失败: {e}")

    log("")
    log("=" * 50)
    log("处理结果")
    log("=" * 50)
    log(f"  成功匹配: {stats['processed']} 组")
    log(f"  子行跳过: {stats.get('sub_skipped', 0)} 组")
    log(f"  未找到匹配: {stats['not_found']} 组")
    log(f"  标记出货行: {stats['rows_marked']} 行")
    if inventory_errors:
        log(f"  *** 异常: {len(inventory_errors)} 条 → 见错误报告")

    return output_path, stats


def _fallback_openpyxl_write(order_path, output_path, cell_edits, log):
    """兜底：用openpyxl写入（格式可能有损失）"""
    wb = load_workbook(order_path)
    ws = wb.active
    for row_num, col_num, value in cell_edits:
        ws.cell(row=row_num, column=col_num).value = value
    wb.save(output_path)
    wb.close()
    log("  (兜底) openpyxl保存完成")


# ============================================================
# 命令行入口
# ============================================================

if __name__ == '__main__':
    import sys

    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s %(message)s',
        datefmt='%H:%M:%S',
    )

    if len(sys.argv) < 3:
        print("用法: python shipment_processor.py <接单表.xlsx> <出货文件夹>")
        sys.exit(1)

    order_file = sys.argv[1]
    ship_folder = sys.argv[2]

    if not os.path.isfile(order_file):
        print(f"错误: 接单表不存在: {order_file}")
        sys.exit(1)

    if not os.path.isdir(ship_folder):
        print(f"错误: 出货文件夹不存在: {ship_folder}")
        sys.exit(1)

    out, info = process(order_file, ship_folder)
    if out:
        print(f"\n完成 → {out}")
    else:
        print(f"\n处理失败: {info}")
        sys.exit(1)
