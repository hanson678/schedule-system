# -*- coding: utf-8 -*-
"""排期Excel自动处理 v5 - SKU映射 + 缓存 + 模糊搜索 + 并发锁 + 进度条"""
import os, re, shutil, json, threading, time as _time, logging
from datetime import datetime, timedelta, date
from contextlib import contextmanager
import openpyxl
from openpyxl.utils import column_index_from_string
from base_path import get_app_dir

DESKTOP = os.path.join(os.environ.get('USERPROFILE', r'C:\Users\Administrator'), 'Desktop')
DATA_DIR = os.path.join(get_app_dir(), 'data')
HISTORY_FILE = os.path.join(DATA_DIR, 'history.json')
RETRY_FILE = os.path.join(DATA_DIR, 'pending_retries.json')
BATCH_DIR = os.path.join(DESKTOP, 'batch_temp')
CFG_FILE = os.path.join(DATA_DIR, 'config.json')

# COM 颜色常量 (RGB: R + G*256 + B*65536)
BLUE_COM = 15773696     # RGB(0, 176, 240) = 0xF0B000 → 浅蓝 FF00B0F0
RED_COM = 255           # RGB(255, 0, 0) → 红色字体
BLACK_COM = 0           # RGB(0, 0, 0) → 黑色字体
YELLOW_COM = 65535      # RGB(255, 255, 0) → 黄色填充
_HAS_DIGIT = re.compile(r'\d')  # 有效SKU必须含数字
GREEN_COM = 5296274     # RGB(146, 208, 80) → 浅绿色填充（备注列空白标记）
_DEFAULT_ZURU_PATH = r'Z:\各客排期\ZURU生产排期'  # ISM直查表默认路径
UNDO_DIR = os.path.join(BATCH_DIR, 'undo')
UNDO_HISTORY = os.path.join(DATA_DIR, 'undo_history.json')

def _col_cn(col):
    """列字母→中文字段名（用于修改详情展示）"""
    if col in 'IJKL': return 'PO数量'
    if col in 'MNOP': return '出货日期'
    if col in 'DEFG': return '客户PO'
    n = ord(col) - 64 if len(col) == 1 else 26 + (ord(col[1]) - 64)
    if 18 <= n <= 32: return '单价USD'
    return col

# =================== 全局缓存 ===================
_sku_map_cache = {}          # SKU→排期文件关键词映射
_sku_map_mtime = 0           # 总排期文件修改时间
_yellow_cache = {}           # {filepath: {'mtime': float, 'rows': [...]}}
_yellow_cache_time = 0       # 上次全量扫描时间

# =================== 文件操作锁 ===================
_file_locks = {}             # {filepath: threading.Lock()}
_file_locks_lock = threading.Lock()

# =================== 批量进度 ===================
_batch_progress = {'running': False, 'current': '', 'done': 0, 'total': 0, 'details': []}


def _get_file_lock(filepath):
    """获取文件级别的锁"""
    with _file_locks_lock:
        if filepath not in _file_locks:
            _file_locks[filepath] = threading.Lock()
        return _file_locks[filepath]


@contextmanager
def file_lock(filepath):
    """文件操作锁上下文管理器"""
    lock = _get_file_lock(filepath)
    acquired = lock.acquire(timeout=30)
    if not acquired:
        raise TimeoutError(f'文件 {os.path.basename(filepath)} 正在被系统处理，请稍后重试')
    try:
        yield
    finally:
        lock.release()


def _is_yellow_fill(cell):
    """检查单元格是否有黄色填充"""
    try:
        fill = cell.fill
        if not fill or fill.patternType is None or fill.patternType == 'none':
            return False
        fg = fill.fgColor
        if fg and fg.rgb and str(fg.rgb) not in ('00000000', '0'):
            rgb_str = str(fg.rgb).upper()
            if len(rgb_str) == 8:
                r, g, b = int(rgb_str[2:4], 16), int(rgb_str[4:6], 16), int(rgb_str[6:8], 16)
            elif len(rgb_str) == 6:
                r, g, b = int(rgb_str[0:2], 16), int(rgb_str[2:4], 16), int(rgb_str[4:6], 16)
            else:
                return False
            # 黄色：R高, G高, B低
            if r > 200 and g > 180 and b < 100:
                return True
    except:
        pass
    return False


def _sku_key(sku):
    return re.sub(r'[^0-9]', '', str(sku))[:5]


def _is_ma_sheet(sn):
    """判断是否为MA材料sheet（应跳过）。
    跳过：纯"MA"、"彩盒MA"、"布料MA"、"MA包装"等
    保留：带产品前缀的排期sheet如"游水MA彩盒"（去掉MA+材料词后还有内容）"""
    snu = sn.upper()
    if 'MA' not in snu:
        return False
    cleaned = sn.replace('MA', '').replace('ma', '').strip()
    if not cleaned:
        return True  # 纯"MA"名 → 材料sheet
    temp = cleaned
    for _kw in ('彩盒', '半成品', '包装', '包裝', '产品', '產品', '客版', '布料', '成品'):
        temp = temp.replace(_kw, '')
    if not temp.strip():
        return True  # 纯材料sheet（如"布料MA"、"MA包装"、"彩盒MA"）
    return False  # 有产品前缀（如"游水MA彩盒"→"游水"），保留


def _should_skip_file(fn):
    """判断是否应跳过此文件（总排期、样板、旧排期）"""
    return '总' in fn or '样板' in fn or '樣板' in fn or '旧' in fn


def _should_skip_sheet(sn):
    """判断是否应跳过此sheet（样板、MA材料、旧排期、导出、转版本等）。注意：不含'取消'检查，需调用方按需判断"""
    return '样板' in sn or '樣板' in sn or '旧' in sn or '导出' in sn or '導出' in sn or '转' in sn or '轉' in sn or _is_ma_sheet(sn)


# =================== 繁体→简体转换（表头检测用）===================
_TRAD_TO_SIMP = str.maketrans(
    '貨驗國備註辦單數產業務號額價條碼總內類據際計劃種組廠區',
    '货验国备注办单数产业务号额价条码总内类据际计划种组厂区'
)

def _t2s(text):
    """繁体转简体（仅常用字，用于表头检测）"""
    return text.translate(_TRAD_TO_SIMP)


def _filter_remark_for_sku(remark, sku):
    """按货号过滤备注：只保留与当前SKU相关的行 + 通用行。
    货号特定行后面的续行（无货号前缀）归属于前面那个货号。
    例如备注含 '1.77772GQ2:MA\nEAN版本包装\n2.77858:MA\nEAN版本包装'，
    SKU为77772GQ2-S001时只返回第1-2行。"""
    if not remark or not sku:
        return remark
    sku_base = _item_code(sku).upper()
    if not sku_base:
        return remark
    lines = remark.split('\n')
    result = []
    # 状态：None=通用区域, True=当前货号匹配区域, False=其他货号区域
    current_owner = None
    has_any_item_line = False  # 整篇备注是否含货号特定行

    for line in lines:
        stripped = line.strip()
        if not stripped:
            continue
        # 检测行是否以 "N.货号:" 或 "货号：" 开头（货号特定行）
        m = re.match(r'\d+\.\s*(\d+[A-Za-z]*\d*[A-Za-z]*)\s*[:：]', stripped)
        if not m:
            m = re.match(r'(\d+[A-Za-z]+\d*[A-Za-z]*)\s*[:：]', stripped)
        if m:
            has_any_item_line = True
            line_item = m.group(1).upper()
            if line_item == sku_base or sku_base.startswith(line_item):
                current_owner = True
                result.append(stripped)
            else:
                current_owner = False
            continue
        # 非货号特定行：根据归属状态决定
        if current_owner is True:
            result.append(stripped)  # 归属当前货号
        elif current_owner is False:
            pass  # 归属其他货号，跳过
        else:
            result.append(stripped)  # 通用区域（还没遇到任何货号行）

    # 如果整篇备注都没有货号特定行，说明全是通用备注，原样返回
    if not has_any_item_line:
        return remark
    return '\n'.join(result) if result else remark


def _item_code(s):
    """提取基础商品代码: '125160H-S001' → '125160H', '15760UQ1' → '15760UQ1',
    'MEC457-77772-S001' → 'MEC457-77772'"""
    if not s:
        return ''
    s = re.sub(r'[\s\n]+', '', str(s).strip())
    # 取第一段（'-'之前）的数字+字母部分
    base = s.split('-')[0]
    m = re.match(r'(\d+[A-Za-z]*\d*)', base)
    if m:
        return m.group(1).upper()
    # 字母前缀货号（如MEC457-77772-S001）：取到-SXXX之前的所有段
    if re.match(r'[A-Za-z]+\d+', base):
        parts = s.split('-')
        result_parts = [parts[0]]
        for p in parts[1:]:
            if re.match(r'S\d+', p, re.I):
                break  # -SXXX后缀不属于基础代码
            if re.match(r'P\d+', p, re.I):
                break  # -P1等后缀也不属于基础代码
            result_parts.append(p)
        return '-'.join(result_parts).upper()
    return ''


def _sku_spec(s):
    """提取完整SKU规格码（含-SXXX后缀）: '92105-S001' → '92105-S001', '125160H-S001' → '125160H-S001',
    'MEC457-77772-S001' → 'MEC457-77772-S001'（字母前缀代码原样返回）"""
    if not s:
        return ''
    s = re.sub(r'[\s\n]+', '', str(s).strip()).upper()
    # 标准: 数字开头，支持可选年份段(-2025等)，完整保留 -SXXX 后面的所有字母后缀（如 -DISPC、-PKC、-MTS、-FS）
    m = re.match(r'(\d+[A-Za-z]*\d*(?:-20\d{2})?(?:-S\d+(?:-[A-Za-z]+\d*)*)?)', s, re.IGNORECASE)
    if m:
        return m.group(1).upper()
    # 字母前缀货号（如MEC457-77772-S001）：原样返回完整代码
    if re.match(r'[A-Za-z]+\d+', s.split('-')[0]):
        return s
    return _item_code(s)


def _po_type(po_str):
    """分类PO号类型：4500/ZMPO/MPO/其他
    用于参考行按PO类型匹配，同类型PO的参考行优先"""
    s = str(po_str or '').strip().upper()
    if s.startswith('4500'):
        return '4500'
    elif s.startswith('ZMPO'):
        return 'ZMPO'
    elif s.startswith('MPO'):
        return 'MPO'
    else:
        return 'other'


def _po_boundary_match(needle, haystack):
    """PO号数字边界匹配：确保前后不是数字，避免子串误匹配。
    例如：PO '4500194' 不匹配 '4500194447'，但匹配 'PO:4500194/line20'"""
    if not needle or not haystack:
        return False
    if needle == haystack:
        return True
    return needle in haystack and bool(re.search(r'(?<!\d)' + re.escape(needle) + r'(?!\d)', haystack))


def _normalize_date(s):
    """将各种日期格式统一为YYYY-MM-DD，支持YYYY-MM-DD、DD-MM-YYYY、MM-DD-YYYY"""
    if not s:
        return ''
    s = str(s).strip().replace('/', '-')
    # 已经是YYYY-MM-DD
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    # DD-MM-YYYY 或 MM-DD-YYYY
    m = re.match(r'(\d{1,2})-(\d{1,2})-(\d{4})', s)
    if m:
        a, b, year = int(m.group(1)), int(m.group(2)), m.group(3)
        if a > 12:   # a一定是日，格式DD-MM-YYYY
            return f"{year}-{b:02d}-{a:02d}"
        elif b > 12:  # b一定是日，格式MM-DD-YYYY
            return f"{year}-{a:02d}-{b:02d}"
        else:         # 都<=12，默认MM-DD-YYYY（商业惯用）
            return f"{year}-{a:02d}-{b:02d}"
    return s


def _parse_date(s):
    """解析日期字符串为datetime，无法解析时返回None（不再默认返回now()）"""
    if isinstance(s, datetime):
        return s
    if isinstance(s, date):
        return datetime(s.year, s.month, s.day)
    if not s:
        return None
    ns = _normalize_date(str(s))
    try:
        return datetime.strptime(ns, '%Y-%m-%d')
    except:
        return None


def _date_serial(dt):
    """将datetime转为Excel序列号（整数天数），避免pywin32时区偏移问题
    pywin32写datetime时会自动做UTC转换（CST-8h），导致日期差一天。
    直接写序列号不经过时区转换，结果精确。"""
    if not dt or not hasattr(dt, 'year'):
        return None
    # Excel序列号: 天数从1899/12/30起算（兼容Lotus 1-2-3 bug）
    return (dt - datetime(1899, 12, 30)).days


def _calc_inspection(ship_dt, wb_name='', sheet_name=''):
    """根据出货日期计算验货日期
    Fuggler系列: 出货-2天; 其他: 出货-4天
    河源(HY): 周六/周日不能验货; 其他: 周日不能验货
    返回datetime或None"""
    if not ship_dt or not hasattr(ship_dt, 'year'):
        return None
    _fn_sn = (str(wb_name) + ' ' + str(sheet_name)).lower()
    is_fuggler = ('fuggler' in _fn_sn or '125160' in _fn_sn or '125169' in _fn_sn
                  or any(f'1{x}' in _fn_sn for x in
                         ['5746','5749','5751','5754','5755','5760',
                          '5714','5726','5729','5747','5752','5779','5792','5783','5785',
                          '5704','5706','5733','5743','5756','5758','5759','5789','5797',
                          '5710','5728','5731','5727','5732','5750','5757','5774','5780','5787','5788']))
    is_hy = any(k in _fn_sn for k in ('hy', '河源'))
    days_before = 2 if is_fuggler else 4
    insp_dt = ship_dt - timedelta(days=days_before)
    if is_hy:
        if insp_dt.weekday() == 5:  # Saturday → Friday
            insp_dt -= timedelta(days=1)
        elif insp_dt.weekday() == 6:  # Sunday → Monday
            insp_dt += timedelta(days=1)
    else:
        if insp_dt.weekday() == 6:  # Sunday → Monday
            insp_dt += timedelta(days=1)
    if insp_dt >= ship_dt:
        insp_dt = ship_dt - timedelta(days=1)
        if insp_dt.weekday() == 6:
            insp_dt -= timedelta(days=2)
        elif is_hy and insp_dt.weekday() == 5:
            insp_dt -= timedelta(days=1)
    return insp_dt


def _sv_com(ws, r, c, v, d=False):
    """通过COM设置单元格值"""
    cell = ws.Cells(r, c)
    if d and v:
        if isinstance(v, str) and v:
            try:
                dt = datetime.strptime(v.replace('/', '-'), '%Y-%m-%d')
                cell.Value = _date_serial(dt)
                return
            except:
                pass
        if isinstance(v, datetime):
            cell.Value = _date_serial(v)
            return
    if v is not None and v != '':
        cell.Value = v


def _col_num(letter):
    """列字母→数字: A=1, B=2, ..., Z=26, AA=27"""
    r = 0
    for c in letter.upper():
        r = r * 26 + (ord(c) - ord('A') + 1)
    return r


class ExcelHandler:
    # 角色配比字典（加载一次后缓存，重启自动重载）
    _ratio_map = None
    _ratio_map_lock = threading.Lock()
    # 配比文件夹路径（放置所有 *配比*.xlsx 文件）
    _RATIO_DIR = os.path.join(os.environ.get('USERPROFILE', r'C:\Users\Administrator'),
                              r'Desktop\配比表')

    def __init__(self, config):
        self.z_path = config.get('z_drive_path', _DEFAULT_ZURU_PATH)
        self._cn_name_cache = None  # 中文名直查表缓存，首次使用时加载

    @staticmethod
    def _normalize_fn(s):
        """文件名标准化：统一全角/半角括号、大小写、空白，用于模糊匹配"""
        s = s.lower().replace('\xa0', ' ')
        s = s.replace('（', '(').replace('）', ')').replace('＃', '#')
        return s

    def _remap_ism_path(self, p):
        """ISM直查表路径→实际z_path映射（测试版路径可能与默认Z盘不同）
        + 子目录不存在时回退到父目录（如FUGGLER河源排期→ZURU生产排期）
        + 原文件被WPS重命名为.xlsx0.xlsx时自动查找替代
        + 文件名全角/半角/大小写差异时模糊匹配"""
        if self.z_path.rstrip('\\') != _DEFAULT_ZURU_PATH.rstrip('\\') \
           and p.startswith(_DEFAULT_ZURU_PATH):
            p = self.z_path + p[len(_DEFAULT_ZURU_PATH):]
        if not os.path.exists(p):
            # 尝试子目录回退
            alt = os.path.join(os.path.dirname(os.path.dirname(p)), os.path.basename(p))
            if os.path.exists(alt):
                return alt
            # 查找.xlsx0.xlsx替代（WPS编辑时会把原文件重命名为.xlsx0.xlsx）
            _dir = os.path.dirname(p)
            _base = os.path.basename(p)
            if _base.endswith('.xlsx') and os.path.isdir(_dir):
                _prefix = _base[:-5]  # 去掉.xlsx后缀
                for f in os.listdir(_dir):
                    if f.startswith(_prefix) and re.search(r'\.xlsx\d+\.xlsx$', f) and not f.startswith('~$'):
                        return os.path.join(_dir, f)
            # 模糊匹配：全角/半角括号、大小写差异（如ZURU→zuru，（B车间）→(B车间)）
            if os.path.isdir(_dir):
                _norm = self._normalize_fn(_base)
                for f in os.listdir(_dir):
                    if self._normalize_fn(f) == _norm:
                        return os.path.join(_dir, f)
        return p

    # =================== 只读搜索 (openpyxl read_only) ===================

    _xlsx_list_cache = None
    _xlsx_list_path = None
    _xlsx_list_time = 0  # 上次刷新时间戳

    # auto_find 持久缓存：key=sku, value=result dict（跨请求复用）
    _auto_find_cache = {}
    _auto_find_cache_time = 0  # 上次清除时间

    # item_schedule_map.json 缓存
    _ism_cache = None
    _ism_prefix = None   # 前缀索引：数字前缀 → 第一个匹配的entries
    _ism_mtime = 0

    # 请求级workbook缓存：同一批处理中复用已打开的openpyxl workbook，避免重复load_workbook
    _wb_cache = {}           # {filepath: wb}
    _wb_cache_lock = threading.Lock()
    _wb_fail_cache = {}      # 负面缓存：打开失败的文件不再重试（避免被锁文件每次hang 60秒）

    # 增量内容缓存（跨请求持久化）：缓存每个文件的货号列表，避免重复load_workbook
    # 格式：{filepath: {'mtime': float, 'items': {sheet_name: set(item_upper)}, 'sheetnames': [...]}}
    _content_cache = {}
    _content_cache_lock = threading.Lock()

    @classmethod
    def _get_item_schedule_map(cls):
        """加载item_schedule_map.json直查表（1964个货号→文件/sheet映射）"""
        _path = os.path.join(get_app_dir(), 'data', 'item_schedule_map.json')
        if not os.path.exists(_path):
            return None
        try:
            mt = os.path.getmtime(_path)
            if cls._ism_cache is not None and mt == cls._ism_mtime:
                return cls._ism_cache
            with open(_path, 'r', encoding='utf-8') as f:
                cls._ism_cache = json.load(f)
            cls._ism_mtime = mt
            # 构建前缀索引：从"15731-S001"/"#25257"等key提取数字前缀
            # 合并同前缀所有entries（按path去重），避免遗漏不同变体对应的文件
            cls._ism_prefix = {}
            for key in cls._ism_cache:
                m = re.match(r'#?(\d{3,})', key)
                if m:
                    prefix = m.group(1)
                    if prefix not in cls._ism_prefix:
                        cls._ism_prefix[prefix] = list(cls._ism_cache[key])
                    else:
                        existing_paths = {e.get('path') for e in cls._ism_prefix[prefix]}
                        for e in cls._ism_cache[key]:
                            if e.get('path') not in existing_paths:
                                cls._ism_prefix[prefix].append(e)
                                existing_paths.add(e.get('path'))
            return cls._ism_cache
        except Exception:
            return cls._ism_cache

    def _list_xlsx(self):
        # 缓存文件列表，300秒内同一路径不重复遍历
        # 只扫描根目录 + FUGGLER河源排期子目录（与scan_schedules.py一致）
        import time as _t
        now = _t.time()
        if (ExcelHandler._xlsx_list_cache is not None
                and ExcelHandler._xlsx_list_path == self.z_path
                and now - ExcelHandler._xlsx_list_time < 300):
            return ExcelHandler._xlsx_list_cache
        files = []
        if not os.path.isdir(self.z_path):
            return files
        for root, dirs, fnames in os.walk(self.z_path):
            # 与scan_schedules.py保持一致：根目录只进入FUGGLER河源排期
            if root == self.z_path:
                dirs[:] = [d for d in dirs if d == 'FUGGLER河源排期']
            else:
                dirs[:] = []  # 不再深入子目录
            for item in fnames:
                if item.endswith('.xlsx') and not item.startswith('~$'):
                    files.append(os.path.join(root, item))
        ExcelHandler._xlsx_list_cache = files
        ExcelHandler._xlsx_list_path = self.z_path
        ExcelHandler._xlsx_list_time = now
        return files

    @classmethod
    def _ism_lookup(cls, ism, sku, spec=None):
        """ISM直查表共享查找逻辑：按spec/sku_upper/item/item_digits优先级匹配。
        返回匹配的entries列表，未命中返回None。"""
        if spec is None:
            spec = _sku_spec(sku)
        item = _item_code(sku)
        sku_upper = re.sub(r'[^A-Za-z0-9]', '', str(sku)).upper()
        _m = re.match(r'\d+', item) if item else None
        item_digits = _m.group() if _m else ''
        for key in [spec.upper() if spec else '', sku_upper,
                    item.upper() if item else '', item_digits]:
            if key and key in ism:
                return ism[key]
        if cls._ism_prefix and item_digits and len(item_digits) >= 3:
            return cls._ism_prefix.get(item_digits)
        return None

    @classmethod
    def check_sku_locked_file(cls, sku, fail_cache=None):
        """检查SKU对应的ISM文件是否被锁定。
        fail_cache: 可选，外部传入的_wb_fail_cache快照（缓存已清空后使用）。
        返回 (is_locked, file_name) 或 (False, None)"""
        ism = cls._get_item_schedule_map()
        if not ism:
            return False, None
        entries = cls._ism_lookup(ism, sku)
        if not entries:
            return False, None  # ISM没有此货号，不是"被占用"
        _cache = fail_cache if fail_cache is not None else cls._wb_fail_cache
        for entry in entries:
            fn = entry.get('file', '')
            fp = entry.get('path', '')
            if fp and fp in _cache:
                return True, fn
        return False, None

    @classmethod
    def get_ism_candidate_files(cls, lines):
        """通过item_schedule_map.json直查表，获取候选排期文件→sheet映射。
        返回 {file_path: set(sheet_names)}，用于缩小PO搜索范围。
        lines: [{'sku_spec': ..., 'item_code': ..., 'sku': ...}, ...]"""
        ism = cls._get_item_schedule_map()
        if not ism:
            return {}
        candidates = {}  # {file_path: set(sheet_names)}
        _path_fix_cache = {}  # ISM路径修正缓存
        for ln in lines:
            sku = ln.get('item_code') or ln.get('sku', '')
            spec = _sku_spec(ln.get('sku_spec', '') or sku)
            entries = cls._ism_lookup(ism, sku, spec)
            if entries:
                for entry in entries:
                    p = entry.get('path', '')
                    s = entry.get('sheet', '')
                    if not p or _should_skip_file(os.path.basename(p)):
                        continue
                    if s and _should_skip_sheet(s):
                        continue
                    # ISM路径修正
                    if p in _path_fix_cache:
                        p = _path_fix_cache[p]
                    else:
                        _orig_p = p
                        fname = os.path.basename(p)
                        # WPS .xlsx0.xlsx → 正常.xlsx映射
                        if re.search(r'\.xlsx\d+\.xlsx$', fname):
                            normal_name = re.sub(r'\.xlsx\d+\.xlsx$', '.xlsx', fname)
                            normal_path = os.path.join(os.path.dirname(p), normal_name)
                            if os.path.exists(normal_path):
                                p = normal_path
                            else:
                                # 尝试父目录
                                parent = os.path.dirname(os.path.dirname(p))
                                alt = os.path.join(parent, normal_name)
                                if os.path.exists(alt):
                                    p = alt
                        if not os.path.exists(p):
                            parent = os.path.dirname(os.path.dirname(_orig_p))
                            alt = os.path.join(parent, os.path.basename(_orig_p))
                            if os.path.exists(alt):
                                p = alt
                        _path_fix_cache[_orig_p] = p
                    candidates.setdefault(p, set())
                    if s:
                        candidates[p].add(s)
        return candidates

    @classmethod
    def clear_cache(cls):
        """清除文件列表缓存和ISM缓存（路径切换或刷新时调用）"""
        cls._xlsx_list_cache = None
        cls._xlsx_list_path = None
        cls._xlsx_list_time = 0
        cls._ism_cache = None
        cls._ism_prefix = None
        cls._ism_mtime = 0

    @classmethod
    def clear_auto_find_cache(cls):
        """清除auto_find缓存（排期文件变更后调用）"""
        cls._auto_find_cache = {}
        cls._auto_find_cache_time = 0

    @classmethod
    def _get_cached_wb(cls, fp):
        """从缓存获取或新建openpyxl workbook（read_only模式）。
        batch处理期间同一文件只打开一次，节省4-10秒/文件的load_workbook开销。
        注意：~$锁文件和.xlsx0.xlsx文件openpyxl read_only模式仍可读取数据，不跳过。"""
        with cls._wb_cache_lock:
            if fp in cls._wb_cache:
                return cls._wb_cache[fp]
            if fp in cls._wb_fail_cache:
                raise OSError(f"文件已知无法打开(负面缓存): {os.path.basename(fp)}")
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        except Exception as e:
            with cls._wb_cache_lock:
                cls._wb_fail_cache[fp] = True
            raise OSError(f"无法打开文件: {os.path.basename(fp)}: {e}")
        with cls._wb_cache_lock:
            if fp in cls._wb_cache:
                try:
                    wb.close()
                except Exception:
                    pass
                return cls._wb_cache[fp]
            cls._wb_cache[fp] = wb
        return wb

    @classmethod
    def _clear_wb_cache(cls):
        """关闭并清空workbook缓存（每次batch处理结束后必须调用）"""
        with cls._wb_cache_lock:
            for fp, wb in cls._wb_cache.items():
                try:
                    # Windows上read_only模式需要显式关闭底层archive释放文件句柄
                    if hasattr(wb, '_archive') and wb._archive:
                        try:
                            wb._archive.close()
                        except Exception:
                            pass
                    wb.close()
                except Exception:
                    pass
            cls._wb_cache.clear()
            cls._wb_fail_cache.clear()
        import gc
        gc.collect()

    @classmethod
    def _get_content_cache(cls, fp):
        """获取文件的增量内容缓存（货号列表）。命中返回dict，未命中返回None。"""
        try:
            mtime = os.path.getmtime(fp)
        except OSError:
            return None
        with cls._content_cache_lock:
            cached = cls._content_cache.get(fp)
            if cached and cached['mtime'] == mtime:
                return cached
        return None

    @classmethod
    def _set_content_cache(cls, fp, sheetnames, items_by_sheet):
        """写入文件的增量内容缓存。items_by_sheet: {sheet_name: set(item_upper)}"""
        try:
            mtime = os.path.getmtime(fp)
        except OSError:
            return
        with cls._content_cache_lock:
            cls._content_cache[fp] = {
                'mtime': mtime,
                'sheetnames': list(sheetnames),
                'items': items_by_sheet,
            }

    @classmethod
    def _content_cache_has_sku(cls, fp, sku_upper, item_upper, spec_upper):
        """快速检查content cache中某文件是否含某SKU。返回(found, sheet_name)或(False, None)。"""
        cached = cls._get_content_cache(fp)
        if not cached:
            return None, None  # None=缓存未命中，需打开文件
        for sn, items in cached['items'].items():
            for check in [spec_upper, sku_upper, item_upper]:
                if check and check in items:
                    return True, sn
            # 前缀匹配：item_digits匹配（如15780匹配15780D）
            if item_upper:
                _digits = re.match(r'\d+', item_upper)
                if _digits:
                    d = _digits.group()
                    for it in items:
                        it_digits = re.match(r'\d+', it)
                        if it_digits and it_digits.group() == d:
                            return True, sn
        return False, None  # False=缓存确认不在此文件

    # =================== SKU→排期映射 ===================

    def _get_sku_mapping(self):
        """获取SKU→排期映射（优先JSON文件，带缓存）"""
        global _sku_map_cache, _sku_map_mtime
        # 优先从 data/sku_mapping.json 读取
        json_path = os.path.join(get_app_dir(), 'data', 'sku_mapping.json')
        if os.path.exists(json_path):
            try:
                mtime = os.path.getmtime(json_path)
                if _sku_map_cache and mtime == _sku_map_mtime:
                    return _sku_map_cache
                _sku_map_cache = self._load_sku_mapping_json(json_path)
                _sku_map_mtime = mtime
                logging.info(f"[SKU映射] 从JSON加载 {len(_sku_map_cache)} 个映射项")
                return _sku_map_cache
            except Exception as e:
                logging.warning(f"[SKU映射] JSON读取失败: {e}")
        # 回退：从总排期Excel读取
        master = self.find_master_schedule()
        if not master:
            return _sku_map_cache or {}
        try:
            mtime = os.path.getmtime(master)
        except:
            return _sku_map_cache or {}
        if _sku_map_cache and mtime == _sku_map_mtime:
            return _sku_map_cache
        _sku_map_cache = self._load_sku_mapping_excel(master)
        _sku_map_mtime = mtime
        logging.info(f"[SKU映射] 从Excel加载 {len(_sku_map_cache)} 个映射项")
        return _sku_map_cache

    def _load_sku_mapping_json(self, json_path):
        """从 data/sku_mapping.json 加载映射"""
        import json as _json
        with open(json_path, 'r', encoding='utf-8') as f:
            data = _json.load(f)
        return data.get('mapping', {})

    def _get_sheet_mapping(self):
        """获取货号→工作簿名称映射（从sku_mapping.json的sheet_mapping段）"""
        json_path = os.path.join(get_app_dir(), 'data', 'sku_mapping.json')
        if not os.path.exists(json_path):
            return {}
        try:
            import json as _json
            with open(json_path, 'r', encoding='utf-8') as f:
                data = _json.load(f)
            return data.get('sheet_mapping', {})
        except:
            return {}

    def _load_sku_mapping_excel(self, master_fp):
        """从总排期"对应排期-货号"Sheet加载 SKU→排期文件关键词 映射"""
        try:
            wb = openpyxl.load_workbook(master_fp, read_only=True, data_only=True)
        except:
            return {}
        target = None
        for sn in wb.sheetnames:
            if '对应' in sn and '货号' in sn:
                target = sn
                break
        if not target:
            wb.close()
            return {}
        ws = wb[target]
        mapping = {}
        current_keywords = []

        for row in ws.iter_rows(min_row=2, max_col=20):
            first_val = str(row[0].value or '').strip()
            if re.search(r'[Zz][Uu][Rr][Uu]', first_val) and re.search(r'20\d{2}', first_val):
                current_keywords = re.findall(r'(\d{4,5})', first_val)
                continue
            if not current_keywords:
                continue
            for cell in row:
                val = str(cell.value or '').strip()
                if not val:
                    continue
                val = re.sub(r'.*?明[细細][:：]\s*', '', val)
                for token in re.split(r'[\s,;，；]+', val):
                    token = token.strip()
                    if not token:
                        continue
                    if re.match(r'^\d{4,6}$', token):
                        mapping[token] = current_keywords
                    elif re.match(r'^[A-Za-z]+\d+$', token):
                        mapping[token.upper()] = current_keywords
                    elif re.match(r'^\d+[A-Za-z]+\d*$', token):
                        mapping[token.upper()] = current_keywords
        wb.close()
        return mapping

    def get_sku_mapping_info(self):
        """返回SKU映射信息（供API调用）"""
        mapping = self._get_sku_mapping()
        # 按排期文件分组
        grouped = {}
        for sku, keywords in mapping.items():
            key = ','.join(keywords)
            if key not in grouped:
                grouped[key] = {'keywords': keywords, 'skus': []}
            grouped[key]['skus'].append(sku)
        return {
            'total': len(mapping),
            'groups': len(grouped),
            'detail': [{'keywords': g['keywords'], 'skus': sorted(g['skus']),
                         'count': len(g['skus'])} for g in grouped.values()]
        }

    def auto_find(self, sku, current_po=''):
        """自动查找SKU对应的排期文件和工作表（优先使用映射表）
        current_po: 当前PO号，用于参考行按PO类型匹配"""
        # 有效SKU必须含数字（排除纯中文如"更新包装备注"）
        if not sku or not _HAS_DIGIT.search(str(sku)):
            logging.info(f"[auto_find] SKU '{sku}' 无数字，跳过")
            return None
        # 持久缓存：同SKU结果300秒内复用（排期文件短期内不会变）
        import time as _t
        _now = _t.time()
        # 每60秒清理过期条目，防止_auto_find_cache无限增长
        if _now - ExcelHandler._auto_find_cache_time > 60:
            ExcelHandler._auto_find_cache_time = _now
            _expired = [k for k, v in ExcelHandler._auto_find_cache.items()
                        if _now - v.get('_ts', 0) > 300]
            for k in _expired:
                del ExcelHandler._auto_find_cache[k]
        _cache_key = f"{sku}|{_po_type(current_po) if current_po else ''}"
        if _cache_key in ExcelHandler._auto_find_cache:
            _cached = ExcelHandler._auto_find_cache[_cache_key]
            if _now - _cached.get('_ts', 0) < 300:
                logging.debug(f"[auto_find] 缓存命中: {sku}")
                return {k: v for k, v in _cached.items() if k != '_ts'}

        num = _sku_key(sku)
        sku_upper = re.sub(r'[^A-Za-z0-9]', '', str(sku)).upper()
        item = _item_code(sku)  # 基础货号，如 '125160H'
        spec = _sku_spec(sku)   # 完整规格码，如 '125160H-S001'（区分变体）
        # 提取item的前导数字部分作为备选文件名匹配（解决9548-S001→95480不匹配9548文件名）
        _m = re.match(r'\d+', item) if item else None
        item_digits = _m.group() if _m else ''

        # 检查SKU是否含年份版本（如"9298-2025-S001-NB"中的"2025"）
        sku_year = ''
        for part in str(sku).split('-'):
            if re.match(r'^20\d{2}$', part):
                sku_year = part
                break

        # ===== 第0步：通过item_schedule_map.json直查表快速定位 =====
        _ism = self._get_item_schedule_map()
        _ism_found_files = False  # ISM是否找到候选文件（含前缀匹配）
        _ism_all_locked = True    # ISM候选文件是否全部被锁
        _ism_locked_file = ''     # 被锁文件名（用于错误提示）
        if _ism:
            _entries_to_try = self._ism_lookup(_ism, sku, spec)
            if _entries_to_try:
                _ism_found_files = True
                # SKU含年份时(如92104-2025-S001/9298-2025-S001)，
                # 珠片公仔系列应去"无手脚毛绒"文件而非"rainbocorn"文件
                if sku_year:
                    _has_wushj = [e for e in _entries_to_try if '无手脚' in e.get('file', '')]
                    _no_wushj = [e for e in _entries_to_try if '无手脚' not in e.get('file', '')]
                    if _has_wushj:
                        _entries_to_try = _has_wushj + _no_wushj
                        logging.info(f"[auto_find] ISM年份感知: '{sku}' 含'{sku_year}', 优先无手脚毛绒文件")
                    elif _no_wushj and any('rainbocorn' in e.get('file', '').lower() for e in _no_wushj):
                        # ISM只有rainbocorn文件但SKU含年份→跳过ISM，让mapping步骤重新路由
                        logging.info(f"[auto_find] ISM年份感知: '{sku}' 含'{sku_year}', ISM仅有rainbocorn→跳过ISM走mapping")
                        _entries_to_try = []
                        _ism_found_files = False
                if _entries_to_try:
                    _ism_found_files = True
                for _entry in _entries_to_try:
                    fp = self._remap_ism_path(_entry.get('path', ''))
                    fn = _entry.get('file', '')
                    ts = _entry.get('sheet', '')
                    if fp:
                        result = self._search_sku_in_file(fp, fn, num, sku_upper, item,
                                                          target_sheet=ts, sku_spec=spec,
                                                          current_po=current_po)
                        if result:
                            _ism_all_locked = False
                            result['_ts'] = _t.time()
                            ExcelHandler._auto_find_cache[_cache_key] = result
                            logging.info(f"[auto_find] 直查表命中: {sku} → {fn}/{ts}")
                            return {k: v for k, v in result.items() if k != '_ts'}
                        # 检查是否因为文件被锁而失败（非真正的"未找到"）
                        if fp in ExcelHandler._wb_fail_cache:
                            _ism_locked_file = fn
                        else:
                            _ism_all_locked = False  # 至少有一个文件可以打开
        # ISM找到候选文件但全部被锁 → 直接返回，不浪费时间做全文件搜索
        if _ism_found_files and _ism_all_locked:
            logging.info(f"[auto_find] SKU '{sku}' ISM定位到文件但被占用({_ism_locked_file})，跳过全文件搜索")
            return None

        # ===== 第1步：通过SKU映射表查找 =====
        mapping = self._get_sku_mapping()
        sheet_map = self._get_sheet_mapping()
        file_keywords = None
        target_sheet = None  # 从sheet_mapping获取的目标工作簿
        lookup_keys = [sku_upper, num, item.upper() if item else '', item_digits]
        if mapping:
            for key in lookup_keys:
                if key and mapping.get(key):
                    file_keywords = mapping[key]
                    break
        # 查找目标工作簿名称
        if sheet_map:
            for key in lookup_keys:
                if key and key in sheet_map and not key.startswith('_'):
                    target_sheet = sheet_map[key]
                    break

        if file_keywords:
            # 有target_sheet时，优先搜索含"排期"的文件（生产排期），再搜其他
            all_matched = []
            for fp in self._list_xlsx():
                fn = os.path.basename(fp)
                if _should_skip_file(fn):
                    continue
                # 有target_sheet时，跳过旧排期目录（优先当年排期）
                if target_sheet and ('旧排期' in fp or '旧排期' in fn):
                    continue
                if not any(kw in fn for kw in file_keywords):
                    continue
                all_matched.append((fp, fn))
            # SKU含年份(如92104-2025-S001/9298-2025-S001)时，优先匹配文件名含"无手脚毛绒"的文件
            # 珠片公仔系列：rainbocorn=有手脚毛绒(旧版), 无手脚毛绒=2025新版
            if target_sheet and sku_year and len(all_matched) > 1:
                _year_preferred = [(fp, fn) for fp, fn in all_matched
                                   if '无手脚' in fn or (num and num in fn)]
                if _year_preferred:
                    all_matched = _year_preferred
                    logging.info(f"[auto_find] 年份感知: SKU含'{sku_year}', 缩窄到 {[fn for _,fn in all_matched]}")
            if target_sheet:
                all_matched.sort(key=lambda x: (0 if '排期' in x[1] else 1, x[1]))
            for fp, fn in all_matched:
                result = self._search_sku_in_file(fp, fn, num, sku_upper, item,
                                                  target_sheet=target_sheet, sku_spec=spec,
                                                  current_po=current_po)
                if result:
                    return result

        # ===== 第2步：原有逻辑（按文件名包含数字匹配）=====
        # 构建候选匹配词：num + item纯数字（去重）
        candidates = set()
        if num and len(num) >= 4:
            candidates.add(num)
        if item_digits and len(item_digits) >= 4:
            candidates.add(item_digits)

        if not candidates:
            # 无候选词时直接进入兜底搜索
            pass
        else:
            best = None
            for fp in self._list_xlsx():
                fn = os.path.basename(fp)
                if _should_skip_file(fn):
                    continue
                if not any(c in fn for c in candidates):
                    continue
                # 含年份版本时，优先匹配同时含年份的文件
                if sku_year and sku_year not in fn:
                    continue
                result = self._search_sku_in_file(fp, fn, num, sku_upper, item, sku_spec=spec,
                                                  current_po=current_po)
                if result:
                    if not best or result['cnt'] > best['cnt']:
                        best = result
            # 若年份过滤太严格未匹配到，放宽年份限制重试
            if not best and sku_year:
                for fp in self._list_xlsx():
                    fn = os.path.basename(fp)
                    if _should_skip_file(fn):
                        continue
                    if not any(c in fn for c in candidates):
                        continue
                    result = self._search_sku_in_file(fp, fn, num, sku_upper, item, sku_spec=spec,
                                                  current_po=current_po)
                    if result:
                        if not best or result['cnt'] > best['cnt']:
                            best = result
            if not best and not sku_year:
                # 无年份也无匹配，正常逻辑
                for fp in self._list_xlsx():
                    fn = os.path.basename(fp)
                    if _should_skip_file(fn):
                        continue
                    if not any(c in fn for c in candidates):
                        continue
                    result = self._search_sku_in_file(fp, fn, num, sku_upper, item, sku_spec=spec,
                                                  current_po=current_po)
                    if result:
                        if not best or result['cnt'] > best['cnt']:
                            best = result
            if best:
                best['_ts'] = _t.time()
                ExcelHandler._auto_find_cache[_cache_key] = best
                return {k: v for k, v in best.items() if k != '_ts'}

        # ===== 第3步：兜底搜索（前两步均未匹配，搜索所有排期文件内容）=====
        # ISM已定位候选文件但打开失败 → 不做全文件搜索（该SKU的文件已知，只是被锁）
        if _ism_found_files:
            logging.info(f"[auto_find] SKU '{sku}' ISM有候选但未匹配到（文件可能被锁或内容不符），跳过全文件搜索")
            return None
        logging.info(f"[auto_find] SKU '{sku}' 前两步未匹配，启动全文件搜索")
        best = None
        _full_scan_deadline = _t.time() + 20  # 全文件搜索最多20秒
        _spec_upper = spec.upper() if spec else ''
        _item_upper = item.upper() if item else ''
        _cache_skip = 0
        for fp in self._list_xlsx():
            if _t.time() > _full_scan_deadline:
                logging.warning(f"[auto_find] SKU '{sku}' 全文件搜索超时(>20s)，停止搜索")
                break
            fn = os.path.basename(fp)
            if _should_skip_file(fn):
                continue
            # 增量缓存快速排除：如果缓存确认该文件不含此SKU，跳过
            _cc_found, _cc_sheet = ExcelHandler._content_cache_has_sku(
                fp, sku_upper, _item_upper, _spec_upper)
            if _cc_found is False:
                _cache_skip += 1
                continue
            result = self._search_sku_in_file(fp, fn, num, sku_upper, item, sku_spec=spec,
                                                  current_po=current_po)
            if result:
                if not best or result['cnt'] > best['cnt']:
                    best = result
        if _cache_skip:
            logging.debug(f"[auto_find] 全文件搜索: content cache跳过{_cache_skip}个文件")
        if not best:
            logging.warning(f"[auto_find] SKU '{sku}' (num={num}, item={item}, spec={spec}) 未匹配到任何排期文件")
        if best:
            best['_ts'] = _t.time()
            ExcelHandler._auto_find_cache[_cache_key] = best
            return {k: v for k, v in best.items() if k != '_ts'}
        return best

    def _search_sku_in_file(self, fp, fn, num, sku_upper, item_code='', target_sheet=None, sku_spec='',
                            current_po=''):
        """在指定文件中搜索SKU，返回匹配结果
        只在ITEM#列（G列=index6）搜索货号，匹配优先级：
        1. SKU-SPEC精确匹配（如92105-S001），区分同货号不同变体
        2. 基础货号精确匹配（如92105）
        3. 前缀匹配（如9548匹配9548G4，长度差≤3）
        target_sheet: 从sheet_mapping指定的目标工作簿名称，有则只搜该sheet
        sku_spec: 完整规格码（含-SXXX后缀），用于精确区分变体
        current_po: 当前PO号，用于参考行按PO类型(4500/ZMPO/MPO/其他)优先匹配"""
        try:
            wb = ExcelHandler._get_cached_wb(fp)
        except Exception as e:
            # 记入负面缓存，同一batch内不再重试此文件（避免被锁文件每次hang 60秒）
            with ExcelHandler._wb_cache_lock:
                ExcelHandler._wb_fail_cache[fp] = True
            logging.warning(f"[auto_find] 无法打开 {fn}: {e}")
            return None
        best = None
        _items_by_sheet = {}  # 收集货号用于content cache：{sheet: set(item_upper)}
        # 当有target_sheet时，优先精确匹配sheet名称
        sheets_to_search = wb.sheetnames
        matched_target_sheets = set()
        if target_sheet:
            # 1.精确匹配
            matched = [sn for sn in wb.sheetnames if sn == target_sheet]
            if not matched:
                # 2.双向包含匹配（如"15701明细"匹配"15701"，或"恐龙"匹配"恐龙明细"）
                matched = [sn for sn in wb.sheetnames if target_sheet in sn or sn in target_sheet]
            if not matched:
                # 3.前导数字匹配（如target_sheet="9548明细"匹配sheet名含"9548"的）
                ts_digits = re.match(r'\d+', target_sheet)
                if ts_digits:
                    ts_num = ts_digits.group()
                    matched = [sn for sn in wb.sheetnames if ts_num in sn and '取消' not in sn and not _should_skip_sheet(sn)]
            if matched:
                sheets_to_search = matched
                matched_target_sheets = set(matched)
                logging.info(f"[auto_find] 使用sheet_mapping定位: {fn} → {matched[0]}")
            else:
                # target_sheet指定但此文件中找不到 → 跳过此文件，不搜索其他sheet
                logging.info(f"[auto_find] {fn} 中未找到目标sheet '{target_sheet}'，跳过此文件")
                return None

        for sn in sheets_to_search:
            if '取消' in sn or '对应' in sn or '总' in sn:
                continue
            if _should_skip_sheet(sn):
                continue
            try:
                ws = wb[sn]
                # 三级匹配：spec精确 > base精确 > 前缀
                # 每级同时跟踪：同PO类型(spo) vs 任意PO类型
                cur_pt = _po_type(current_po) if current_po else ''
                # 参考行选择：同级别中取非空数据最多的行
                # 每个ref存(row_num, data_count)，新匹配只在data更多时替换
                ref_spec_named = (0, 0)    # SKU-SPEC精确匹配（有产品名）
                ref_spec_any = (0, 0)      # SKU-SPEC精确匹配（无产品名）
                ref_spec_named_spo = (0, 0)  # 同PO类型+spec+named
                ref_spec_any_spo = (0, 0)    # 同PO类型+spec
                cnt_spec = 0
                ref_exact_named = (0, 0)   # 基础货号精确匹配（有产品名）
                ref_exact_any = (0, 0)     # 基础货号精确匹配（无产品名）
                ref_exact_named_spo = (0, 0)
                ref_exact_any_spo = (0, 0)
                cnt_exact = 0
                ref_prefix_named = (0, 0)
                ref_prefix_any = (0, 0)
                ref_prefix_named_spo = (0, 0)
                ref_prefix_any_spo = (0, 0)
                cnt_prefix = 0
                last_data_row = None
                def _rdc(r):
                    """计算行的非空单元格数（Row Data Count），用于选最佳参考行"""
                    return sum(1 for ci in range(min(10, len(r))) if r[ci].value)
                def _better(old_tup, dc):
                    """新行数据更多时替换，相同数据量取后面的行"""
                    return dc >= old_tup[1]
                for row in ws.iter_rows(min_row=2, max_col=10):
                    row_num = getattr(row[0], 'row', None)
                    if row_num is None:
                        for c in row[1:]:
                            row_num = getattr(c, 'row', None)
                            if row_num is not None:
                                break
                    if row_num is None:
                        continue
                    has_any_data = any(ci < len(row) and row[ci].value for ci in range(min(8, len(row))))
                    if has_any_data:
                        last_data_row = row_num
                    # 只在ITEM#列（G=6）和备选列（F=5, H=7）搜索货号
                    # 不搜索D(3)、E(4)列（PO号/客户PO，会产生子串误匹配）
                    for ci in [6, 5, 7]:
                        if ci >= len(row) or not row[ci].value:
                            continue
                        cv = str(row[ci].value).strip()
                        cv_item = _item_code(cv)
                        if not cv_item:
                            continue
                        # 收集货号用于content cache
                        if sn not in _items_by_sheet:
                            _items_by_sheet[sn] = set()
                        _items_by_sheet[sn].add(cv_item.upper())
                        has_name = len(row) > 7 and row[7].value and str(row[7].value).strip()
                        cv_spec = _sku_spec(cv)
                        dc = _rdc(row)  # 本行数据量
                        # 检测此行的PO类型（D列=index 3）
                        row_po_val = ''
                        if cur_pt:
                            for pi in [3, 4]:  # D列、E列查找PO号
                                if pi < len(row) and row[pi].value:
                                    row_po_val = str(row[pi].value).strip()
                                    if row_po_val:
                                        break
                        same_po = bool(cur_pt and _po_type(row_po_val) == cur_pt)
                        # 第1级：SKU-SPEC精确匹配（如92105-S001精确匹配92105-S001）
                        if sku_spec and cv_spec == sku_spec:
                            if _better(ref_spec_any, dc):
                                ref_spec_any = (row_num, dc)
                            cnt_spec += 1
                            if has_name and _better(ref_spec_named, dc):
                                ref_spec_named = (row_num, dc)
                            if same_po:
                                if _better(ref_spec_any_spo, dc):
                                    ref_spec_any_spo = (row_num, dc)
                                if has_name and _better(ref_spec_named_spo, dc):
                                    ref_spec_named_spo = (row_num, dc)
                            break
                        # 第2级：基础货号精确匹配（如92105精确匹配92105）
                        elif item_code and cv_item == item_code:
                            if _better(ref_exact_any, dc):
                                ref_exact_any = (row_num, dc)
                            cnt_exact += 1
                            if has_name and _better(ref_exact_named, dc):
                                ref_exact_named = (row_num, dc)
                            if same_po:
                                if _better(ref_exact_any_spo, dc):
                                    ref_exact_any_spo = (row_num, dc)
                                if has_name and _better(ref_exact_named_spo, dc):
                                    ref_exact_named_spo = (row_num, dc)
                            break
                        # 第3级：前缀匹配（如"9548"匹配"9548G4"，长度差≤3防止误匹配）
                        elif item_code and cv_item and \
                             abs(len(cv_item) - len(item_code)) <= 3 and \
                             (cv_item.startswith(item_code) or item_code.startswith(cv_item)):
                            if _better(ref_prefix_any, dc):
                                ref_prefix_any = (row_num, dc)
                            cnt_prefix += 1
                            if has_name and _better(ref_prefix_named, dc):
                                ref_prefix_named = (row_num, dc)
                            if same_po:
                                if _better(ref_prefix_any_spo, dc):
                                    ref_prefix_any_spo = (row_num, dc)
                                if has_name and _better(ref_prefix_named_spo, dc):
                                    ref_prefix_named_spo = (row_num, dc)
                            break
                # 优先级：spec精确 > base精确 > 前缀（从tuple中提取行号[0]，0表示无匹配）
                def _pick(*candidates):
                    """从多个(row_num, dc) tuple中取第一个非零的"""
                    for t in candidates:
                        if t[0]:
                            return t
                    return (0, 0)
                ref_t = _pick(ref_spec_named, ref_spec_any, ref_exact_named, ref_exact_any, ref_prefix_named, ref_prefix_any)
                ref_spo_t = _pick(ref_spec_named_spo, ref_spec_any_spo,
                           ref_exact_named_spo, ref_exact_any_spo,
                           ref_prefix_named_spo, ref_prefix_any_spo)
                cnt = cnt_spec or cnt_exact or cnt_prefix
                # 如果有同PO类型匹配，优先使用；否则用任意匹配
                po_matched = False
                if ref_spo_t[0]:
                    ref_t = ref_spo_t
                    po_matched = True
                    logging.info(f"[auto_find] {fn}/{sn} 同PO类型({cur_pt})匹配: ref={ref_spo_t[0]}(dc={ref_spo_t[1]})")
                elif ref_t[0] and cur_pt:
                    logging.info(f"[auto_find] {fn}/{sn} 无同PO类型({cur_pt})参考行，使用任意匹配: ref={ref_t[0]}(dc={ref_t[1]})")
                if ref_t[0] and sku_spec and ref_spec_named[0]:
                    logging.info(f"[auto_find] {fn}/{sn} SKU-SPEC精确匹配: spec={sku_spec}, ref={ref_t[0]}(dc={ref_t[1]})")
                elif ref_t[0] and not ref_spec_named[0] and not ref_spec_any[0] and item_code:
                    logging.info(f"[auto_find] {fn}/{sn} 仅基础货号匹配: item={item_code}(spec={sku_spec}), ref={ref_t[0]}(dc={ref_t[1]})")
                ref = ref_t[0]  # 提取行号给下游使用
                if ref:
                    result = {'file': fp, 'fname': fn, 'sheet': sn, 'ref': ref,
                              'cnt': cnt, 'mcol': max(ws.max_column or 30, 50),  # 至少50列
                              'po_type_match': po_matched}
                    # sheet名与货号匹配时优先（如搜77794，sheet"77794"优先于"77772"）
                    sn_digits = re.match(r'\d+', sn)
                    sn_matches_item = bool(sn_digits and item_code and sn_digits.group() == item_code)
                    if not best:
                        best = result
                        best['_sheet_match'] = sn_matches_item
                    elif sn_matches_item and not best.get('_sheet_match'):
                        best = result
                        best['_sheet_match'] = True
                    elif not sn_matches_item and best.get('_sheet_match'):
                        pass  # 保留已匹配sheet名的best
                    elif cnt > best['cnt']:
                        best = result
                        best['_sheet_match'] = sn_matches_item
                elif target_sheet and sn in matched_target_sheets:
                    if last_data_row:
                        logging.info(f"[auto_find] {fn}/{sn} 无精确匹配行，使用最后数据行{last_data_row}作参考")
                        best = {'file': fp, 'fname': fn, 'sheet': sn, 'ref': last_data_row,
                                'cnt': 1, 'mcol': max(ws.max_column or 30, 50), 'po_type_match': False}
            except Exception as e:
                logging.debug(f"[SKU搜索] {fn}/{sn} read_only模式跳过: {e}")
                # 如果是目标sheet且read_only模式失败，用非只读模式重试
                if target_sheet and sn in matched_target_sheets:
                    logging.info(f"[auto_find] {fn}/{sn} read_only失败，尝试非只读模式重试")
                    wb2 = None
                    try:
                        wb2 = openpyxl.load_workbook(fp, read_only=False, data_only=True)
                        ws2 = wb2[sn]
                        ref_spec2 = None
                        ref_retry = None
                        cnt_retry = 0
                        last_row2 = None
                        for r in range(2, ws2.max_row + 1):
                            has_data = any(ws2.cell(r, c).value for c in range(1, min(9, ws2.max_column + 1)))
                            if has_data:
                                last_row2 = r
                            for ci in [7, 6, 8]:  # G, F, H (1-based)
                                cv = ws2.cell(r, ci).value
                                if not cv:
                                    continue
                                cv_str = str(cv).strip()
                                cv_item = _item_code(cv_str)
                                if not cv_item:
                                    continue
                                cv_sp = _sku_spec(cv_str)
                                if sku_spec and cv_sp == sku_spec:
                                    ref_spec2 = r
                                    cnt_retry += 1
                                    break
                                elif item_code and cv_item == item_code:
                                    ref_retry = r
                                    cnt_retry += 1
                                    break
                                elif item_code and cv_item and \
                                     abs(len(cv_item) - len(item_code)) <= 3 and \
                                     (cv_item.startswith(item_code) or item_code.startswith(cv_item)):
                                    if not ref_retry:
                                        ref_retry = r
                                    cnt_retry += 1
                                    break
                        ref_retry = ref_spec2 or ref_retry
                        if ref_retry:
                            best = {'file': fp, 'fname': fn, 'sheet': sn, 'ref': ref_retry,
                                    'cnt': cnt_retry, 'mcol': ws2.max_column or 30,
                                    'po_type_match': False}
                            logging.info(f"[auto_find] 非只读重试成功: {fn}/{sn} 行{ref_retry}")
                        elif last_row2:
                            best = {'file': fp, 'fname': fn, 'sheet': sn, 'ref': last_row2,
                                    'cnt': 1, 'mcol': ws2.max_column or 30,
                                    'po_type_match': False}
                            logging.info(f"[auto_find] 非只读重试: {fn}/{sn} 使用最后数据行{last_row2}")
                    except Exception as e2:
                        logging.warning(f"[auto_find] {fn}/{sn} 非只读重试也失败: {e2}")
                    finally:
                        if wb2:
                            try: wb2.close()
                            except: pass
                        if not best:
                            # openpyxl完全无法读取，使用WPS COM后备搜索
                            try:
                                logging.info(f"[auto_find] {fn}/{sn} 启用WPS COM后备搜索")
                                com_result = self._search_sku_com(fp, fn, sn, item_code, sku_spec=sku_spec)
                                if com_result:
                                    best = com_result
                            except Exception as e3:
                                logging.warning(f"[auto_find] {fn}/{sn} COM后备也失败: {e3}")
                continue
        # wb由_wb_cache统一管理生命周期，不在此处close
        # 填充content cache（下次搜索同文件时可快速排除）
        if _items_by_sheet:
            ExcelHandler._set_content_cache(fp, wb.sheetnames, _items_by_sheet)
        if best:
            best.pop('_sheet_match', None)  # 清理临时标记
        return best

    def _search_sku_com(self, fp, fn, sheet_name, item_code, sku_spec=''):
        """WPS COM后备搜索：当openpyxl无法读取时用COM打开文件搜索ITEM#列
        sku_spec: 完整规格码（含-SXXX后缀），用于精确区分变体"""
        import pythoncom
        import win32com.client
        pythoncom.CoInitialize()
        app = None
        wb = None
        try:
            for pid in ['Ket.Application', 'Et.Application', 'Excel.Application']:
                try:
                    app = win32com.client.DispatchEx(pid)
                    break
                except:
                    continue
            if not app:
                return None
            app.Visible = False
            app.DisplayAlerts = False
            wb = app.Workbooks.Open(fp, ReadOnly=True)
            ws = None
            for i in range(1, wb.Sheets.Count + 1):
                if wb.Sheets(i).Name == sheet_name:
                    ws = wb.Sheets(i)
                    break
            if not ws:
                return None
            max_row = ws.UsedRange.Rows.Count + ws.UsedRange.Row - 1
            max_col = ws.UsedRange.Columns.Count + ws.UsedRange.Column - 1
            ref_spec = None
            ref = None
            cnt = 0
            last_data_row = None
            for r in range(2, min(max_row + 1, 2000)):
                has_data = False
                for c in range(1, min(9, max_col + 1)):
                    v = ws.Cells(r, c).Value
                    if v:
                        has_data = True
                        break
                if has_data:
                    last_data_row = r
                # 搜索G(7), F(6), H(8)列
                for ci in [7, 6, 8]:
                    cv = ws.Cells(r, ci).Value
                    if not cv:
                        continue
                    cv_str = str(cv).strip()
                    cv_item = _item_code(cv_str)
                    if not cv_item:
                        continue
                    cv_sp = _sku_spec(cv_str)
                    # SKU-SPEC精确匹配（最高优先级）
                    if sku_spec and cv_sp == sku_spec:
                        ref_spec = r
                        cnt += 1
                        break
                    elif item_code and cv_item == item_code:
                        ref = r
                        cnt += 1
                        break
                    elif item_code and cv_item and \
                         abs(len(cv_item) - len(item_code)) <= 3 and \
                         (cv_item.startswith(item_code) or item_code.startswith(cv_item)):
                        if not ref:
                            ref = r
                        cnt += 1
                        break
            ref = ref_spec or ref
            if ref:
                logging.info(f"[auto_find] COM后备成功: {fn}/{sheet_name} 行{ref}")
                return {'file': fp, 'fname': fn, 'sheet': sheet_name, 'ref': ref,
                        'cnt': cnt, 'mcol': max_col or 30, 'po_type_match': False}
            elif last_data_row:
                logging.info(f"[auto_find] COM后备: {fn}/{sheet_name} 使用最后数据行{last_data_row}")
                return {'file': fp, 'fname': fn, 'sheet': sheet_name, 'ref': last_data_row,
                        'cnt': 1, 'mcol': max_col or 30, 'po_type_match': False}
            return None
        except Exception as e:
            logging.warning(f"[auto_find] COM后备搜索异常: {e}")
            return None
        finally:
            try:
                if wb:
                    wb.Close(False)
            except:
                pass
            try:
                if app:
                    app.Quit()
            except:
                pass
            try:
                pythoncom.CoUninitialize()
            except:
                pass

    def search_po(self, po_number):
        po_s = str(po_number)
        po_i = int(po_s) if po_s.isdigit() else None
        results = []
        for fp in self._list_xlsx():
            fn = os.path.basename(fp)
            if _should_skip_file(fn):
                continue
            try:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            except:
                continue
            try:
                for sn in wb.sheetnames:
                    if '取消' in sn or _should_skip_sheet(sn):
                        continue
                    try:
                        ws = wb[sn]
                        for row in ws.iter_rows(min_row=2, max_col=30):
                            row_num = getattr(row[0], 'row', None)
                            if row_num is None:
                                for c in row[1:]:
                                    row_num = getattr(c, 'row', None)
                                    if row_num is not None:
                                        break
                            if row_num is None:
                                continue
                            d = row[3].value if len(row) > 3 else None
                            e = row[4].value if len(row) > 4 else None
                            hit = False
                            if d and (str(d) == po_s or (po_i and d == po_i) or _po_boundary_match(po_s, str(d))):
                                hit = True
                            elif e and (str(e) == po_s or (po_i and e == po_i) or _po_boundary_match(po_s, str(e))):
                                hit = True
                            if hit:
                                data = {}
                                for c in row:
                                    try:
                                        col_num = getattr(c, 'column', None)
                                        if col_num is None:
                                            continue
                                        cl = openpyxl.utils.get_column_letter(col_num)
                                        v = c.value
                                        if isinstance(v, datetime):
                                            v = v.strftime('%Y-%m-%d')
                                        data[cl] = v
                                    except:
                                        pass
                                results.append({'file': fp, 'fname': fn,
                                                'sheet': sn, 'row': row_num, 'data': data})
                    except Exception as e:
                        logging.debug(f"[PO搜索] {fn}/{sn} 跳过: {e}")
                        continue
            finally:
                try: wb.close()
                except: pass
        return results

    def combined_search(self, po_list, all_skus, candidate_map):
        """合并搜索：每个候选文件只打开一次，同时搜PO和SKU参考行。
        po_list: PO号列表
        all_skus: {sku: (po, ln_key)} 需要查找的SKU
        candidate_map: {file_path: set(sheet_names)} ISM候选
        返回 (po_results, sku_refs):
          po_results = {po_str: [{'file','fname','sheet','row','data'}, ...]}
          sku_refs = {sku: {'file','fname','sheet','ref','cnt','mcol','po_type_match'} or None}
        """
        import time as _t
        from collections import defaultdict
        po_set = set()
        po_int_set = {}
        for p in po_list:
            if not p: continue
            ps = str(p).strip()
            po_set.add(ps)
            if ps.isdigit():
                po_int_set[int(ps)] = ps
        po_results = defaultdict(list)

        # ISM路径→实际z_path路径映射（测试版副本路径不同于Z:盘）
        if self.z_path.rstrip('\\') != _DEFAULT_ZURU_PATH.rstrip('\\'):
            candidate_map = {self._remap_ism_path(fp): sheets
                             for fp, sheets in candidate_map.items()}

        # 预处理SKU：按候选文件分组
        sku_by_file = defaultdict(list)  # {fp: [(sku, target_sheet, item_code, sku_spec, po)]}
        _ism = self._get_item_schedule_map()
        _remap_cache = {}  # 缓存_remap_ism_path结果，避免重复磁盘检查
        for sku, (po, ln_key) in all_skus.items():
            entries = self._ism_lookup(_ism, sku)
            if entries:
                for entry in entries:
                    _raw_path = entry.get('path', '')
                    if _raw_path in _remap_cache:
                        fp = _remap_cache[_raw_path]
                    else:
                        fp = self._remap_ism_path(_raw_path)
                        _remap_cache[_raw_path] = fp
                    ts = entry.get('sheet', '')
                    if fp and fp in candidate_map:
                        item = _item_code(sku)
                        spec = _sku_spec(sku)
                        sku_by_file[fp].append((sku, ts, item, spec, po))
                        break
        sku_refs = {}

        # ISM找不到SKU时，把所有SKU分配到所有候选文件（全盘搜索模式）
        _is_fallback = False
        if not sku_by_file and all_skus:
            _fallback_tasks = []
            for sku, (po, ln_key) in all_skus.items():
                item = _item_code(sku)
                spec = _sku_spec(sku)
                _fallback_tasks.append((sku, '', item, spec, po))
            for fp in candidate_map:
                sku_by_file[fp] = _fallback_tasks  # 只读共享引用，勿修改
            _is_fallback = True

        # 预计算列名 A-L（避免循环调用get_column_letter）
        _COL_LETTERS = [''] + [chr(64 + i) for i in range(1, 27)] + ['A' + chr(64 + i) for i in range(1, 7)]  # A-AF(32列)，覆盖smart_diff所需

        # 预构建SKU快速查找表（spec→sku、item→多个sku）
        _spec_to_sku = {}  # {spec_upper: sku}
        _item_to_skus = {}  # {item_upper: [sku1, sku2, ...]} 一对多，同item不同spec
        _sku_spec_cache = {}  # {sku: spec_upper} 预算缓存，避免热路径重复调用
        # fallback模式所有文件共享同一任务列表，只需迭代一次
        _unique_tasks = [_fallback_tasks] if _is_fallback else sku_by_file.values()
        for tasks in _unique_tasks:
            for (sku, ts, item, spec, po) in tasks:
                if spec:
                    _spec_to_sku.setdefault(spec.upper(), sku)
                _sku_spec_cache[sku] = spec.upper() if spec else ''
                if item:
                    _item_to_skus.setdefault(item.upper(), [])
                    if sku not in _item_to_skus[item.upper()]:
                        _item_to_skus[item.upper()].append(sku)

        def _scan_combined(fp):
            """单个文件：同时搜PO和SKU。返回 (_po_hits, _sku_hits, _skipped_reason)"""
            _t0 = _t.time()
            fn = os.path.basename(fp)
            _po_hits = []
            _sku_hits = {}  # {sku: best_ref_dict}
            # 被占用文件（.xlsx0.xlsx或~$锁文件）openpyxl read_only仍可读取，不跳过
            try:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            except Exception:
                return _po_hits, _sku_hits, fn  # 打开失败才跳过
            try:
                _ts_set = candidate_map.get(fp, set())
                _sku_tasks = sku_by_file.get(fp, [])
                mc = 30  # smart_diff需要M-P(出货期)和R+(单价)，必须≥30列
                mcol_result = 50  # 返回给_do_new_com用的列数，至少50列防止_detect_cols截断

                for sn in wb.sheetnames:
                    if '取消' in sn or _should_skip_sheet(sn):
                        continue
                    _scan_po = True
                    _need_sku = [t for t in _sku_tasks
                                 if not t[1] or t[1] == sn or t[1] in sn or sn in t[1]]
                    if _ts_set and sn not in _ts_set and not any(t in sn or sn in t for t in _ts_set):
                        _scan_po = False
                    if not _scan_po and not _need_sku:
                        continue

                    try:
                        ws = wb[sn]
                        _empty_streak = 0
                        _total_sku_need = sum(1 for t in _sku_tasks if t[0] not in _sku_hits)
                        # 检测表头列位置（出货期/数量/单价/客户PO），供smart_diff使用
                        _sheet_col_map = {}  # {字段名: 列字母}
                        try:
                            for _hr in ws.iter_rows(min_row=1, max_row=4, max_col=mc):
                                for _hc in _hr:
                                    _hv = _hc.value
                                    if not _hv or not isinstance(_hv, str):
                                        continue
                                    _hvl = _hv.strip().replace(' ', '')
                                    _hvu = _hvl.upper()
                                    _cn = getattr(_hc, 'column', 0)
                                    if _cn <= 0 or _cn >= len(_COL_LETTERS):
                                        continue
                                    _cl = _COL_LETTERS[_cn]
                                    if ('出货' in _hvl or '出期' in _hvl or _hvu == 'CRD') and '情况' not in _hvl and 'ship_date' not in _sheet_col_map:
                                        _sheet_col_map['ship_date'] = _cl
                                    elif ('数量' in _hvl and '合计' not in _hvl and '箱' not in _hvl) or _hvu in ('QTY', 'QTYPCS', 'PO数量'):
                                        if 'qty' not in _sheet_col_map:
                                            _sheet_col_map['qty'] = _cl
                                    elif ('单价' in _hvl or _hvu in ('PRICE', 'PRICEUSD', '单价USD')) and 'price' not in _sheet_col_map:
                                        _sheet_col_map['price'] = _cl
                                    elif ('客PO' in _hvl or '客户PO' in _hvl or '小PO' in _hvl or _hvu in ('CUSTOMERPO', 'CUSTPO')) and '期' not in _hvl and 'customer_po' not in _sheet_col_map:
                                        _sheet_col_map['customer_po'] = _cl
                                    elif ('系统' in _hvl and '货号' in _hvl) and 'system_code' not in _sheet_col_map:
                                        _sheet_col_map['system_code'] = _cl
                                    elif (_hvu in ('ITEM#', 'ITEM', 'ITEM＃') or ('货号' in _hvl and '系统' not in _hvl)) and 'item' not in _sheet_col_map:
                                        _sheet_col_map['item'] = _cl
                        except Exception:
                            pass  # 表头检测失败不影响搜索
                        for row in ws.iter_rows(min_row=2, max_col=mc, max_row=2000):
                            row_num = getattr(row[0], 'row', None)
                            if row_num is None:
                                for c in row[1:]:
                                    row_num = getattr(c, 'row', None)
                                    if row_num is not None:
                                        break
                            if row_num is None:
                                continue
                            # 连续空行检测：30行无数据则停止本sheet
                            if all(c.value is None for c in row[:8]):
                                _empty_streak += 1
                                if _empty_streak >= 30:
                                    break
                                continue
                            _empty_streak = 0

                            # === PO匹配（列D,E）===
                            if _scan_po and po_set:
                                d = row[3].value if len(row) > 3 else None
                                e = row[4].value if len(row) > 4 else None
                                matched_po = None
                                for val in [d, e]:
                                    if val is None: continue
                                    vs = str(val).strip()
                                    if vs in po_set:
                                        matched_po = vs; break
                                    if isinstance(val, (int, float)) and int(val) in po_int_set:
                                        matched_po = po_int_set[int(val)]; break
                                    for ps in po_set:
                                        if _po_boundary_match(ps, vs):
                                            matched_po = ps; break
                                    if matched_po: break
                                if matched_po:
                                    data = {}
                                    for i, c in enumerate(row):
                                        if i >= len(_COL_LETTERS): break
                                        v = c.value
                                        if v is not None:
                                            if isinstance(v, datetime):
                                                v = v.strftime('%Y-%m-%d')
                                            data[_COL_LETTERS[i + 1]] = v
                                    _po_hits.append((matched_po, {'file': fp, 'fname': fn,
                                                                   'sheet': sn, 'row': row_num, 'data': data,
                                                                   'col_map': _sheet_col_map}))

                            # === SKU参考行匹配（列F,G,H）===
                            if _need_sku and _total_sku_need > 0:
                                for ci in [6, 5, 7]:  # G=7, F=6, H=8 (0-based: 6,5,7)
                                    cv = row[ci].value if len(row) > ci else None
                                    if not cv: continue
                                    cv_str = str(cv).strip()
                                    cv_sp = _sku_spec(cv_str)
                                    cv_item = _item_code(cv_str)
                                    if not cv_item: continue
                                    # dict查找替代线性扫描
                                    _matched_any = False
                                    # 1) spec精确匹配（如15789UQ1-S002）
                                    if cv_sp:
                                        _ms = _spec_to_sku.get(cv_sp.upper())
                                        if _ms and _ms not in _sku_hits:
                                            _sku_hits[_ms] = {
                                                'file': fp, 'fname': fn, 'sheet': sn, 'ref': row_num,
                                                'cnt': 1, 'mcol': mcol_result, 'po_type_match': True
                                            }
                                            _total_sku_need -= 1
                                            _matched_any = True
                                    # 2) item匹配：同item_code的行可作为参考行
                                    #    spec完全一致优先，spec不同但item相同也允许（如77858-S001→77858-S002）
                                    #    但排除复合货号（如77785-S001-77673-INT含嵌入产品号77673）
                                    if cv_item and not _matched_any:
                                        # 检查是否复合货号（含嵌入的≥4位纯数字段）
                                        _cv_parts = cv_str.upper().split('-')
                                        _cv_is_compound = any(p.isdigit() and len(p) >= 4 for p in _cv_parts[1:])
                                        if _cv_is_compound:
                                            break  # 复合货号只用spec精确匹配，不走item兜底
                                        _cv_item_up = cv_item.upper()
                                        _cv_sp_up = cv_sp.upper() if cv_sp else ''
                                        for _ms in _item_to_skus.get(_cv_item_up, []):
                                            _ms_sp = _sku_spec_cache.get(_ms, '')
                                            if _cv_sp_up and _ms_sp and _cv_sp_up == _ms_sp:
                                                # spec完全一致 — 最佳匹配，覆盖弱匹配
                                                _was_new = _ms not in _sku_hits
                                                _sku_hits[_ms] = {
                                                    'file': fp, 'fname': fn, 'sheet': sn, 'ref': row_num,
                                                    'cnt': 1, 'mcol': mcol_result,
                                                    'po_type_match': True
                                                }
                                                if _was_new:
                                                    _total_sku_need -= 1
                                                _matched_any = True
                                            elif _ms not in _sku_hits:
                                                # item_code相同但spec不同 — 弱匹配，可被后续精确匹配覆盖
                                                _sku_hits[_ms] = {
                                                    'file': fp, 'fname': fn, 'sheet': sn, 'ref': row_num,
                                                    'cnt': 1, 'mcol': mcol_result,
                                                    'po_type_match': False
                                                }
                                                _total_sku_need -= 1
                                                _matched_any = True
                                    break  # 只检查第一个有值的列
                    except Exception:
                        continue
            finally:
                try:
                    if hasattr(wb, '_archive') and wb._archive:
                        wb._archive.close()
                    wb.close()
                except:
                    pass
            logging.info(f"[合并搜索] {fn}: {_t.time()-_t0:.1f}s (PO命中{len(_po_hits)} SKU命中{len(_sku_hits)})")
            return _po_hits, _sku_hits, None  # None=未跳过

        # 并行处理所有候选文件（每个文件独立打开自己的workbook，无共享冲突）
        from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FuturesTimeoutError
        _files = list(candidate_map.keys())
        workers = min(3, len(_files))  # 网络Z:盘并发过高反而慢
        skipped_files = []  # 因锁定/打开失败而跳过的文件名列表
        with ThreadPoolExecutor(max_workers=max(workers, 1)) as pool:
            _futs = {pool.submit(_scan_combined, fp): fp for fp in _files}
            try:
                for fut in as_completed(_futs, timeout=120):
                    try:
                        _po_hits, _sku_hits, _skipped = fut.result(timeout=15)
                        if _skipped:
                            skipped_files.append(_skipped)
                        for po_str, rec in _po_hits:
                            po_results[po_str].append(rec)
                        for sku, ref in _sku_hits.items():
                            if sku not in sku_refs:
                                sku_refs[sku] = ref
                    except Exception as e:
                        logging.warning(f"[合并搜索] 文件处理失败: {e}")
            except FuturesTimeoutError:
                logging.warning("[合并搜索] 超时(>60s)")

        if skipped_files:
            logging.info(f"[合并搜索] 跳过{len(skipped_files)}个被占用文件: {', '.join(skipped_files[:5])}")
        # 排除总排期/样板文件
        for k in po_results:
            po_results[k] = [r for r in po_results[k] if not _should_skip_file(r.get('fname', ''))]
        return dict(po_results), sku_refs, skipped_files

    def batch_search_pos(self, po_list, target_files=None, target_sheets=None):
        """并行搜索多个PO号。target_files指定只搜特定文件，target_sheets={fp: set(sheet_names)}限定sheet。"""
        from concurrent.futures import ThreadPoolExecutor, as_completed
        po_set = set()
        po_int_set = {}
        for p in po_list:
            if not p:
                continue
            ps = str(p).strip()
            po_set.add(ps)
            if ps.isdigit():
                po_int_set[int(ps)] = ps
        if not po_set:
            return {}
        all_results = {p: [] for p in po_set}

        # ZIP预筛选：先检查xlsx内sharedStrings.xml是否包含PO号，跳过不含的文件
        def _zip_has_po(fp):
            """快速检查xlsx(ZIP)内是否包含任一PO号（文本型查sharedStrings，数字型查worksheet XML）"""
            import zipfile
            try:
                with zipfile.ZipFile(fp, 'r') as zf:
                    names = zf.namelist()
                    # 1. 文本型：sharedStrings.xml
                    if 'xl/sharedStrings.xml' in names:
                        content = zf.read('xl/sharedStrings.xml').decode('utf-8', errors='ignore')
                        for ps in po_set:
                            if ps in content:
                                return True
                    # 2. 数字型：worksheet XML中的 <v>4500xxxxxx</v>
                    ws_files = sorted(n for n in names
                                      if n.startswith('xl/worksheets/sheet') and n.endswith('.xml'))
                    for ws_file in ws_files:
                        ws_content = zf.read(ws_file).decode('utf-8', errors='ignore')
                        for ps in po_set:
                            if f'>{ps}<' in ws_content:
                                return True
                    return False
            except Exception:
                return True  # 出错则保守处理，仍然扫描

        if target_files:
            _raw = [(fp, os.path.basename(fp)) for fp in target_files
                     if not _should_skip_file(os.path.basename(fp))]
        else:
            _raw = [(fp, os.path.basename(fp)) for fp in self._list_xlsx()
                     if not _should_skip_file(os.path.basename(fp))]
        # 并行ZIP预筛选
        if len(_raw) > 3:
            _t_filter = __import__('time').time()
            from concurrent.futures import ThreadPoolExecutor as _TPE
            with _TPE(max_workers=min(8, len(_raw))) as _fpool:
                _checks = list(_fpool.map(lambda x: (x, _zip_has_po(x[0])), _raw))
            files = [item for item, has in _checks if has]
            _skipped_count = len(_raw) - len(files)
            logging.info(f"[batch_search_pos] ZIP预筛选: {len(_raw)}→{len(files)}个文件({_skipped_count}个不含PO号), 耗时{__import__('time').time()-_t_filter:.1f}s")
        else:
            files = _raw

        def _scan_file(args):
            fp, fn = args
            hits = []
            try:
                wb = ExcelHandler._get_cached_wb(fp)
            except Exception:
                return hits
            # 如果有目标sheet列表，只扫描匹配的sheet（大幅减少IO）
            _ts = target_sheets.get(fp) if target_sheets else None
            for sn in wb.sheetnames:
                if '取消' in sn or _should_skip_sheet(sn):
                    continue
                if _ts and sn not in _ts:
                    if not any(t in sn or sn in t for t in _ts):
                        continue
                try:
                    ws = wb[sn]
                    # 检测表头列位置（供smart_diff取消检测用）
                    _scm = {}
                    try:
                        for _hr in ws.iter_rows(min_row=1, max_row=4, max_col=15):
                            for _hc in _hr:
                                _hv = _hc.value
                                if not _hv or not isinstance(_hv, str):
                                    continue
                                _hvl = _hv.strip().replace('\n','').replace(' ','').lower()
                                _hvu = _hvl.upper()
                                _cn = getattr(_hc, 'column', None)
                                if _cn is None or _cn > 26:
                                    continue
                                _cl = chr(64 + _cn)
                                if ('系统' in _hvl and '货号' in _hvl) and 'system_code' not in _scm:
                                    _scm['system_code'] = _cl
                                elif (_hvu in ('ITEM#', 'ITEM', 'ITEM＃') or ('货号' in _hvl and '系统' not in _hvl)) and 'item' not in _scm:
                                    _scm['item'] = _cl
                    except Exception:
                        pass
                    for row in ws.iter_rows(min_row=2, max_col=30):
                        row_num = getattr(row[0], 'row', None)
                        if row_num is None:
                            for c in row[1:]:
                                row_num = getattr(c, 'row', None)
                                if row_num is not None:
                                    break
                        if row_num is None:
                            continue
                        d = row[3].value if len(row) > 3 else None
                        e = row[4].value if len(row) > 4 else None
                        matched = None
                        for val in [d, e]:
                            if val is None:
                                continue
                            vs = str(val).strip()
                            if vs in po_set:
                                matched = vs
                                break
                            if isinstance(val, (int, float)) and int(val) in po_int_set:
                                matched = po_int_set[int(val)]
                                break
                            for ps in po_set:
                                if _po_boundary_match(ps, vs):
                                    matched = ps
                                    break
                            if matched:
                                break
                        if matched:
                            data = {}
                            for c in row:
                                try:
                                    cn = getattr(c, 'column', None)
                                    if cn is None:
                                        continue
                                    cl = openpyxl.utils.get_column_letter(cn)
                                    v = c.value
                                    if isinstance(v, datetime):
                                        v = v.strftime('%Y-%m-%d')
                                    data[cl] = v
                                except:
                                    pass
                            hits.append((matched, {'file': fp, 'fname': fn,
                                                   'sheet': sn, 'row': row_num, 'data': data,
                                                   'col_map': _scm}))
                except:
                    continue
            # wb由_wb_cache统一管理生命周期，不在此处close
            return hits

        if not files:
            return all_results
        workers = min(6, len(files))
        from concurrent.futures import as_completed, TimeoutError as FuturesTimeoutError
        with ThreadPoolExecutor(max_workers=workers) as pool:
            _futs = {pool.submit(_scan_file, fp): fp for fp in files}
            try:
                for fut in as_completed(_futs, timeout=120):
                    try:
                        result = fut.result(timeout=15)
                        for po_str, rec in result:
                            all_results[po_str].append(rec)
                    except Exception as e:
                        logging.warning(f"[batch_search_pos] 文件扫描失败: {e}")
            except FuturesTimeoutError:
                _pending = [fp[1] if isinstance(fp, tuple) else os.path.basename(fp)
                            for fut, fp in _futs.items() if not fut.done()]
                logging.warning(f"[batch_search_pos] 超时(>120s)，跳过: {_pending}")
        return all_results

    def search_by_skus(self, lines):
        """当PO号搜不到时，通过SKU/商品代码在排期中搜索现有记录
        1. 用auto_find定位排期文件和工作表
        2. 在该工作表中搜索所有包含PDF商品代码的行"""
        results = []
        # 收集所有item code
        code_set = set()
        for ln in lines:
            for field in ('sku', 'item_code'):
                v = ln.get(field, '')
                code = _item_code(v)
                if code:
                    code_set.add(code)
        if not code_set:
            return results

        # 用第一个SKU定位排期文件/工作表
        target_files = {}  # file_path → set of sheet names
        for ln in lines:
            sku = ln.get('item_code') or ln.get('sku', '')
            found = self.auto_find(sku)
            if found:
                fp = found['file']
                if fp not in target_files:
                    target_files[fp] = set()
                target_files[fp].add(found['sheet'])
        if not target_files:
            return results

        # 在目标文件的目标sheet中搜索包含任一商品代码的行
        for fp, sheets in target_files.items():
            fn = os.path.basename(fp)
            if _should_skip_file(fn):
                continue
            try:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            except:
                continue
            for sn in sheets:
                if sn not in wb.sheetnames or '取消' in sn or _should_skip_sheet(sn):
                    continue
                try:
                    ws = wb[sn]
                    for row in ws.iter_rows(min_row=2, max_col=30):
                        row_num = getattr(row[0], 'row', None)
                        if row_num is None:
                            for c in row[1:]:
                                row_num = getattr(c, 'row', None)
                                if row_num is not None:
                                    break
                        if row_num is None:
                            continue
                        hit = False
                        for c in row[:10]:
                            if c.value:
                                code = _item_code(str(c.value))
                                if code and code in code_set:
                                    hit = True
                                    break
                        if hit:
                            data = {}
                            for c in row:
                                try:
                                    col_num = getattr(c, 'column', None)
                                    if col_num is None:
                                        continue
                                    cl = openpyxl.utils.get_column_letter(col_num)
                                    v = c.value
                                    if isinstance(v, datetime):
                                        v = v.strftime('%Y-%m-%d')
                                    data[cl] = v
                                except:
                                    pass
                            results.append({'file': fp, 'fname': fn,
                                            'sheet': sn, 'row': row_num, 'data': data})
                except:
                    continue
            try:
                wb.close()
            except:
                pass
        return results

    def fuzzy_search(self, keyword):
        """模糊搜索：支持PO号、SKU、客户名等"""
        kw = str(keyword).strip()
        kw_lower = kw.lower()
        kw_num = re.sub(r'[^0-9]', '', kw)
        results = []
        for fp in self._list_xlsx():
            fn = os.path.basename(fp)
            if _should_skip_file(fn):
                continue
            try:
                wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            except:
                continue
            for sn in wb.sheetnames:
                if '取消' in sn or _should_skip_sheet(sn):
                    continue
                try:
                    ws = wb[sn]
                    for row in ws.iter_rows(min_row=2, max_col=30):
                        row_num = getattr(row[0], 'row', None)
                        if row_num is None:
                            for c in row[1:]:
                                row_num = getattr(c, 'row', None)
                                if row_num is not None:
                                    break
                        if row_num is None:
                            continue
                        hit = False
                        hit_col = ''
                        for c in row[:10]:
                            if c.value:
                                cv = str(c.value)
                                col_num = getattr(c, 'column', 0)
                                if kw_lower in cv.lower() or (kw_num and len(kw_num) >= 4 and kw_num in cv):
                                    hit = True
                                    col_names = {1:'接单日期', 2:'客户', 3:'目的地', 4:'PO号',
                                                 5:'客户PO', 6:'SKU', 7:'品名', 9:'数量', 13:'出货日期'}
                                    hit_col = col_names.get(col_num, f'列{col_num}')
                                    break
                        if hit:
                            data = {}
                            for c in row:
                                try:
                                    cn = getattr(c, 'column', None)
                                    if cn is None:
                                        continue
                                    cl = openpyxl.utils.get_column_letter(cn)
                                    v = c.value
                                    if isinstance(v, datetime):
                                        v = v.strftime('%Y-%m-%d')
                                    data[cl] = v
                                except:
                                    pass
                            results.append({
                                'file': fp, 'fname': fn, 'sheet': sn,
                                'row': row_num, 'data': data, 'hit_col': hit_col
                            })
                            if len(results) >= 100:
                                try:
                                    wb.close()
                                except:
                                    pass
                                return results
                except:
                    continue
            try:
                wb.close()
            except:
                pass
        return results

    # =================== 智能对比（纯逻辑） ===================

    @staticmethod
    def _find_rec_po(rd, po_hint=''):
        """从记录data中提取PO号，搜索D/E列（不同排期PO列位置不同）
        po_hint: PDF的PO号，用于判断哪一列包含真正的PO号"""
        for c in 'DE':
            v = str(rd.get(c, '') or '').strip()
            if v and po_hint and _po_boundary_match(po_hint, v):
                return v
        # 没有精确匹配时，返回D列值（兼容原逻辑）
        return str(rd.get('D', '') or '').strip()

    def smart_diff(self, pdf_data, existing_records):
        """比对PDF与现有记录，生成新增/修改操作
        原则：
        1. 同PO记录：检测变化→修改；无变化→跳过
        2. 不同PO记录：只标记该商品代码已存在→不重复添加，不修改
        3. 未匹配记录：不自动取消（取消由用户手动操作）
        匹配策略：PO-line精确匹配 → 完整商品代码匹配"""
        actions = []
        new_lines = pdf_data.get('lines', [])
        po = pdf_data.get('po_number', '')

        # ===== 1. 构建PDF行查找表 =====
        new_by_code = {}       # 完整商品代码 → [lines...] (同码多行用列表)
        new_by_spec = {}       # 完整sku_spec(大写) → [lines...] (同spec多行用列表)
        new_by_poline = {}     # "PO-lineNo" → line

        _poline_conflicts = set()  # 多组件共享同一PO-line的冲突key
        # 预计算：PDF中是否有混装行（同line>=2行 或 is_mixed_carton标记）
        _has_mixed_lines = any(l.get('is_mixed_carton') for l in new_lines)
        for ln in new_lines:
            sku = ln.get('sku', '')
            code = _item_code(sku)
            if code:
                new_by_code.setdefault(code, []).append(ln)
            ic = ln.get('item_code', '')
            ic_code = _item_code(ic)
            if ic_code and ic_code != code:
                new_by_code.setdefault(ic_code, []).append(ln)
            # 严格sku_spec匹配表
            spec = (ln.get('sku_spec', '') or sku).strip().upper()
            if spec:
                new_by_spec.setdefault(spec, []).append(ln)
            line_no = ln.get('line_no', '')
            if po and line_no:
                key = f"{po}-{line_no}"
                if key in new_by_poline:
                    del new_by_poline[key]
                    _poline_conflicts.add(key)
                elif key not in _poline_conflicts:
                    new_by_poline[key] = ln

        matched_polines = set()      # 已匹配的PO-line键
        matched_pdf_line_nos = set() # 已匹配到existing record的PDF行line_no

        # ===== 2. 逐条匹配已有记录 =====
        for rec in existing_records:
            rd = rec['data']
            matched_ln = None
            match_code = None

            # 检查D-J列寻找匹配
            for col in 'FGHDEIJ':
                v = rd.get(col)
                if not v:
                    continue
                vs = str(v).strip()
                if not vs:
                    continue
                # 策略1：PO-line精确匹配（如 "4500193745-10"）
                if vs in new_by_poline:
                    _candidate_ln = new_by_poline[vs]
                    _cand_item = _item_code(_candidate_ln.get('sku', ''))
                    # 交叉验证：PDF行货号必须在该行F/G/H列中出现，防止跨货号误匹配
                    # （例如F列存"4500195690-30"作为SKU，与7133的PO-line key巧合相同）
                    _row_items = set()
                    for _ic in 'FGH':
                        _iv = rd.get(_ic)
                        if _iv:
                            _ri = _item_code(str(_iv).strip())
                            if _ri:
                                _row_items.add(_ri)
                    if not _row_items or (_cand_item and _cand_item in _row_items):
                        matched_ln = _candidate_ln
                        match_code = _cand_item
                        matched_polines.add(vs)
                        break
                # 策略2：货号严格完全匹配（如排期"15746A-S001-PKC"必须完全等于PDF的sku_spec）
                vs_upper = vs.strip().upper()
                if vs_upper in new_by_spec and new_by_spec[vs_upper]:
                    matched_ln = new_by_spec[vs_upper][0]
                    match_code = _item_code(matched_ln.get('sku', ''))
                    break
                # 策略3兜底：完整商品代码匹配 — 只在货号列(F/G/H)做，避免PO号/客PO号误匹配
                if col in ('F', 'G', 'H'):
                    code = _item_code(vs)
                    if code and code in new_by_code and new_by_code[code]:
                        # 防止复合货号误匹配：如排期"77785-S001-77673-INT"不应匹配PDF"77785-S001-INT"
                        # 检查排期值是否含嵌入的其他产品编号（-后跟纯数字段且非S/P前缀）
                        _vs_parts = vs_upper.split('-')
                        _has_embedded = any(p.isdigit() and len(p) >= 4 for p in _vs_parts[1:])
                        if _has_embedded:
                            # 排期值是复合货号，必须用sku_spec精确匹配（策略2已处理），跳过策略3
                            pass
                        else:
                            matched_ln = new_by_code[code][0]
                            match_code = code
                            break

            if matched_ln:
                # 判断是否同一PO：只有同PO才比较和修改（数字边界匹配，避免子串误判）
                rec_po = self._find_rec_po(rd, po)
                same_po = bool(po and rec_po and _po_boundary_match(po, rec_po))

                if not same_po:
                    # 不同PO → 只标记已存在，不做任何修改
                    continue

                # 同PO → 记录已匹配的PDF行line_no，并从查找表中移除已消费的行
                _ml_line_no = matched_ln.get('line_no', '')
                if _ml_line_no:
                    matched_pdf_line_nos.add(_ml_line_no)
                # 从列表中移除已消费的PDF行，避免后续existing record重复匹配同一行
                if match_code and match_code in new_by_code:
                    new_by_code[match_code] = [l for l in new_by_code[match_code]
                                                if l.get('line_no') != _ml_line_no]
                _ml_spec = (matched_ln.get('sku_spec', '') or matched_ln.get('sku', '')).strip().upper()
                if _ml_spec and _ml_spec in new_by_spec:
                    new_by_spec[_ml_spec] = [l for l in new_by_spec[_ml_spec]
                                              if l.get('line_no') != _ml_line_no]

                # 同PO → 检测实际变化（不检查接单日期，接单日期不做修改）
                changes = {}
                new_qty = matched_ln.get('qty', 0)
                new_price = matched_ln.get('price', 0)
                # 出货期：优先取每行独立delivery，兜底用PO头部ship_date
                _ln_delivery = matched_ln.get('delivery', '') or ''
                new_ship = _normalize_date(_ln_delivery) or _normalize_date(pdf_data.get('ship_date', ''))
                new_cpo = matched_ln.get('customer_po', '')
                # 获取表头列映射（由combined_search阶段检测）
                _col_map = rec.get('col_map', {})

                # 数量：优先用col_map检测的qty列，兜底扫描I-L
                if new_qty:
                    _qty_col_detected = _col_map.get('qty', '')
                    _qty_search = [_qty_col_detected] if _qty_col_detected else []
                    _qty_search += [c for c in 'IJKL' if c not in _qty_search]
                    for c in _qty_search:
                        ov = rd.get(c)
                        if ov is not None:
                            try:
                                if int(float(ov)) > 0:
                                    if int(float(ov)) != int(new_qty):
                                        changes[c] = new_qty
                                    break
                            except:
                                continue

                # 卡板货号：qty变化时同步重算外箱（外箱 = 产品件数 ÷ 卡板数）
                if matched_ln.get('is_pallet') and new_qty:
                    _pallet_count = matched_ln.get('pallet_count', 0) or 0
                    if _pallet_count > 0:
                        _new_outer = int(new_qty / _pallet_count)
                        if _new_outer > 0:
                            _ob_col = _col_map.get('outer_box', 'K')
                            _cur_outer = rd.get(_ob_col)
                            if _cur_outer is not None:
                                try:
                                    if int(float(_cur_outer)) != _new_outer:
                                        changes[_ob_col] = _new_outer
                                        logging.info(f"[卡板修改单] 重算外箱: col={_ob_col} {int(float(_cur_outer))}→{_new_outer} (qty={new_qty}÷卡板={_pallet_count})")
                                except:
                                    pass

                # 单价：优先用col_map检测的price列，兜底扫描R-AF
                _price_empty = False
                if new_price:
                    _found_price = False
                    _price_col_detected = _col_map.get('price', '')
                    _price_search = []
                    if _price_col_detected:
                        _price_search.append(_price_col_detected)
                    for ci in range(18, 33):
                        c = chr(64 + ci) if ci <= 26 else 'A' + chr(64 + ci - 26)
                        if c not in _price_search:
                            _price_search.append(c)
                    for c in _price_search:
                        ov = rd.get(c)
                        if ov is not None:
                            try:
                                if 0 < float(ov) < 10000:
                                    _found_price = True
                                    if abs(float(ov) - new_price) > 0.001:
                                        changes[c] = new_price
                                    break
                            except:
                                continue
                    if not _found_price:
                        _price_empty = True

                # 出货日期：优先用表头检测到的列，兜底扫描K-P列找日期
                if new_ship:
                    _ship_col_detected = _col_map.get('ship_date', '')
                    _ship_search_cols = [_ship_col_detected] if _ship_col_detected else []
                    _ship_search_cols += [c for c in 'KLMNOP' if c not in _ship_search_cols]
                    for c in _ship_search_cols:
                        ov = rd.get(c)
                        if ov is not None:
                            old_str = ''
                            if hasattr(ov, 'year'):
                                old_str = ov.strftime('%Y-%m-%d')
                            elif isinstance(ov, str) and ov:
                                old_str = _normalize_date(ov)
                            # 验证结果是否为有效日期格式，防止备注等文本被误判为日期
                            if old_str and re.match(r'^\d{4}-\d{2}-\d{2}$', old_str):
                                if old_str != new_ship:
                                    changes[c] = new_ship
                                break

                # 客户PO：在D-G列中查找匹配，只有确实不同才标记修改
                # 跳过小数值（防止CBM等误判为客PO）
                if new_cpo and not re.match(r'^\d+\.\d+$', new_cpo):
                    cpo_changed = True
                    # 标准化空白/换行后比较，避免仅空格差异导致误判
                    _cpo_clean = re.sub(r'\s+', '', new_cpo)
                    _cpo_col_detected = _col_map.get('customer_po', '')
                    _cpo_check = [_cpo_col_detected] if _cpo_col_detected else []
                    _cpo_check += [c for c in 'DEFG' if c not in _cpo_check]
                    for c in _cpo_check:
                        ov = str(rd.get(c, '') or '').strip()
                        _ov_clean = re.sub(r'\s+', '', ov)
                        if _ov_clean == _cpo_clean:
                            cpo_changed = False
                            break
                    # 只有PDF有明确客PO、且排期所有候选列都不匹配时才标修改
                    if cpo_changed:
                        # 优先用col_map检测的客PO列，兜底E/F
                        _cpo_write = [_cpo_col_detected] if _cpo_col_detected else []
                        _cpo_write += [c for c in 'EF' if c not in _cpo_write]
                        for c in _cpo_write:
                            ov = str(rd.get(c, '') or '').strip()
                            if not ov:
                                continue
                            # 跳过PO-line格式（PO号-行号）：这是SKU列不能覆写
                            if re.match(r'^\d{7,}-\d+$', ov):
                                continue
                            # 跳过包含当前PO号的值（可能是PO列或SKU列），数字边界匹配避免子串误判
                            if po and _po_boundary_match(po, ov):
                                continue
                            if ov != new_cpo:
                                changes[c] = new_cpo
                                break

                # 货号(ITEM#)：PDF的sku_spec与排期货号列比较，不同则标记修改
                # 77785/92123/77869系列是复合货号（如77785-S001-77673-INT / 77869-77772-S001），PDF会截断，跳过不检测
                _new_item = (matched_ln.get('sku_spec', '') or '').strip()
                _item_change_col = ''
                _is_composite = bool(re.match(r'^(77785|92123|77869)', _new_item, re.I))
                if _new_item and not _is_composite:
                    _item_col = _col_map.get('items', '') or _col_map.get('item', '')
                    _sys_col = _col_map.get('system_code', '')
                    _item_search_cols = [_item_col] if _item_col else []
                    _item_search_cols += [c for c in 'FGH' if c not in _item_search_cols and c != _sys_col]
                    for c in _item_search_cols:
                        ov = str(rd.get(c, '') or '').strip()
                        if not ov:
                            continue
                        if not re.search(r'[A-Za-z]', ov):
                            continue
                        if re.match(r'^\d{7,}-\d+$', ov):
                            continue
                        if ov.upper() != _new_item.upper():
                            changes[c] = _new_item
                            _item_change_col = c
                        break

                # 修改单不检测备注变化，留给用户自己检查
                _note_changed = False
                _pdf_note = ''

                # 检测需要填充的空字段（排期为空但PDF有数据）
                _fill_empty = {}
                if _price_empty and new_price:
                    _fill_empty['price'] = new_price
                _from_person = pdf_data.get('from_person', '') or ''
                if _from_person:
                    _fill_empty['from_person'] = _from_person

                # 混装检测：matched_ln是混装组件或其line_no有冲突（多组件共享），跳过修改
                _ml_is_mixed = matched_ln.get('is_mixed_carton', False)
                _ml_poline_key = f"{po}-{matched_ln.get('line_no', '')}" if po and matched_ln.get('line_no') else ''
                if _ml_is_mixed or (_ml_poline_key and _ml_poline_key in _poline_conflicts):
                    actions.append({
                        'type': 'mixed_skip', 'record': rec,
                        'sku': matched_ln.get('sku', ''),
                        'detail': f"混装跳过 {matched_ln.get('sku','')}（同line混装，需手动处理）",
                        'line': matched_ln
                    })
                    continue

                if changes or _fill_empty:
                    # 有字段修改、备注变化或空字段需填充 → modify 类型
                    detail_parts = []
                    old_vals = {}
                    for col, new_val in changes.items():
                        old_val = rd.get(col, '')
                        if hasattr(old_val, 'strftime'):
                            old_val = old_val.strftime('%Y-%m-%d')
                        old_vals[col] = old_val
                        # 货号列用"货号"标签，避免被_col_cn误显示为"客户PO"
                        _label = '货号' if col == _item_change_col else _col_cn(col)
                        detail_parts.append(f"{_label}: {old_val}→{new_val}")
                    if _fill_empty:
                        _fe_names = []
                        if 'price' in _fill_empty: _fe_names.append('单价')
                        if 'from_person' in _fill_empty: _fe_names.append('跟单')
                        detail_parts.append(f"补填空字段({'/'.join(_fe_names)})")
                    actions.append({
                        'type': 'modify', 'record': rec, 'changes': changes,
                        'old_vals': old_vals,
                        'sku': matched_ln.get('sku', ''),
                        'detail': f"修改 {matched_ln.get('sku','')} {', '.join(detail_parts)}",
                        'note_changed': _note_changed,
                        'pdf_note': _pdf_note if _note_changed else '',
                        'fill_empty': _fill_empty,
                        'item_col': _item_change_col
                    })
                else:
                    # 同PO无变化 → 生成"无变化"标记，前端展示但不写入排期
                    actions.append({
                        'type': 'unchanged', 'record': rec,
                        'sku': matched_ln.get('sku', ''),
                        'detail': f"无变化 {matched_ln.get('sku','')} {matched_ln.get('qty',0)}pcs"
                    })
            # 取消检测已废除：跨文件检测不可靠，留给用户自己检查

        # ===== 3. PDF中有但排期中没有的行 → 新增 =====
        _added_keys = set()  # 去重：(sku_spec, line_no)避免真正重复，同SKU不同line允许
        for ln in new_lines:
            line_no = ln.get('line_no', '')
            po_line = f"{po}-{line_no}" if po and line_no else ''
            # 去重1：按line_no/po_line（已在modify/unchanged中匹配过的行）
            already = (line_no and line_no in matched_pdf_line_nos) or (po_line and po_line in matched_polines)
            if already:
                continue
            # 去重2：同一sku_spec+line_no只新增一次（同SKU不同line各自独立入单）
            _spec = (ln.get('sku_spec', '') or ln.get('sku', '')).strip().upper()
            _dedup_key = (_spec, line_no) if line_no else (_spec,)
            if _dedup_key in _added_keys:
                logging.warning(f"[smart_diff去重] 跳过重复新增: {_spec} line={line_no}")
                continue
            _added_keys.add(_dedup_key)
            sched = self.auto_find(ln.get('sku_spec', '') or ln.get('item_code', '') or ln.get('sku', ''),
                                   current_po=po)
            actions.append({
                'type': 'new', 'line': ln, 'schedule': sched,
                'sku': ln.get('sku', ''),
                'detail': f"新增 {ln.get('sku','')} {ln.get('qty',0)}pcs"
            })

        return actions

    # =================== COM 启动 ===================

    @staticmethod
    def _com_app():
        """启动WPS/Excel COM进程"""
        import win32com.client
        import pythoncom
        pythoncom.CoInitialize()
        for pid in ['Ket.Application', 'Et.Application', 'Excel.Application']:
            try:
                app = win32com.client.DispatchEx(pid)
                app.Visible = False
                app.DisplayAlerts = False
                return app
            except:
                continue
        raise RuntimeError("无法启动WPS或Excel，请确认已安装WPS Office")

    @staticmethod
    def _com_quit(app):
        """安全退出COM"""
        if app:
            try:
                app.DisplayAlerts = False
                app.Quit()
            except:
                pass
        try:
            import pythoncom
            pythoncom.CoUninitialize()
        except:
            pass

    # =================== 批量处理 (COM写入) ===================

    @staticmethod
    def get_batch_progress():
        """获取批量处理进度"""
        return dict(_batch_progress)

    def batch_process(self, orders):
        global _batch_progress
        os.makedirs(BATCH_DIR, exist_ok=True)
        os.makedirs(UNDO_DIR, exist_ok=True)
        ExcelHandler._detect_cols_cache = {}  # 新批次清除列检测缓存
        results = []
        failed = []
        _batch_progress = {'running': True, 'current': '分析中...', 'done': 0, 'total': 0, 'details': []}

        # 生成批次ID
        batch_id = datetime.now().strftime('%Y%m%d-%H%M%S')

        # 按排期文件分组（含去重：同一货号+PO+行号不重复写入）
        file_ops = {}
        _seen_new = set()       # (sku_spec, po, line_no) 去重
        _seen_modify = set()    # (file, row, po) 去重
        for order in orders:
            _po = order.get('header', {}).get('po_number', '')
            for act in order.get('actions', []):
                if act['type'] == 'new' and act.get('schedule'):
                    fkey = act['schedule']['file']
                    # 去重：同货号+同PO+同line_no只入一次
                    _ln = act.get('line', {})
                    _dedup_key = (_ln.get('sku_spec', '') or _ln.get('sku', ''),
                                  _po, _ln.get('line_no', ''))
                    if _dedup_key in _seen_new:
                        logging.warning(f"[batch去重] 跳过重复新单: {_dedup_key}")
                        continue
                    _seen_new.add(_dedup_key)
                    if fkey not in file_ops:
                        file_ops[fkey] = {'file': fkey, 'new': [], 'modify': [], 'cancel': []}
                    # 每个action携带自己的header，不共享
                    act['_header'] = order.get('header', {})
                    file_ops[fkey]['new'].append(act)
                elif act['type'] == 'modify':
                    fkey = act['record']['file']
                    _row = act['record'].get('row', 0)
                    _mod_key = (fkey, _row, _po)
                    if _mod_key in _seen_modify:
                        logging.warning(f"[batch去重] 跳过重复修改: {_mod_key}")
                        continue
                    _seen_modify.add(_mod_key)
                    if fkey not in file_ops:
                        file_ops[fkey] = {'file': fkey, 'new': [], 'modify': [], 'cancel': []}
                    act['_header'] = order.get('header', {})
                    file_ops[fkey]['modify'].append(act)
                elif act['type'] == 'cancel':
                    fkey = act['record']['file']
                    if fkey not in file_ops:
                        file_ops[fkey] = {'file': fkey, 'new': [], 'modify': [], 'cancel': []}
                    act['_header'] = order.get('header', {})
                    file_ops[fkey]['cancel'].append(act)

        # 收集每个订单的PO和客户信息，用于按文件构建撤销记录
        order_info = {}
        for order in orders:
            po = order.get('header', {}).get('po_number', '')
            customer = order.get('header', {}).get('customer', '')
            order_info[po] = customer

        # 逐文件处理前：预检文件锁，被占用的直接放入failed跳过
        locked_files = {}   # {fkey: reason}
        locked_ops   = {}   # {fkey: ops_dict} 保留被锁文件的完整操作数据供重试用
        for fkey in file_ops:
            is_locked, user = self._check_lock_file(fkey)
            if is_locked:
                locked_files[fkey] = f'文件正在被{user or "其他人"}编辑'
        for fkey, reason in locked_files.items():
            ops = file_ops.pop(fkey)  # 从待处理列表移除
            locked_ops[fkey] = ops    # 保留完整操作数据供调用方重试
            fname = os.path.basename(fkey)
            # 收集该文件涉及的PO和货号，展示给用户
            _po_skus = []
            for act in ops.get('new', []) + ops.get('modify', []) + ops.get('cancel', []):
                _sku = act.get('sku', '') or (act.get('line', {}) or {}).get('sku', '') or ''
                _rd = (act.get('record', {}) or {}).get('data', {})
                _po = (act.get('line', {}) or {}).get('po_number', '') or ExcelHandler._find_rec_po(_rd) if _rd else ''
                if _sku:
                    _po_skus.append(f"{_sku}(PO {_po})" if _po else _sku)
            _detail = '、'.join(_po_skus[:10])
            if len(_po_skus) > 10:
                _detail += f' 等{len(_po_skus)}项'
            failed.append({'file': fkey, 'fname': fname, 'reason': reason,
                           'local': '', 'z': fkey, 'detail': _detail})
            logging.warning(f"[batch预检] {fname} 被占用，跳过: {reason} | 涉及: {_detail}")

        _batch_progress['total'] = len(file_ops)
        _skipped_skus = set()  # 被去重跳过的SKU，自检时排除
        app = None
        try:
            app = self._com_app()

            for file_idx, (fkey, ops) in enumerate(file_ops.items()):
                fname = os.path.basename(fkey)
                _batch_progress['current'] = fname
                _batch_progress['done'] = file_idx
                local = os.path.join(BATCH_DIR, fname)
                try:
                    # 只备份当前修改的排期文件（不全量备份）
                    undo_fp = os.path.join(UNDO_DIR, f"{batch_id}_{fname}")
                    shutil.copy2(fkey, undo_fp)
                    shutil.copy2(fkey, local)
                    wb = app.Workbooks.Open(os.path.abspath(local))
                    msg_parts = []
                    file_ratio_warnings = []  # 本文件的比例提醒列表
                    file_note_warnings = []   # 本文件的备注变化提醒
                    file_modify_details = []  # 本文件的修改单详情

                    # --- 1) 先取消（从大行号往小删，避免行号偏移）---
                    cancel_ops = sorted(ops['cancel'],
                                        key=lambda x: x['record']['row'], reverse=True)
                    deleted_rows = []  # (sheet_name, row) 元组列表，区分不同sheet
                    for act in cancel_ops:
                        sn = act['record']['sheet']
                        rn = act['record']['row']
                        ws = wb.Sheets(sn)
                        mc = min(ws.UsedRange.Columns.Count + ws.UsedRange.Column, 100)
                        _row_deleted = self._do_cancel_com(wb, ws, rn, mc)
                        if _row_deleted:
                            deleted_rows.append((sn, rn))  # 只有真正删除了才跟踪偏移
                        msg_parts.append(f"取消{act['sku']}")

                    # --- 2) 修改（调整行号，只对同sheet的删除做偏移）---
                    for act in ops['modify']:
                        sn = act['record']['sheet']
                        orig_row = act['record']['row']
                        # 只对同sheet的删除行做偏移调整（不同sheet行号独立）
                        shift = sum(1 for dsn, d in deleted_rows if dsn == sn and d < orig_row)
                        adj_row = orig_row - shift
                        ws = wb.Sheets(sn)
                        mc = min(ws.UsedRange.Columns.Count + ws.UsedRange.Column, 100)
                        self._do_modify_com(ws, adj_row, mc, act['changes'],
                                            note_changed=act.get('note_changed', False),
                                            pdf_note=act.get('pdf_note', ''),
                                            fill_empty=act.get('fill_empty', {}))
                        _mod_tag = f"修改{act['sku']}"
                        if act.get('note_changed'):
                            _mod_tag += "(备注变化)"
                            file_note_warnings.append({
                                'type': 'modify',
                                'sku': act.get('sku', ''),
                                'sheet': sn,
                                'row': adj_row,
                                'pdf_note': act.get('pdf_note', '')
                            })
                        msg_parts.append(_mod_tag)
                        # 收集修改详情供前端展示
                        if act.get('changes') or act.get('fill_empty'):
                            _detail = {
                                'sku': act.get('sku', ''),
                                'sheet': sn,
                                'row': adj_row,
                                'changes': {},
                                'fill_empty': list((act.get('fill_empty') or {}).keys()),
                                'note_changed': act.get('note_changed', False)
                            }
                            # 翻译列字母为中文字段名
                            _col_names = {'H': '单价', 'I': '金额', 'J': '接单期',
                                          'K': '出货期', 'L': '验货期', 'M': '出货期',
                                          'N': '出货期', 'O': '出货期', 'P': '出货期',
                                          'E': '客PO', 'F': '客PO'}
                            _act_item_col = act.get('item_col', '')
                            for _ck, _cv in (act.get('changes') or {}).items():
                                # 货号列优先用"货号"标签，不被客PO覆盖
                                _cn = '货号' if (_act_item_col and _ck == _act_item_col) else _col_names.get(_ck, f'列{_ck}')
                                _detail['changes'][_cn] = str(_cv)
                            file_modify_details.append(_detail)

                    # --- 3) 新增（动态查找插入位置，按sheet独立跟踪偏移）---
                    # 按出货期升序排列，确保start_after机制下日期顺序正确
                    def _new_sort_key(a):
                        d = a.get('line', {}).get('delivery', '') or a.get('line', {}).get('ship_date', '')
                        return _normalize_date(d) or '9999-99-99'
                    ops['new'] = sorted(ops['new'], key=_new_sort_key)
                    inserted_positions = {}  # {sheet_name: [pos_list]}，按sheet独立跟踪
                    last_insert_pos = {}     # {sheet_name: last_pos}，按sheet独立跟踪
                    _batch_written_keys = set()  # PO-line_no去重键集合
                    for act in ops['new']:
                        if not act.get('schedule'):
                            continue
                        sn = act['schedule']['sheet']
                        ref = act['schedule']['ref']
                        mc = min(act['schedule'].get('mcol', 100), 100)
                        # === 重复入单检查：基于PO+line_no去重 ===
                        # smart_diff的_seen_new已用(sku_spec, po, line_no)做批内去重
                        # 此处仅做最终安全检查：同PO同line_no才算重复
                        _act_hdr = act.get('_header', {})
                        _dup_po = str(_act_hdr.get('po_number', '') or act['line'].get('po', '')).strip()
                        _dup_line = str(act['line'].get('line_no', '') or '').strip()
                        _dup_key = f"{_dup_po}-{_dup_line}" if _dup_po and _dup_line else ''
                        if _dup_key and _dup_key in _batch_written_keys:
                            logging.info(f"[去重] 跳过批内重复: {_dup_key} sheet={sn}")
                            msg_parts.append(f"跳过(已存在){act['sku']}")
                            _skipped_skus.add(act.get('sku', ''))
                            continue
                        if _dup_key:
                            _batch_written_keys.add(_dup_key)
                        # 只对同sheet的删除行做偏移调整
                        shift_del = sum(1 for dsn, d in deleted_rows if dsn == sn and d < ref)
                        adj_ref = ref - shift_del
                        # 只对同sheet的已插入行做偏移调整
                        for p in inserted_positions.get(sn, []):
                            if p <= adj_ref:
                                adj_ref += 1
                        ws = wb.Sheets(sn)
                        _ptm = act.get('schedule', {}).get('po_type_match', True) if act.get('schedule') else True
                        pos, w, ratio_w = self._do_new_com(ws, adj_ref, mc, _act_hdr, act['line'],
                                               start_after=last_insert_pos.get(sn, 0),
                                               po_type_match=_ptm)
                        inserted_positions.setdefault(sn, []).append(pos)
                        last_insert_pos[sn] = pos
                        warn_tag = ''
                        if w:
                            warn_tag = f" [空字段: {', '.join(w)}]"
                        if ratio_w:
                            _po_num = act['line'].get('po', '') or _act_hdr.get('po', '')
                            file_ratio_warnings.append({
                                'po': _po_num,
                                'sku': act.get('sku', ''),
                                'pattern': ratio_w,
                                'file': fname,
                                'sheet': sn,
                                'row': pos
                            })
                        msg_parts.append(f"新增{act['sku']}{warn_tag}")
                        # 收集新单备注信息（供前端提醒"备注已写入"）
                        _hdr = _act_hdr
                        _has_note = bool((_hdr.get('tracking_code') or '') or
                                         (_hdr.get('packaging_info') or '') or
                                         (_hdr.get('remark') or ''))
                        if _has_note:
                            file_note_warnings.append({
                                'type': 'new',
                                'sku': act.get('sku', ''),
                                'sheet': sn,
                                'row': pos
                            })

                    wb.Save()
                    wb.Close(False)

                    # 尝试保存到Z盘
                    z_ok = False
                    z_err = ''
                    try:
                        self._try_save_z(local, fkey)
                        z_ok = True
                    except Exception as e:
                        z_err = str(e)
                        logging.warning(f"[batch] {fname} 保存Z盘失败: {z_err}")

                    r = {'file': fname, 'local': local, 'z': fkey, 'z_saved': z_ok,
                         'msg': ' | '.join(msg_parts),
                         'ratio_warnings': file_ratio_warnings,
                         'note_warnings': file_note_warnings,
                         'modify_details': file_modify_details,
                         'counts': {'new': len(ops['new']), 'modify': len(ops['modify']),
                                    'cancel': len(ops['cancel'])}}
                    if z_ok:
                        results.append(r)
                        # 每个排期文件单独保存一条撤销记录
                        type_names = {'new': '新增', 'modify': '修改', 'cancel': '取消'}
                        file_ops_list = []
                        for t in ('new', 'modify', 'cancel'):
                            for act in ops[t]:
                                op = {'type': t, 'sku': act.get('sku', ''),
                                      'detail': act.get('detail', '')}
                                if t == 'new' and act.get('line'):
                                    op['qty'] = act['line'].get('qty', 0)
                                file_ops_list.append(op)
                        labels = []
                        for op in file_ops_list:
                            tn = type_names.get(op['type'], op['type'])
                            labels.append(f"{tn} {op['sku']}" +
                                          (f" {op.get('qty','')}pcs" if op.get('qty') else ''))
                        self._save_undo_entry({
                            'id': f"{batch_id}_{fname}",
                            'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                            'operations': file_ops_list,
                            'files': [{'name': fname, 'backup': undo_fp, 'z_path': fkey}],
                            'label': f"[{fname}] " + (' | '.join(labels[:3]) +
                                     (f' 等{len(labels)}项' if len(labels) > 3 else ''))
                        })
                    else:
                        r['reason'] = self._classify_save_error(z_err)
                        failed.append(r)

                except Exception as e:
                    logging.error(f"[batch] 处理 {fname} 异常: {e}")
                    # 确保workbook关闭
                    try:
                        wb.Close(False)
                    except:
                        pass
                    reason = self._classify_save_error(str(e))
                    failed.append({'file': fname, 'local': local, 'z': fkey,
                                   'z_saved': False, 'reason': reason,
                                   'msg': f'处理失败: {e}',
                                   'counts': {'new': len(ops['new']),
                                              'modify': len(ops['modify']),
                                              'cancel': len(ops['cancel'])}})
        finally:
            self._com_quit(app)
            _batch_progress = {'running': False, 'current': '完成', 'done': len(file_ops),
                               'total': len(file_ops), 'details': []}

        # 入单后自检已移除 — Step 10已内置空字段检查，无需重复读取Z盘验证

        return {'results': results, 'failed': failed, 'locked_ops': locked_ops}

    # =================== 入单后自检 ===================

    @staticmethod
    def _detect_cols_openpyxl(ws):
        """openpyxl版列检测（只检测自检需要的关键列）"""
        cols = {}
        for r in range(1, 11):
            for c in range(1, 40):
                try:
                    v = ws.cell(r, c).value
                    if not v:
                        continue
                    vl = _t2s(str(v).strip())
                    vlu = vl.upper().replace(' ', '')
                    if ('PO' in vlu and ('号' in vl or '#' in vl or vl == 'PO')
                            and '客' not in vl and '小' not in vl and 'po_number' not in cols):
                        cols['po_number'] = c
                    elif (('ITEM' in vlu and ('#' in vl or vlu.endswith('ITEM') or vlu in ('ITEMS', 'ITEM', 'ITEMCODE', 'ITEM#')))
                          or vl == '货号') and 'items' not in cols:
                        cols['items'] = c
                    elif 'SKU' in vlu and 'sku' not in cols:
                        cols['sku'] = c
                    elif (('数量' in vl and '合计' not in vl and '箱' not in vl)
                          or vlu in ('QTY', 'QTYPCS', 'PO数量')) and 'qty' not in cols:
                        cols['qty'] = c
                    elif '单价' in vl and 'price' not in cols:
                        cols['price'] = c
                    elif ('出货' in vl or '出期' in vl or vlu == 'CRD') and '情况' not in vl and 'ship_date' not in cols:
                        cols['ship_date'] = c
                    elif ('客户名' in vl or '第三方' in vl or vl == '客户') and 'PO' not in vlu and 'customer' not in cols:
                        cols['customer'] = c
                    elif ('中文' in vl or '品名' in vl or vl == '名称') and 'product_name' not in cols:
                        cols['product_name'] = c
                except Exception:
                    pass
        return cols

    def _post_validate(self, z_path, ops, skipped_skus=None):
        """入单后自检：用openpyxl只读打开Z盘文件，验证写入数据与PDF一致
        检查项：新增行是否存在、数量是否正确；修改行是否存在"""
        skipped = skipped_skus or set()
        issues = []
        try:
            wb = openpyxl.load_workbook(z_path, read_only=True, data_only=True)
        except Exception as e:
            return [{'level': 'error', 'msg': f'无法打开验证: {e}'}]

        # 缓存每个sheet的列检测和行数据
        _sheet_cache = {}  # {sheet_name: {'cols': {}, 'rows': [(po, items_set, row_data)]}}

        def _get_sheet_data(sn):
            if sn in _sheet_cache:
                return _sheet_cache[sn]
            if sn not in wb.sheetnames:
                _sheet_cache[sn] = None
                return None
            ws = wb[sn]
            cols = ExcelHandler._detect_cols_openpyxl(ws)
            po_c = cols.get('po_number')
            item_c = cols.get('items')
            sku_c = cols.get('sku')
            qty_c = cols.get('qty')
            price_c = cols.get('price')
            ship_c = cols.get('ship_date')
            cust_c = cols.get('customer')
            pname_c = cols.get('product_name')
            rows = []
            try:
                for row in ws.iter_rows(min_row=3, max_row=min(ws.max_row or 500, 500)):
                    po_val = ''
                    if po_c and po_c <= len(row):
                        po_val = str(row[po_c - 1].value or '').strip()
                    # 收集item列的值
                    items_set = set()
                    for ci in (item_c, sku_c):
                        if ci and ci <= len(row):
                            cv = str(row[ci - 1].value or '').strip()
                            ic = _item_code(cv)
                            if ic:
                                items_set.add(ic)
                    # 也检查前后几列（F/G/H常规位置）
                    for ci in range(6, 9):
                        if ci <= len(row):
                            cv = str(row[ci - 1].value or '').strip()
                            ic = _item_code(cv)
                            if ic:
                                items_set.add(ic)
                    rd = {}
                    if qty_c and qty_c <= len(row):
                        rd['qty'] = row[qty_c - 1].value
                    if price_c and price_c <= len(row):
                        rd['price'] = row[price_c - 1].value
                    if ship_c and ship_c <= len(row):
                        rd['ship_date'] = row[ship_c - 1].value
                    if cust_c and cust_c <= len(row):
                        rd['customer'] = row[cust_c - 1].value
                    if pname_c and pname_c <= len(row):
                        rd['product_name'] = row[pname_c - 1].value
                    rows.append((po_val, items_set, rd))
            except (ValueError, StopIteration):
                pass  # WPS AutoFilter兼容
            except Exception as _e:
                logging.warning(f"[自检] 读取sheet {sn} 异常: {_e}")
            data = {'cols': cols, 'rows': rows}
            _sheet_cache[sn] = data
            return data

        try:
            # --- 验证新增 ---
            for act in ops.get('new', []):
                if not act.get('schedule'):
                    continue
                sku = act.get('sku', '')
                if sku in skipped:
                    continue
                sn = act['schedule']['sheet']
                hdr = act.get('_header', {})
                ln = act.get('line', {})
                po = str(hdr.get('po_number', '')).strip()
                item = _item_code(ln.get('sku_spec', '') or ln.get('sku', ''))
                if not item:
                    continue

                sd = _get_sheet_data(sn)
                if sd is None:
                    issues.append({'level': 'error', 'type': 'new', 'sku': sku,
                                   'msg': f'工作表"{sn}"不存在'})
                    continue

                # 搜索PO+item
                found_rd = None
                for r_po, r_items, r_data in sd['rows']:
                    if item not in r_items:
                        continue
                    if po and not _po_boundary_match(po, r_po):
                        continue
                    found_rd = r_data
                    break

                if not found_rd:
                    issues.append({'level': 'error', 'type': 'new', 'sku': sku,
                                   'msg': f'未找到: {sn} PO={po} 货号={item}'})
                    continue

                # 验证数量
                expected_qty = ln.get('qty', 0)
                if expected_qty and found_rd.get('qty') is not None:
                    try:
                        aq = float(found_rd['qty'])
                        eq = float(expected_qty)
                        if abs(aq - eq) > 0.5:
                            issues.append({'level': 'error', 'type': 'new', 'sku': sku,
                                           'msg': f'数量不匹配: 期望{int(eq)} 实际{int(aq)}'})
                    except (ValueError, TypeError):
                        pass

                # 验证中文名不为空
                if not found_rd.get('product_name'):
                    issues.append({'level': 'warning', 'type': 'new', 'sku': sku,
                                   'msg': f'中文名为空'})

            # --- 验证修改 ---
            for act in ops.get('modify', []):
                sku = act.get('sku', '')
                sn = act['record']['sheet']
                hdr = act.get('_header', {})
                po = str(hdr.get('po_number', '')).strip() if hdr else ''
                item = _item_code(sku)
                if not item:
                    continue

                sd = _get_sheet_data(sn)
                if sd is None:
                    issues.append({'level': 'error', 'type': 'modify', 'sku': sku,
                                   'msg': f'工作表"{sn}"不存在'})
                    continue

                found = False
                for r_po, r_items, r_data in sd['rows']:
                    if item in r_items:
                        if po and r_po:
                            if _po_boundary_match(po, r_po):
                                found = True
                                break
                        else:
                            found = True
                            break
                if not found:
                    issues.append({'level': 'error', 'type': 'modify', 'sku': sku,
                                   'msg': f'修改后未找到: {sn} PO={po} 货号={item}'})

            # --- 验证取消（检查是否已从生产sheet移除）---
            for act in ops.get('cancel', []):
                sku = act.get('sku', '')
                sn = act['record']['sheet']
                item = _item_code(sku)
                if not item:
                    continue
                sd = _get_sheet_data(sn)
                if sd is None:
                    continue  # sheet不存在=已删除，OK
                # 检查item是否仍在该sheet中（取消应该删除或标记）
                # 注意：只标记红字的情况下item仍会被找到，这是正常的
                # 所以取消验证不做严格检查

        except Exception as e:
            issues.append({'level': 'error', 'msg': f'自检过程异常: {e}'})
        finally:
            try:
                wb.close()
            except Exception:
                pass

        return issues

    def _save_undo_entry(self, entry):
        """保存撤销历史条目"""
        os.makedirs(DATA_DIR, exist_ok=True)
        history = []
        if os.path.exists(UNDO_HISTORY):
            try:
                with open(UNDO_HISTORY, 'r', encoding='utf-8') as f:
                    history = json.load(f)
                if not isinstance(history, list):
                    history = []
            except:
                history = []
        history.append(entry)
        # 只保留最近30条
        history = history[-30:]
        with open(UNDO_HISTORY, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=1)

    # =================== COM内部操作 ===================

    _detect_cols_cache = {}  # 缓存key = (file_path, sheet_name)

    def _detect_cols(self, ws, max_col):
        """从表头行自动检测所有关键列位置，适配不同排期文件布局
        不同排期文件列顺序不同（如15760有额外的系统货号/ITEM#列），
        必须通过表头关键词检测，不能硬编码列号
        支持繁体中文表头（如出貨→出货、驗貨→验货等）"""
        # 同Sheet缓存：同一Sheet多行操作只检测一次
        try:
            _ck = (ws.Parent.FullName, ws.Name)
            if _ck in ExcelHandler._detect_cols_cache:
                return ExcelHandler._detect_cols_cache[_ck]
        except:
            _ck = None
        cols = {}
        mc = min(max(max_col, 50), 100)  # 至少扫50列，防止openpyxl的max_column=30截断
        # 先识别表头行（含>=3个典型表头关键词的行），避免数据行的内容被误匹配
        _hdr_kws = ('接单', '客户', 'PO', '货号', 'ITEM', '数量', 'QTY', '国家', '走货',
                     '出货', '出期', '验货', '备注', '单价', '金额', '箱', 'SKU', '名称', '品名', 'CRD')
        _header_rows = set()
        for r in range(1, 11):
            _kw_cnt = 0
            for c in range(1, min(mc + 1, 40)):
                try:
                    v = ws.Cells(r, c).Value
                    if v:
                        _vs = _t2s(str(v).strip()).upper()
                        if len(_vs) <= 30 and any(kw in _vs for kw in _hdr_kws):
                            _kw_cnt += 1
                except:
                    pass
            if _kw_cnt >= 3:
                _header_rows.add(r)
        if not _header_rows:
            _header_rows = set(range(1, 6))  # 兜底：扫描所有行
        # 只扫描表头行（避免数据行中"7/25 出货"等值被误匹配为列头）
        for r in sorted(_header_rows):
            for c in range(1, mc + 1):
                try:
                    v = ws.Cells(r, c).Value
                    if not v:
                        continue
                    vl_raw = str(v).strip()
                    # 跳过长文本（>50字符几乎肯定是数据而非表头，避免备注等列的数据被误匹配）
                    if len(vl_raw) > 50:
                        continue
                    vl = _t2s(vl_raw)  # 繁体→简体
                    vlu = vl.upper().replace(' ', '')

                    # 接单日期（含"首办"变体）
                    if ('接单' in vl or '首办' in vl) and 'po_date' not in cols:
                        cols['po_date'] = c
                    # 客户名（排除客户PO，支持"第二方"/"第三方"变体）
                    elif ('客户名' in vl or '第三方' in vl or '第二方' in vl) and 'PO' not in vlu and 'customer' not in cols:
                        cols['customer'] = c
                    elif vl == '客户' and 'customer' not in cols:
                        cols['customer'] = c
                    # 走货国（排除走货日期/走货期/走货情况等非国家列）
                    elif (('走货' in vl and '日' not in vl and '期' not in vl and '情况' not in vl)
                          or vl in ('国家', '目的国')) and 'destination' not in cols:
                        cols['destination'] = c
                    # PO号（排除客户PO/小PO）
                    elif 'PO' in vlu and ('号' in vl or '#' in vl or vl == 'PO') and '客户' not in vl and '客' not in vl.split('PO')[0][-1:] and '小' not in vl and '数量' not in vl and 'po_number' not in cols:
                        cols['po_number'] = c
                    # 客户PO / 小PO / 客PO（含"第三方客PO"变体）
                    # 注意："客PO期"是出货日期列，不是客户PO号列
                    elif (('客户' in vl and 'PO' in vlu) or ('小' in vl and 'PO' in vlu) or
                          ('客PO' in vl_raw or '客PO' in vl) or
                          vl in ('小PO号', '小PO', '客PO号', '客PO',
                                 '第三方客户PO NO#', '第三方客PO NO#')) and '期' not in vl and 'customer_po' not in cols:
                        cols['customer_po'] = c
                    # SKU（精确匹配及变体，含"SKU号"）
                    elif ('SKU' in vlu and vlu not in ('SKUCODE',) and
                          'sku' not in cols and 'ITEM' not in vlu):
                        cols['sku'] = c
                    # 系统货号（标记但不写入）
                    elif ('系统' in vl or vlu in ('SYSTEMCODE', 'SYSTEMNO')) and 'system_code' not in cols:
                        cols['system_code'] = c
                    # ITEM#/ITEMS/货号
                    elif 'items' not in cols and (
                        ('ITEM' in vlu and ('#' in vl or vlu.endswith('ITEM'))) or
                        vlu in ('ITEMS', 'ITEM', 'ITEMCODE', 'ITEM#')
                    ):
                        cols['items'] = c
                    elif vl == '货号' and 'items' not in cols:
                        cols['items'] = c
                    # 中文名/品名/产品名/名称
                    elif 'product_name' not in cols and (
                        '中文' in vl or '品名' in vl or vl == '名称' or
                        ('产品' in vl and ('名' in vl or '描述' in vl))
                    ):
                        cols['product_name'] = c
                    # PO数量/数量
                    elif 'qty' not in cols and (
                        ('数量' in vl and '合计' not in vl and '计划' not in vl and
                         '箱' not in vl and '外' not in vl) or
                        vlu in ('QTY', 'QTYPCS', 'PO数量')
                    ):
                        cols['qty'] = c
                    # 内箱
                    elif '内箱' in vl and 'inner_box' not in cols:
                        cols['inner_box'] = c
                    # 外箱（排除"外箱贴纸"等非数量列）
                    elif ('外箱' in vl or ('装箱' in vl and '内箱' not in vl)) and '贴纸' not in vl and 'outer_box' not in cols:
                        cols['outer_box'] = c
                    # 总箱/箱数
                    elif ('总箱' in vl or vl == '箱数') and 'total_box' not in cols:
                        cols['total_box'] = c
                    # 卡板/柜
                    elif '卡板' in vl and 'pallets' not in cols:
                        cols['pallets'] = c
                    # 出货期/出货日期/出期/走货日期/CRD（排除"计算走货""走货情况""走货国"）
                    elif ('出货' in vl or '出期' in vl or vlu == 'CRD' or
                          ('走货' in vl and '计算' not in vl and '情况' not in vl and '国' not in vl)) and 'ship_date' not in cols:
                        cols['ship_date'] = c
                    # 客PO期（部分排期有独立的客PO期列，填同样的出货日期）
                    elif ('客PO' in vl and '期' in vl) and 'cpo_date' not in cols:
                        cols['cpo_date'] = c
                    # 验货期/验货日期/计划验货期
                    elif '验货' in vl and 'inspection' not in cols:
                        cols['inspection'] = c
                    # 业务/跟单
                    elif ('业务' in vl or '跟单' in vl) and 'from_person' not in cols:
                        cols['from_person'] = c
                    # 单价
                    elif '单价' in vl and 'price' not in cols:
                        cols['price'] = c
                    # 金额
                    elif '金额' in vl and 'total_usd' not in cols:
                        cols['total_usd'] = c
                    # 备注
                    elif vl in ('备注', '备注专栏', 'Remark', 'REMARK') and 'remark' not in cols:
                        cols['remark'] = c
                    # 条码
                    elif ('条码' in vl or 'BARCODE' in vlu or 'UPC' in vlu or 'EAN' in vlu) and 'barcode' not in cols:
                        cols['barcode'] = c
                except:
                    pass
        # 兜底：如果ship_date未检测到但cpo_date存在，用cpo_date作为ship_date
        if 'ship_date' not in cols and 'cpo_date' in cols:
            cols['ship_date'] = cols['cpo_date']
        _sn = ''
        try:
            _sn = ws.Name or ''
        except:
            pass
        if 'ship_date' not in cols:
            logging.warning(f"[_detect_cols] {_sn}: ship_date未检测到! 表头行={list(_header_rows)}, 检测到列={list(cols.keys())}")
        else:
            logging.debug(f"[_detect_cols] {_sn}: 检测到列: {cols}")
        if _ck:
            ExcelHandler._detect_cols_cache[_ck] = cols
        return cols

    # ===== 11962双产品专用写入器 =====

    @staticmethod
    def _scan_11962_cols(ws):
        """扫描11962排期第1行表头，返回专用列位置字典（支持双产品结构）"""
        cols = {}
        _qty_seen = _price_seen = _usd_seen = _item_seen = _name_seen = _inner_seen = _outer_seen = _tbox_seen = False
        for c in range(1, 50):
            try:
                v = str(ws.Cells(1, c).Value or '').strip()
            except:
                continue
            if not v:
                continue
            if '接单期' in v:
                cols.setdefault('po_date', c)
            elif '客PO期' in v or ('客' in v and 'PO' in v.upper() and '期' in v):
                cols.setdefault('ship_date', c)
            elif '计划验货期' in v:
                cols.setdefault('inspection', c)
            elif 'PO号' == v or v == 'PO号':
                cols.setdefault('po_number', c)
            elif '跟单' in v:
                cols.setdefault('from_person', c)
            elif v == 'SKU':
                cols.setdefault('sku', c)
            elif '备注' in v:
                cols.setdefault('remark', c)
            # 双产品列：按出现顺序区分第1组(公仔)/第2组(浴巾)
            elif v == '货号':
                if not _item_seen:
                    cols['item1'] = c; _item_seen = True
                else:
                    cols.setdefault('item2', c)
            elif v == '名称':
                if not _name_seen:
                    cols['name1'] = c; _name_seen = True
                else:
                    cols.setdefault('name2', c)
            elif '内箱' in v and '数量' in v:
                if not _inner_seen:
                    cols['inner1'] = c; _inner_seen = True
                else:
                    cols.setdefault('inner2', c)
            elif '外箱' in v and '数量' in v:
                if not _outer_seen:
                    cols['outer1'] = c; _outer_seen = True
                else:
                    cols.setdefault('outer2', c)
            elif '总箱' in v:
                if not _tbox_seen:
                    cols['tbox1'] = c; _tbox_seen = True
                else:
                    cols.setdefault('tbox2', c)
            elif '数量' in v and '箱' not in v and '合计' not in v:
                if not _qty_seen:
                    cols['qty1'] = c; _qty_seen = True
                else:
                    cols.setdefault('qty2', c)
            elif '公仔' in v and '单价' in v:
                cols.setdefault('price1', c)
            elif '公仔' in v and '金额' in v:
                cols.setdefault('usd1', c)
            elif '浴巾' in v and '单价' in v:
                cols.setdefault('price2', c)
            elif '浴巾' in v and '金额' in v:
                cols.setdefault('usd2', c)
        return cols

    def _do_new_11962_com(self, ws, ref_row, max_col, header, ln, start_after=0):
        """11962猫头鹰专用新录入：每行同时写入公仔(第1组)和浴巾(第2组)"""
        from datetime import timedelta
        BLUE_COM = 15773696

        ship_str = ln.get('delivery', '') or header.get('ship_date', '')
        ship_dt = _parse_date(ship_str)

        # 检测11962专用列位置
        c = self._scan_11962_cols(ws)
        ship_col = c.get('ship_date', 19)
        mc = min(max(max_col, 50), 100)

        # 确定插入位置
        pos = self._insert_pos_com(ws, ship_dt, col=ship_col, start_after=start_after)
        actual_ref = ref_row + 1 if ref_row >= pos else ref_row
        logging.info(f"[11962新录入] pos={pos} actual_ref={actual_ref} PO={header.get('po_number')}")

        # 复制参考行格式
        try:
            ws.Range(ws.Cells(actual_ref, 1), ws.Cells(actual_ref, mc)).Copy()
            ws.Range(ws.Cells(pos, 1), ws.Cells(pos, mc)).PasteSpecial(Paste=-4122)
        except:
            pass
        try:
            ws.Application.CutCopyMode = False
        except:
            pass

        # 复制参考行值/公式（到最后已知数据列）
        end_col = max(c.get('usd2', 30), 30)
        for col in range(1, end_col + 1):
            try:
                rc = ws.Cells(actual_ref, col)
                if rc.HasFormula:
                    ws.Cells(pos, col).FormulaR1C1 = rc.FormulaR1C1
                elif rc.Value is not None:
                    ws.Cells(pos, col).Value = rc.Value
            except:
                pass

        # 蓝色填充
        try:
            ws.Range(ws.Cells(pos, 1), ws.Cells(pos, end_col)).Interior.Color = BLUE_COM
            ws.Range(ws.Cells(pos, 1), ws.Cells(pos, end_col)).Font.Color = 0
        except:
            pass

        # 写入订单特定字段
        po = header.get('po_number', '')
        po_dt = _parse_date(header.get('po_date', ''))

        if po_dt and c.get('po_date'):
            ws.Cells(pos, c['po_date']).Value = _date_serial(po_dt)
            try:
                ws.Cells(pos, c['po_date']).NumberFormat = 'yyyy/m/daaa'
            except:
                pass
        if po and c.get('po_number'):
            try:
                ws.Cells(pos, c['po_number']).ClearContents()
            except:
                pass
            _sv_com(ws, pos, c['po_number'], po)
        if ship_dt and ship_col:
            ws.Cells(pos, ship_col).Value = _date_serial(ship_dt)
        if ship_dt and c.get('inspection'):
            insp_dt = ship_dt - timedelta(days=4)
            if insp_dt.weekday() == 6:
                insp_dt -= timedelta(days=1)
            ws.Cells(pos, c['inspection']).Value = _date_serial(insp_dt)

        # 公仔数量/单价/金额
        qty1 = ln.get('qty', 0) or 0
        if qty1 and c.get('qty1'):
            _sv_com(ws, pos, c['qty1'], int(qty1))
        if ln.get('price') and c.get('price1'):
            _sv_com(ws, pos, c['price1'], ln['price'])
        if ln.get('total_usd') and c.get('usd1'):
            _sv_com(ws, pos, c['usd1'], ln['total_usd'])

        # 浴巾数量/单价/金额
        qty2 = ln.get('qty2', 0) or 0
        if qty2 and c.get('qty2'):
            _sv_com(ws, pos, c['qty2'], int(qty2))
        if ln.get('price2') and c.get('price2'):
            _sv_com(ws, pos, c['price2'], ln['price2'])
        if ln.get('total_usd2') and c.get('usd2'):
            _sv_com(ws, pos, c['usd2'], ln['total_usd2'])

        # 跟单
        fp = header.get('from_person', '')
        if fp and c.get('from_person'):
            _sv_com(ws, pos, c['from_person'], fp)

        # 货号（物料号）：公仔→item1列(C7)，浴巾→item2列(C13)
        mat_no1 = ln.get('mat_no1', '')
        mat_no2 = ln.get('mat_no2', '')
        if mat_no1 and c.get('item1'):
            try:
                ws.Cells(pos, c['item1']).ClearContents()
            except:
                pass
            _sv_com(ws, pos, c['item1'], mat_no1)
        if mat_no2 and c.get('item2'):
            try:
                ws.Cells(pos, c['item2']).ClearContents()
            except:
                pass
            _sv_com(ws, pos, c['item2'], mat_no2)

        # 备注：始终写入（覆盖参考行继承内容）；_parse_mpo_rfq保证最少为"Remark:"
        remark = ln.get('remark', '')
        if remark and c.get('remark'):
            try:
                ws.Cells(pos, c['remark']).ClearContents()
            except:
                pass
            _sv_com(ws, pos, c['remark'], remark)

        # 清除不应继承参考行的列（小PO号）
        if c.get('customer_po') or c.get('sku'):
            pass  # 11962无小PO号、SKU从参考行继承即可

        logging.info(f"[11962新录入] row={pos} PO={po} 公仔qty={qty1} 浴巾qty={qty2} ship={ship_str} mat1={mat_no1} mat2={mat_no2}")
        return pos, [], None

    # ===== 通用新录入（主入口） =====

    def _do_new_com(self, ws, ref_row, max_col, header, ln, start_after=0, po_type_match=True):
        """通过COM插入新行 — 全部列位置通过表头自动检测，适配不同排期文件布局
        流程：检测列位置 → 按出货期找插入位 → 插入空行 → 复制参考行全部内容
             → 仅覆盖订单特定字段 → 清除系统列
        start_after: 传给_insert_pos_com，同批次多条目时保持插入顺序
        po_type_match: 参考行PO类型是否匹配，False时清空中文名、按单价判定金额公式"""
        # 11962双产品格式路由
        if ln.get('is_11962_dual'):
            return self._do_new_11962_com(ws, ref_row, max_col, header, ln, start_after)

        # 出货日期：优先行级delivery（每行可能不同），兜底用header的ship_date
        ship_str = ln.get('delivery', '') or header.get('ship_date', '')
        ship_dt = _parse_date(ship_str)
        mc = min(max(max_col, 50), 100)  # 至少50列，防止mcol=30截断导致单价/金额列丢失

        # 卡板货号标记（供后续外箱反推和金额公式使用）
        _is_pallet = ln.get('is_pallet', False)
        _pallet_count = ln.get('pallet_count', 0) or 0

        # 1. 先检测所有列位置（在插入行之前，表头不受影响）
        dcols = self._detect_cols(ws, mc)

        # 2. 用检测到的出货期列查找插入位置（不同文件出货期列不同）
        ship_col = dcols.get('ship_date', 13)
        pos = self._insert_pos_com(ws, ship_dt, col=ship_col, start_after=start_after)

        # 3. 插入空行
        ws.Rows(pos).Insert()

        # 4. 插入后ref_row可能移位
        actual_ref = ref_row + 1 if ref_row >= pos else ref_row

        # 5. 只复制参考行的格式（不复制值和公式，避免公式引用错乱影响其他行）
        # 使用范围复制（仅mc列）而非整行复制，避免16384列超大sheet导致COM崩溃
        try:
            src_rng = ws.Range(ws.Cells(actual_ref, 1), ws.Cells(actual_ref, mc))
            src_rng.Copy()
            dst_rng = ws.Range(ws.Cells(pos, 1), ws.Cells(pos, mc))
            dst_rng.PasteSpecial(Paste=-4122)  # xlPasteFormats only
        except Exception as _fmt_err:
            logging.warning(f"[新录入] 范围格式复制失败，尝试整行: {_fmt_err}")
            try:
                ws.Rows(actual_ref).Copy()
                ws.Rows(pos).PasteSpecial(Paste=-4122)
            except Exception as _row_err:
                logging.warning(f"[新录入] 整行格式复制也失败: {_row_err}")

        # 6. 清除剪贴板
        try:
            ws.Application.CutCopyMode = False
        except:
            pass

        # 7. 逐列复制值和公式（公式用FormulaR1C1自动调整相对引用，不影响其他行）
        # 截止列：金额列之后的分配列不复制（用户手动处理）
        # 若金额列未检测到，以所有已知标准列中最靠右的为准，防止复制分配列循环引用公式
        _total_usd_c = dcols.get('total_usd', 0)
        _max_known_col = max((v for v in dcols.values() if isinstance(v, int) and v > 0), default=0)
        if _total_usd_c:
            _copy_end = _total_usd_c
        elif _max_known_col:
            _copy_end = _max_known_col  # 以最右已知列为截止，不复制分配列
        else:
            _copy_end = mc
        for c in range(1, _copy_end + 1):
            try:
                ref_cell = ws.Cells(actual_ref, c)
                if ref_cell.HasFormula:
                    # R1C1格式的公式自动适配新行位置
                    ws.Cells(pos, c).FormulaR1C1 = ref_cell.FormulaR1C1
                else:
                    v = ref_cell.Value
                    if v is not None:
                        ws.Cells(pos, c).Value = v
            except:
                pass

        # 7.2 清除金额列之后的空单元格（xlPasteFormats把数字格式带过来，WPS对空数字格式单元格显示0，
        # 需要显式ClearContents避免出现"幽灵0"干扰分配列的公式）
        if _total_usd_c:
            for c in range(_total_usd_c + 1, mc + 1):
                try:
                    cell = ws.Cells(pos, c)
                    if not cell.HasFormula:
                        cell.ClearContents()
                except:
                    pass

        # 7.5 公式修复：直接按已知列位置构建相对公式，彻底避免绝对引用问题
        # 普通货号 total_box = 数量 ÷ 外箱
        if not _is_pallet:
            _tb_c = dcols.get('total_box')
            _qty_c = dcols.get('qty')
            _ob_c = dcols.get('outer_box')
            if _tb_c and _qty_c and _ob_c:
                try:
                    if not ws.Cells(pos, _tb_c).HasFormula:
                        _fml = f"=RC[{_qty_c - _tb_c}]/RC[{_ob_c - _tb_c}]"
                        ws.Cells(pos, _tb_c).FormulaR1C1 = _fml
                        logging.info(f"[7.5总箱] row={pos} {_fml}")
                except Exception as _e:
                    logging.warning(f"[7.5总箱] row={pos} 失败: {_e}")

        # 普通货号 total_usd = 数量 × 单价（卡板由7.5B处理，混装/M开头由8A处理）
        _is_mixed = ln.get('is_mixed_carton', False)
        _sku_upper = str(ln.get('sku_spec', '') or ln.get('sku', '')).strip().upper()
        _is_m_prefix = _sku_upper.startswith('M')
        if not _is_pallet and not _is_mixed and not _is_m_prefix:
            _usd_c = dcols.get('total_usd')
            _qty_c = dcols.get('qty')
            _price_c = dcols.get('price')
            if _usd_c and _qty_c and _price_c:
                try:
                    if not ws.Cells(pos, _usd_c).HasFormula:
                        _fml = f"=RC[{_qty_c - _usd_c}]*RC[{_price_c - _usd_c}]"
                        ws.Cells(pos, _usd_c).FormulaR1C1 = _fml
                        logging.info(f"[7.5金额] row={pos} {_fml}")
                except Exception as _e:
                    logging.warning(f"[7.5金额] row={pos} 失败: {_e}")

        # 7.5C M开头货号金额公式：金额=总箱×单价（MEC/MSLD等，不论是否混装）
        if not _is_pallet and _is_m_prefix:
            _usd_c = dcols.get('total_usd')
            _price_c = dcols.get('price')
            _tbox_c = dcols.get('total_box')
            if _usd_c and _price_c and _tbox_c:
                try:
                    ws.Cells(pos, _usd_c).ClearContents()
                    formula = f"=RC[{_tbox_c - _usd_c}]*RC[{_price_c - _usd_c}]"
                    ws.Cells(pos, _usd_c).FormulaR1C1 = formula
                    logging.info(f"[M开头金额] row={pos} 金额=总箱(col{_tbox_c})*单价(col{_price_c}) sku={_sku_upper}")
                except Exception as _e:
                    logging.warning(f"[M开头金额] row={pos} 公式写入失败: {_e}")

        # pallets列：公式各文件差异大，保留搜索但跳过绝对行引用
        _pallets_c = dcols.get('pallets')
        if _pallets_c:
            try:
                if not ws.Cells(pos, _pallets_c).HasFormula:
                    _found_fml = False
                    for sr in range(pos - 1, max(3, pos - 50), -1):
                        try:
                            if ws.Cells(sr, _pallets_c).HasFormula:
                                _fml = ws.Cells(sr, _pallets_c).FormulaR1C1 or ''
                                if re.search(r'R\d+C', _fml):
                                    continue
                                ws.Cells(pos, _pallets_c).FormulaR1C1 = _fml
                                _found_fml = True
                                break
                        except:
                            pass
                    if not _found_fml:
                        for sr in range(pos + 1, min(pos + 100, 2000)):
                            try:
                                if ws.Cells(sr, _pallets_c).HasFormula:
                                    _fml = ws.Cells(sr, _pallets_c).FormulaR1C1 or ''
                                    if re.search(r'R\d+C', _fml):
                                        continue
                                    ws.Cells(pos, _pallets_c).FormulaR1C1 = _fml
                                    break
                            except:
                                pass
            except:
                pass

        # 7.5A 卡板货号总箱数：直接写入卡板数（不从附近行复制公式，避免绝对引用引用错误行）
        if _is_pallet and _pallet_count > 0:
            _tbox_c = dcols.get('total_box')
            if _tbox_c:
                try:
                    ws.Cells(pos, _tbox_c).Value = _pallet_count
                    logging.info(f"[卡板总箱] row={pos} 总箱(卡板数)={_pallet_count}")
                except Exception as _e:
                    logging.warning(f"[卡板总箱] row={pos} 写入失败: {_e}")

        # 7.5B 卡板货号金额公式修正：金额=总箱×单价（不是PO数量×单价）
        if _is_pallet and _pallet_count > 0:
            _usd_c = dcols.get('total_usd')
            _price_c = dcols.get('price')
            _tbox_c = dcols.get('total_box')
            if _usd_c and _price_c and _tbox_c:
                try:
                    # 金额 = 总箱列 × 单价列（R1C1格式）
                    formula = f"=RC[{_tbox_c - _usd_c}]*RC[{_price_c - _usd_c}]"
                    ws.Cells(pos, _usd_c).FormulaR1C1 = formula
                    logging.info(f"[卡板金额] row={pos} 金额公式=总箱(col{_tbox_c})*单价(col{_price_c})")
                except Exception as _e:
                    logging.warning(f"[卡板金额] row={pos} 公式写入失败: {_e}")

        # 7.6 修正单价/金额列的数字格式（防止参考行格式为日期导致数字显示为日期）
        for _nfk in ('price', 'total_usd'):
            _nfc = dcols.get(_nfk)
            if _nfc:
                try:
                    _cur_fmt = ws.Cells(pos, _nfc).NumberFormat or ''
                    # 检测日期格式标志（含y/m/d或AAAA等日期格式码）
                    if any(x in _cur_fmt.lower() for x in ('y', 'aaa', '年', '月', '日', 'd/m', 'm/d')):
                        ws.Cells(pos, _nfc).NumberFormat = '#,##0.00' if _nfk == 'total_usd' else '0.00##'
                        logging.info(f"[格式修正] row={pos} col={_nfc}({_nfk}): 日期格式'{_cur_fmt}'→数字格式")
                except:
                    pass

        # 7.8 清除非录入列（参考行复制了旧值，新订单不应继承）
        # A) 跟踪列：验货期和备注之间的列（第三方验货日期、验货结果、船发SO等）
        insp_c = dcols.get('inspection', 0)
        remark_c = dcols.get('remark') or self._note_col_com(ws, pos, mc)
        if insp_c and remark_c and remark_c > insp_c + 1:
            for c in range(insp_c + 1, remark_c):
                try:
                    if not ws.Cells(pos, c).HasFormula:
                        ws.Cells(pos, c).ClearContents()
                except:
                    pass
        # B) 清除系统货号/是否入系统列（金额后的分配列已在Step 7中不复制，无需再清）
        sys_col = dcols.get('system_code')
        if sys_col:
            for sc in (sys_col, sys_col + 1):
                try:
                    if not ws.Cells(pos, sc).HasFormula:
                        ws.Cells(pos, sc).ClearContents()
                except:
                    pass

        # ===== 8. 仅覆盖订单特定字段（全部使用检测到的列位置）=====
        # 核心原则：产品固有属性（中文名、内箱、外箱、总箱公式）全部从参考行继承，
        #          只有订单特定字段从PDF覆盖
        po = header.get('po_number', '')

        # 接单日期（用序列号写入，避免pywin32时区偏移，NumberFormat保留参考行格式）
        # 不用默认列号：检测不到接单日期列时跳过，避免写到错误列
        po_dt = _parse_date(header.get('po_date', ''))
        _po_date_col = dcols.get('po_date')
        if po_dt and _po_date_col:
            ws.Cells(pos, _po_date_col).Value = _date_serial(po_dt)
            try:
                ws.Cells(pos, _po_date_col).NumberFormat = 'yyyy/m/daaa'
            except:
                pass

        # 客户名
        cust = header.get('customer', '')
        if cust:
            _sv_com(ws, pos, dcols.get('customer', 2), cust)

        # 走货国
        dest = header.get('destination_cn', '')
        if dest:
            _sv_com(ws, pos, dcols.get('destination', 3), dest)

        # PO号（先清除公式/旧值，再写入新值，确保不被参考行公式覆盖）
        po_col = dcols.get('po_number', 4)
        if po:
            try:
                ws.Cells(pos, po_col).ClearContents()
            except:
                pass
            _po_str = po if not po.endswith('.0') else po[:-2]
            ws.Cells(pos, po_col).NumberFormat = '@'
            _sv_com(ws, pos, po_col, _po_str)
            logging.info(f"[PO写入] row={pos}, col={po_col}, po={po}")

        # 客户PO（有前导零时保留字符串+文本格式，无前导零的纯数字转int避免小数）
        cpo = ln.get('customer_po', '')
        cpo_col = dcols.get('customer_po', 5)
        if cpo:
            cpo_str = str(cpo).strip()
            if cpo_str.endswith('.0'):
                cpo_str = cpo_str[:-2]
            ws.Cells(pos, cpo_col).NumberFormat = '@'
            _sv_com(ws, pos, cpo_col, cpo_str)
        else:
            # 客PO未提供时清空（不保留参考行的旧值）
            try:
                if not ws.Cells(pos, cpo_col).HasFormula:
                    ws.Cells(pos, cpo_col).ClearContents()
            except:
                pass

        # SKU (PO-line format)
        c_sku = dcols.get('sku', 6)
        line_no = ln.get('line_no', '')
        if po and line_no:
            _sv_com(ws, pos, c_sku, f"{po}-{line_no}")
        elif 'sku' in dcols:
            ref_sku_val = ''
            try:
                ref_sku_val = str(ws.Cells(actual_ref, c_sku).Value or '')
            except:
                pass
            if re.match(r'^\d{7,}-\d+$', ref_sku_val):
                pass
            elif ln.get('sku_spec') or ln.get('sku'):
                _sv_com(ws, pos, c_sku, ln.get('sku_spec', '') or ln.get('sku', ''))
        else:
            ref_sku_val = ''
            try:
                ref_sku_val = str(ws.Cells(actual_ref, c_sku).Value or '')
            except:
                pass
            if re.match(r'^\d{7,}-\d+$', ref_sku_val):
                pass
            elif ln.get('sku_spec') or ln.get('sku'):
                _sv_com(ws, pos, c_sku, ln.get('sku_spec', '') or ln.get('sku', ''))

        # ITEM#/货号 — 用PDF的完整sku_spec覆写（如"9296-S001"），大小写跟随排期已有条目
        item_col = dcols.get('items')
        sku_spec_val = ln.get('sku_spec', '') or ln.get('sku', '')
        if item_col and sku_spec_val:
            # 检查参考行的货号大小写，跟随排期已有格式
            try:
                ref_item_val = str(ws.Cells(actual_ref, item_col).Value or '').strip()
                if ref_item_val and ref_item_val.upper() == sku_spec_val.upper():
                    sku_spec_val = ref_item_val  # 使用参考行的大小写（如s001 vs S001）
                    logging.info(f"[货号大小写] 跟随参考行: '{ref_item_val}'")
            except:
                pass
            _sv_com(ws, pos, item_col, sku_spec_val)
            logging.info(f"[货号写入] row={pos}, col={item_col}, 货号={sku_spec_val}")
        # 产品名称/中文名 — 直接取参考行品名，不投票（避免±200范围内含规格的品名污染结果）
        _pn_col = dcols.get('product_name')
        if _pn_col:
            try:
                def _is_valid_pn(s):
                    """判断品名是否有效：含中文字符，或含英文但长度≥4且不是汇总行"""
                    if not s:
                        return False
                    import re as _re
                    # 排除包装数量描述：如"20个空PDQ"、"迷你包包S4个空PDQ"
                    if _re.search(r'\d+个空[A-Za-z]', s):
                        return False
                    # 排除纯数字+PDQ描述（如"4PDQ"、"20 PDQ"开头）
                    if _re.match(r'^\d+\s*(?:个\s*)?(?:空\s*)?PDQ', s, _re.I):
                        return False
                    if any('\u4e00' <= c <= '\u9fff' for c in s):
                        return True  # 含中文，有效
                    # 全英文：排除汇总行关键词
                    _bad = {'total', 'subtotal', 'grand total', 'sum', 'qty', 'amount'}
                    if s.lower().rstrip(':') in _bad or len(s) < 4:
                        return False
                    # 全英文且只含包装规格词（PCS/PDQ/CTN/数字/斜杠），不是产品名
                    if _re.match(r'^[\d\s/\-\.A-Z]+$', s, _re.I) and _re.search(r'(?:PCS?|PDQ|CTN)', s, _re.I):
                        return False
                    return True

                def _read_pn(r):
                    """读品名列，自动处理WPS合并单元格（合并区内格返回空时取首格值）"""
                    try:
                        cell = ws.Cells(r, _pn_col)
                        v = str(cell.Value or '').strip()
                        if not v:
                            try:
                                if cell.MergeCells:
                                    v = str(cell.MergeArea.Cells(1, 1).Value or '').strip()
                            except:
                                pass
                        return v
                    except:
                        return ''

                # 中文名：优先查直查表（毫秒级），查不到再COM Find
                _target_spec = _sku_spec(sku_spec_val)
                _search_col = item_col or dcols.get('items')
                logging.info(f"[中文名诊断] pos={pos} ref_row={ref_row} actual_ref={actual_ref} _pn_col={_pn_col}")

                # 先清除步骤7从参考行复制过来的中文名（可能是错的）
                try:
                    ws.Cells(pos, _pn_col).ClearContents()
                except:
                    pass

                _found = False

                # 方案A：直查表（item_cn_name_map.json），先按完整货号精确匹配，再fallback基础码
                _target_base = _item_code(sku_spec_val)
                _target_full = re.sub(r'[\s\n]+', '', str(sku_spec_val).strip()).upper()
                if (_target_full or _target_base) and not _found:
                    try:
                        if self._cn_name_cache is None:
                            _cn_map_path = os.path.join(os.path.dirname(__file__), 'data', 'item_cn_name_map.json')
                            if os.path.exists(_cn_map_path):
                                with open(_cn_map_path, 'r', encoding='utf-8') as _f:
                                    self._cn_name_cache = json.load(_f)
                                logging.info(f"[中文名] 已加载直查表，{len(self._cn_name_cache)}个货号")
                            else:
                                self._cn_name_cache = {}
                        # 先按完整货号精确匹配，再fallback基础码
                        _cn_entry = self._cn_name_cache.get(_target_full) or self._cn_name_cache.get(_target_base)
                        if _cn_entry and _is_valid_pn(_cn_entry.get('cn_name', '')):
                            _sv_com(ws, pos, _pn_col, _cn_entry['cn_name'])
                            logging.info(f"[中文名] row={pos} <- 直查表 '{_cn_entry['cn_name']}' (key={_target_full or _target_base})")
                            _found = True
                    except Exception as _map_e:
                        logging.warning(f"[中文名] 直查表异常: {_map_e}")

                # 方案B：COM Find兜底（直查表无此货号时）
                if not _found and _target_spec and _search_col:
                    _search_rng = ws.Columns(_search_col)
                    _hit = None
                    try:
                        _hit = _search_rng.Find(
                            What=_target_spec,
                            After=ws.Cells(pos, _search_col),
                            LookAt=2,            # xlPart
                            SearchDirection=2     # xlPrevious
                        )
                    except:
                        pass
                    if _hit and _hit.Row != pos:
                        if _sku_spec(str(_hit.Value or '')) == _target_spec:
                            pn = _read_pn(_hit.Row)
                            if _is_valid_pn(pn):
                                _sv_com(ws, pos, _pn_col, pn)
                                logging.info(f"[中文名] row={pos} <- Find同货号row={_hit.Row} '{pn}'")
                                _found = True
                    if not _found and _hit and _hit.Row != pos:
                        _start = _hit.Row
                        for _attempt in range(20):
                            try:
                                _hit2 = _search_rng.FindNext(After=_hit)
                                if not _hit2 or _hit2.Row == _start or _hit2.Row == pos:
                                    break
                                if _sku_spec(str(_hit2.Value or '')) == _target_spec:
                                    pn = _read_pn(_hit2.Row)
                                    if _is_valid_pn(pn):
                                        _sv_com(ws, pos, _pn_col, pn)
                                        logging.info(f"[中文名] row={pos} <- FindNext同货号row={_hit2.Row} '{pn}'")
                                        _found = True
                                        break
                                _hit = _hit2
                            except:
                                break
                if not _found:
                    logging.info(f"[中文名] row={pos} 无完全相同货号，留空，spec={sku_spec_val}")
            except Exception as _pn_e:
                logging.warning(f"[中文名] row={pos} 异常: {_pn_e}")
        # 内箱/外箱
        _inner = ln.get('inner_pcs', 0) or 0
        if 'inner_box' in dcols:
            try:
                if not ws.Cells(pos, dcols['inner_box']).HasFormula:
                    _sv_com(ws, pos, dcols['inner_box'], _inner)
            except:
                pass
        # 卡板货号：从参考行读取每卡板件数（外箱K列），供外箱写入和PO数量计算共用
        _ref_pcs_per_pallet = 0
        if _is_pallet and _pallet_count > 0:
            _ob_col = dcols.get('outer_box')
            if _ob_col:
                try:
                    _ref_v = ws.Cells(actual_ref, _ob_col).Value
                    if _ref_v and float(_ref_v) > 1:
                        _ref_pcs_per_pallet = int(float(_ref_v))
                except Exception:
                    pass

        if _is_pallet and _pallet_count > 0:
            _qty_val = ln.get('qty', 0) or 0
            if int(_qty_val) > int(_pallet_count):
                # 有PRODUCT子行：qty已是总件数，正常反推外箱
                _outer = int(_qty_val / _pallet_count)
            elif _ref_pcs_per_pallet > 1:
                # 无PRODUCT子行（qty==pallet_count）：从参考行取每卡板件数
                _outer = _ref_pcs_per_pallet
                logging.info(f"[卡板外箱] row={pos} 无PRODUCT行，从参考行取外箱={_outer}")
            else:
                _outer = int(_qty_val / _pallet_count) if _pallet_count else 0
            logging.info(f"[卡板外箱] row={pos} 外箱={_outer} (PO数量{_qty_val}/卡板数{_pallet_count})")
        elif ln.get('is_mixed_carton') and ln.get('carton_count') and ln.get('carton_count') > 0:
            # 混装货号（7154/7153/25257等）：外箱=产品总数÷箱数
            _outer = int(ln.get('qty', 0) / ln['carton_count'])
            logging.info(f"[混装外箱] row={pos} 外箱={_outer} ({ln.get('qty')}/{ln['carton_count']})")
            # 强制写入（清除参考行复制来的公式，不受HasFormula限制）
            if 'outer_box' in dcols:
                try:
                    ws.Cells(pos, dcols['outer_box']).ClearContents()
                    _sv_com(ws, pos, dcols['outer_box'], _outer)
                except Exception as _e:
                    logging.warning(f"[混装外箱] row={pos} 写入失败: {_e}")
        else:
            _outer = ln.get('outer_qty', 0) or 0
        if not ln.get('is_mixed_carton') and 'outer_box' in dcols:
            try:
                if not ws.Cells(pos, dcols['outer_box']).HasFormula:
                    _sv_com(ws, pos, dcols['outer_box'], _outer)
            except:
                pass
        # 总箱数/卡板/金额 — 不覆写（从参考行复制公式，自动计算）
        # 系统货号 — 不填（明确清空，带"系统"的货号列不填）
        if 'system_code' in dcols:
            try:
                ws.Cells(pos, dcols['system_code']).ClearContents()
            except:
                pass

        # PO数量（唯一从PDF取的数量字段）
        qty = ln.get('qty', 0)
        if qty:
            if _is_pallet and _pallet_count > 0 and int(qty) <= int(_pallet_count):
                # qty实际是卡板数（无PRODUCT子行），I列应写总件数 = 卡板数 × 每卡板件数
                if _ref_pcs_per_pallet > 1:
                    qty = _pallet_count * _ref_pcs_per_pallet
                    logging.info(f"[卡板PO数量] row={pos} 从参考行反推总件数: "
                                 f"{_pallet_count}卡板×{_ref_pcs_per_pallet}pcs={qty}")
            _sv_com(ws, pos, dcols.get('qty', 9), qty)

        # 出货日期（用序列号写入，避免pywin32时区偏移）
        if ship_dt:
            ws.Cells(pos, ship_col).Value = _date_serial(ship_dt)
            # 客PO期 = 走货日期 = 出货日期（ZURU订单三者相同）
            cpo_date_col = dcols.get('cpo_date')
            if cpo_date_col and cpo_date_col != ship_col:
                ws.Cells(pos, cpo_date_col).Value = _date_serial(ship_dt)

        # 验货日期（用抽取的_calc_inspection函数计算）
        insp_col = dcols.get('inspection')
        if insp_col and ship_dt:
            try:
                sn = ws.Name if ws.Name else ''
                wb_name = ''
                try:
                    wb_name = ws.Parent.Name if ws.Parent else ''
                except:
                    pass
                insp_dt = _calc_inspection(ship_dt, wb_name, sn)
                if insp_dt:
                    ws.Cells(pos, insp_col).Value = _date_serial(insp_dt)
                    logging.info(f"[验货期] row={pos} 出货={ship_dt.strftime('%Y-%m-%d')} → 验货={insp_dt.strftime('%Y-%m-%d')}")
            except Exception as _insp_e:
                logging.warning(f"[验货期] 计算失败: {_insp_e}")

        # 备注列：写入PDF备注内容（按货号过滤，只写与当前行相关的备注）
        nc = dcols.get('remark') or self._note_col_com(ws, pos, mc)
        if nc:
            _tc = header.get('tracking_code', '') or ''
            _pi = header.get('packaging_info', '') or ''
            _rm = header.get('remark', '') or ''
            _note_parts = []
            if _tc:
                _note_parts.append(_tc)
            if _pi:
                _note_parts.append(f"Packaging Info: {_pi}")
            if _rm:
                _note_parts.append(f"Remark: {_rm}")
            _full_note = '\n'.join(_note_parts)
            if _full_note:
                try:
                    ws.Cells(pos, nc).Value = _full_note
                    logging.info(f"[备注] row={pos} 写入PDF完整备注 ({len(_full_note)}字符)")
                except Exception as e:
                    logging.warning(f"[备注] row={pos} 写入失败: {e}")

        # 跟单人/业务（仅检测到时写入）
        if 'from_person' in dcols:
            fp = header.get('from_person', '')
            if fp:
                _sv_com(ws, pos, dcols['from_person'], fp.strip())

        # 单价USD（仅检测到且有数据时写入）
        if 'price' in dcols and ln.get('price', 0) > 0:
            _sv_com(ws, pos, dcols['price'], ln['price'])
            # 防止参考行格式为日期导致数字显示为日期
            try:
                ws.Cells(pos, dcols['price']).NumberFormat = '0.00##'
            except:
                pass

        # 8A. 混装箱（MEC/7154/7153/25257等）：总箱数=箱数（carton_count），直接写入
        if ln.get('is_mixed_carton') and ln.get('carton_count') and dcols.get('total_box'):
            try:
                ws.Cells(pos, dcols['total_box']).ClearContents()
                ws.Cells(pos, dcols['total_box']).Value = ln['carton_count']
                logging.info(f"[混装总箱数] row={pos} 总箱数={ln['carton_count']}(箱数)")
            except Exception as e:
                logging.warning(f"[混装总箱数] 写入失败: {e}")
            # 混装金额=总箱×单价（不是数量×单价）
            _usd_c = dcols.get('total_usd')
            _price_c = dcols.get('price')
            _tbox_c = dcols.get('total_box')
            if _usd_c and _price_c and _tbox_c:
                try:
                    ws.Cells(pos, _usd_c).ClearContents()
                    formula = f"=RC[{_tbox_c - _usd_c}]*RC[{_price_c - _usd_c}]"
                    ws.Cells(pos, _usd_c).FormulaR1C1 = formula
                    logging.info(f"[混装金额] row={pos} 金额=总箱(col{_tbox_c})*单价(col{_price_c})")
                except Exception as e:
                    logging.warning(f"[混装金额] 公式写入失败: {e}")

        # 8B. 金额公式：始终从同货号参考行复制（步骤7已完成），不再按阈值覆写
        # 仅当金额列无公式时（参考行也无公式），步骤7.5已向上搜索修复

        # 9. 新录入行：蓝色填充 + 黑色字体（只到金额列，分配列绝不填充）
        # 多重保险：金额列 > 备注列+2 > 所有已检测列的最大值 > 硬上限35
        if _total_usd_c:
            _fill_end = _total_usd_c
        else:
            # 金额列未检测到时，用其他已知列推算边界
            _remark_c = dcols.get('remark', 0)
            if _remark_c:
                _fill_end = _remark_c + 2  # 备注后最多再填2列(ZURU预计、单价)
            else:
                # 所有已检测列的最大值 + 2
                _detected_max = max(dcols.values()) if dcols else 0
                _fill_end = _detected_max + 2 if _detected_max else 35
            logging.warning(f"[填充] 金额列未检测到! 使用备选边界={_fill_end} (remark={_remark_c}, dcols_max={max(dcols.values()) if dcols else 0})")
        # 硬上限：绝不超过50列（ZURU排期金额列最远不超过40列）
        _fill_end = min(_fill_end, 50)
        logging.info(f"[填充] row={pos} _fill_end={_fill_end} total_usd={_total_usd_c}")
        try:
            new_rng = ws.Range(ws.Cells(pos, 1), ws.Cells(pos, _fill_end))
            new_rng.Interior.Color = BLUE_COM
            new_rng.Font.Color = 0  # 黑色字体
        except Exception as e:
            logging.warning(f"[新录入] 格式设置失败: {e}")

        # 9A. 备注列：不做特殊填充（新单备注已从PDF写入，跟随整行蓝色）

        # 9B. 比例检测：检查备注中是否含比例模式（如#15733: A:B:D:E=1:1:1:1）
        #     不自动修改分配列，只返回警告让用户自己处理比例
        _pkg = header.get('packaging_info', '') or ''
        _rmk = header.get('remark', '') or ''
        _all_note = f"{_pkg}\n{_rmk}"
        _ratio_pat = re.search(r'#\d{3,}[A-Za-z]*\s*:\s*[A-Z](?:\s*:\s*[A-Z])+\s*=\s*\d', _all_note, re.I)
        _ratio_warning = ''
        if _ratio_pat:
            _ratio_warning = _ratio_pat.group()
            logging.info(f"[比例提醒] row={pos} 备注含比例模式: {_ratio_warning}")

        # 10. 验证：检查所有必填字段是否成功写入
        # 验证前记录列检测结果（调试用，保留INFO级别）
        logging.info(f"[新录入验证] dcols: ship_date={'ship_date' in dcols}, mc={mc}")
        _warnings = []
        _check_fields = [
            ('items', '货号(ITEM#)'),
            ('product_name', '中文名/货名'),
            ('qty', 'PO数量'),
            ('inner_box', '内箱装箱数'),
            ('outer_box', '外箱装箱数'),
            ('ship_date', '出货日期/走货日期'),
            ('customer', '客户名'),
            ('destination', '国家'),
            ('po_number', 'PO号'),
        ]
        for col_key, label in _check_fields:
            if col_key in dcols:
                try:
                    cv = ws.Cells(pos, dcols[col_key]).Value
                    if cv is None or (isinstance(cv, str) and not cv.strip()):
                        _warnings.append(f"{label}为空(col={dcols[col_key]})")
                except:
                    pass
            else:
                _warnings.append(f"{label}列未检测到")
        # 金额异常检查：单价<40但金额异常大时，精确提醒用户确认
        _price_val = ln.get('price', 0) or 0
        if _price_val < 40 and dcols.get('total_usd'):
            try:
                amount_val = ws.Cells(pos, dcols['total_usd']).Value
                if amount_val and isinstance(amount_val, (int, float)) and amount_val > 500000:
                    cell_addr = ws.Cells(pos, dcols['total_usd']).Address.replace('$', '')
                    _warnings.append(f"金额{amount_val:,.0f}异常偏大，请确认单元格{cell_addr}")
            except:
                pass

        if _warnings:
            logging.warning(f"[新录入验证] row={pos}, item={ln.get('sku','')}: {', '.join(_warnings)}")

        # 11. 15706 Fuggler角色配比公式（仅当sku_spec在配比表中有记录时触发）
        try:
            self._apply_15706_ratio(ws, pos, sku_spec_val, dcols)
        except Exception as _re:
            logging.warning(f"[配比] 应用失败 row={pos}: {_re}")

        return pos, _warnings, _ratio_warning

    @classmethod
    def _load_ratio_map(cls):
        """扫描_RATIO_DIR内所有*配比*.xlsx，动态检测字母列，合并为统一配方字典。
        返回 {sku_spec: [(char_letter, ratio), ...]}，重复sku_spec取先加载的文件优先。"""
        with cls._ratio_map_lock:
            if cls._ratio_map is not None:
                return cls._ratio_map
            ratio_map = {}
            try:
                if not os.path.isdir(cls._RATIO_DIR):
                    logging.warning(f"[配比] 目录不存在: {cls._RATIO_DIR}")
                    cls._ratio_map = ratio_map
                    return ratio_map
                ratio_files = sorted(f for f in os.listdir(cls._RATIO_DIR)
                                     if '配比' in f and f.endswith('.xlsx') and not f.startswith('~'))
                for fname in ratio_files:
                    fpath = os.path.join(cls._RATIO_DIR, fname)
                    try:
                        wb = openpyxl.load_workbook(fpath, read_only=True, data_only=True)
                        ws = wb.active
                        # 动态检测字母列：扫描前5行，找含≥3个单大写字母的行
                        char_col_map = {}  # {letter: col_idx}
                        sku_col = 7       # 默认G列=系统货号，扫描时覆盖
                        max_col = ws.max_column or 80
                        for r in range(1, 6):
                            found = {}
                            for c in range(1, min(max_col + 1, 100)):
                                v = ws.cell(r, c).value
                                if v is None:
                                    continue
                                vs = str(v).strip()
                                if re.match(r'^[A-Z]$', vs):
                                    found[vs] = c
                                elif vs == '系统货号':
                                    sku_col = c
                            if len(found) >= 3:
                                char_col_map = found
                                break
                        if not char_col_map:
                            logging.warning(f"[配比] {fname} 未找到字母列，跳过")
                            wb.close()
                            continue
                        loaded = 0
                        for r in range(4, (ws.max_row or 500) + 1):
                            sku = ws.cell(r, sku_col).value
                            if not sku or not isinstance(sku, str):
                                continue
                            sku = sku.strip()
                            if not sku or sku in ('系统货号', 'SKU'):
                                continue
                            if sku in ratio_map:
                                continue  # 重复取第一次（先扫的文件优先）
                            chars = []
                            for letter, col in char_col_map.items():
                                val = ws.cell(r, col).value
                                if val is not None:
                                    try:
                                        chars.append((letter, float(val)))
                                    except (TypeError, ValueError):
                                        pass
                            if chars:
                                ratio_map[sku] = chars
                                loaded += 1
                        wb.close()
                        logging.info(f"[配比] {fname} 加载{loaded}个配方")
                    except Exception as e:
                        logging.warning(f"[配比] {fname} 加载失败: {e}")
                logging.info(f"[配比] 合计{len(ratio_map)}个sku_spec配方")
            except Exception as e:
                logging.warning(f"[配比] 加载失败: {e}")
            cls._ratio_map = ratio_map
            return ratio_map

    def _apply_15706_ratio(self, ws, pos, sku_spec, dcols):
        """对新录入行写入15706 Fuggler角色配比公式。
        只在sku_spec匹配配比表时触发，其余货号完全不受影响。
        公式：=PO数量列 * ratio / 10（R1C1相对引用，不绑定行号）"""
        ratio_map = self._load_ratio_map()
        if not sku_spec or sku_spec not in ratio_map:
            return
        chars = ratio_map[sku_spec]  # [(letter, ratio), ...]
        qty_col = dcols.get('qty')
        if not qty_col:
            logging.warning(f"[配比] row={pos} qty列未检测到，跳过")
            return

        # 扫描排期表头行，找各角色字母对应的列号
        # 角色列在排期第3行，字母值为 'A'/'B '/'D'/'E '/... 等（可能有空格）
        char_col_map = {}  # {letter: col_index}
        _target_letters = {c[0].strip() for c in chars}
        try:
            # 找表头行：向上最多扫20行，找含角色字母的行
            _max_scan = min(pos - 1, 20)
            for scan_r in range(1, _max_scan + 1):
                _found = 0
                for scan_c in range(40, min(ws.UsedRange.Columns.Count + 1, 80)):
                    try:
                        _v = ws.Cells(scan_r, scan_c).Value
                        if _v is None:
                            continue
                        _vs = str(_v).strip()
                        if _vs in _target_letters:
                            char_col_map[_vs] = scan_c
                            _found += 1
                    except Exception:
                        pass
                if _found >= len(_target_letters) * 0.6:  # 至少找到60%的角色列
                    break
        except Exception as e:
            logging.warning(f"[配比] 表头扫描失败: {e}")
            return

        if not char_col_map:
            logging.warning(f"[配比] row={pos} 未找到角色列，跳过")
            return

        written = 0
        for letter, ratio in chars:
            char_col = char_col_map.get(letter)
            if char_col is None:
                continue
            try:
                cell = ws.Cells(pos, char_col)
                # 只填空格，不覆盖已有数据
                if cell.Value is not None and str(cell.Value).strip():
                    continue
                offset = qty_col - char_col
                if ratio == 10.0:
                    # 100%：=RC[offset]
                    fml = f"=RC[{offset}]"
                else:
                    # 通用：=RC[offset]*ratio/10
                    # 如果ratio是整数则去掉小数点
                    r_str = str(int(ratio)) if ratio == int(ratio) else str(ratio)
                    fml = f"=RC[{offset}]*{r_str}/10"
                cell.FormulaR1C1 = fml
                written += 1
            except Exception as e:
                logging.warning(f"[配比] row={pos} 角色{letter} col={char_col} 写入失败: {e}")

        if written:
            logging.info(f"[配比] row={pos} sku={sku_spec} 写入{written}个角色配比公式")

    def _do_modify_com(self, ws, row, max_col, changes, note_changed=False, pdf_note='', fill_empty=None):
        """通过COM修改指定单元格值，改动单元格蓝色填充标记。
        出货期变更时自动联动更新验货日期。
        note_changed=True时：只对备注列标绿色，不写入任何内容（内容保持不变）。
        fill_empty: 需要补填的空字段 {'price': float, 'from_person': str}"""
        fill_empty = fill_empty or {}

        # 写入变更值，同时检测出货期列变更（合并为单次遍历）
        dcols = self._detect_cols(ws, max_col)
        ship_col = dcols.get('ship_date', 0)
        _po_col_nums = {dcols.get('po_number', 0), dcols.get('customer_po', 0)} - {0}
        _ship_changed = False
        _new_ship_dt = None
        for cl, nv in changes.items():
            cn = _col_num(cl)
            cell = ws.Cells(row, cn)

            # 设置值（按值内容检测日期，不限制列号）
            is_date = False
            if isinstance(nv, str) and re.match(r'\d{4}[-/]\d{1,2}[-/]\d{1,2}', str(nv)):
                is_date = True
            if is_date:
                try:
                    dt = _parse_date(nv)
                    if dt:
                        cell.Value = _date_serial(dt)
                    else:
                        cell.Value = nv
                except:
                    cell.Value = nv
                # 顺便检测出货期列变更（省去第二次遍历）
                if cn == ship_col and not _ship_changed:
                    _new_ship_dt = _parse_date(nv) if isinstance(nv, str) else None
                    if _new_ship_dt:
                        _ship_changed = True
            else:
                _nv_str = str(nv).strip()
                if _nv_str.isdigit() and _nv_str.startswith('0') and len(_nv_str) > 1:
                    # 有前导零：设文本格式保留原值
                    cell.NumberFormat = "@"
                    cell.Value = _nv_str
                else:
                    try:
                        cell.Value = int(nv) if _nv_str.isdigit() else float(nv)
                    except:
                        cell.Value = nv
                    # PO号/客PO列：纯数字设数值格式0位小数
                    if cn in _po_col_nums and _nv_str.isdigit():
                        cell.NumberFormat = '0'
            # 改动单元格蓝色填充标记（只标改动的格，不是整行）
            try:
                cell.Interior.Color = BLUE_COM
            except:
                pass

        if _ship_changed and _new_ship_dt:
            insp_col = dcols.get('inspection')
            if insp_col:
                try:
                    wb_name = ws.Parent.Name if ws.Parent else ''
                    sn = ws.Name or ''
                    insp_dt = _calc_inspection(_new_ship_dt, wb_name, sn)
                    if insp_dt:
                        ws.Cells(row, insp_col).Value = _date_serial(insp_dt)
                        # 验货期改动的单元格蓝色填充
                        ws.Cells(row, insp_col).Interior.Color = BLUE_COM
                        logging.info(f"[修改单验货期] row={row} 出货={_new_ship_dt.strftime('%Y-%m-%d')} → 验货={insp_dt.strftime('%Y-%m-%d')}")
                except Exception as e:
                    logging.warning(f"[修改单验货期] 计算失败: {e}")

        # 修改单不改变行格式（保持原有底色和字体颜色，验货期联动除外）

        # 修改单备注不做任何标记，留给用户自己检查

        # ===== 空字段补填：排期已有记录但部分字段为空 =====
        if fill_empty or not _ship_changed:
            try:
                # 1) 单价为空 → 写入PDF单价，蓝色填充
                if fill_empty.get('price'):
                    pc = dcols.get('price')
                    if pc:
                        existing_price = ws.Cells(row, pc).Value
                        if existing_price is None or str(existing_price).strip() == '':
                            ws.Cells(row, pc).Value = fill_empty['price']
                            ws.Cells(row, pc).Interior.Color = BLUE_COM
                            logging.info(f"[修改单补填] row={row} 单价={fill_empty['price']}")

                # 2) 跟单员为空 → 写入PDF跟单员，蓝色填充
                if fill_empty.get('from_person'):
                    fpc = dcols.get('from_person')
                    if fpc:
                        existing_fp = str(ws.Cells(row, fpc).Value or '').strip()
                        if not existing_fp:
                            fp_val = fill_empty['from_person'].strip()
                            ws.Cells(row, fpc).Value = fp_val
                            ws.Cells(row, fpc).Interior.Color = BLUE_COM
                            logging.info(f"[修改单补填] row={row} 跟单={fp_val}")

                # 3) 验货期为空 → 从出货期计算，蓝色填充
                if not _ship_changed:
                    insp_col = dcols.get('inspection')
                    if insp_col:
                        existing_insp = ws.Cells(row, insp_col).Value
                        if existing_insp is None or str(existing_insp).strip() == '':
                            # 读取同行出货期
                            sc = dcols.get('ship_date')
                            if sc:
                                ship_val = ws.Cells(row, sc).Value
                                ship_dt = None
                                if ship_val is not None:
                                    if hasattr(ship_val, 'year'):
                                        ship_dt = datetime(ship_val.year, ship_val.month, ship_val.day)
                                    elif isinstance(ship_val, (int, float)) and int(ship_val) > 40000:
                                        # Excel序列号 → datetime
                                        ship_dt = datetime(1899, 12, 30) + timedelta(days=int(ship_val))
                                    elif isinstance(ship_val, str) and ship_val.strip():
                                        ship_dt = _parse_date(ship_val)
                                if ship_dt:
                                    wb_name = ws.Parent.Name if ws.Parent else ''
                                    sn = ws.Name or ''
                                    insp_dt = _calc_inspection(ship_dt, wb_name, sn)
                                    if insp_dt:
                                        ws.Cells(row, insp_col).Value = _date_serial(insp_dt)
                                        ws.Cells(row, insp_col).Interior.Color = BLUE_COM
                                        logging.info(f"[修改单补填] row={row} 验货期={insp_dt.strftime('%Y-%m-%d')}")

            except Exception as e:
                logging.warning(f"[修改单补填] row={row} 异常: {e}")

        # 4) 金额为空 → 从附近行复制公式（出货期修改单也需补填，不受_ship_changed影响）
        try:
            total_usd_col = dcols.get('total_usd')
            if total_usd_col:
                existing_amount = ws.Cells(row, total_usd_col).Value
                has_formula = False
                try:
                    has_formula = ws.Cells(row, total_usd_col).HasFormula
                except:
                    pass
                if not has_formula and (existing_amount is None or str(existing_amount).strip() == ''):
                    for sr in range(row - 1, max(3, row - 50), -1):
                        try:
                            if ws.Cells(sr, total_usd_col).HasFormula:
                                ws.Cells(row, total_usd_col).FormulaR1C1 = ws.Cells(sr, total_usd_col).FormulaR1C1
                                ws.Cells(row, total_usd_col).Interior.Color = BLUE_COM
                                logging.info(f"[修改单补填] row={row} 金额公式从row={sr}复制")
                                break
                        except:
                            pass
        except Exception as e:
            logging.warning(f"[修改单补填金额] row={row} 异常: {e}")

    def _do_cancel_com(self, wb, ws, row, max_col):
        """通过COM取消行：复制到取消订单Sheet → 标红+蓝 → 删原行
        无取消Sheet时：原行清除填充+红字，不删除行
        返回True表示删除了原行，False表示仅标记未删除"""
        mc = min(max_col, 100)

        # 判断是否为总排期文件（总排期只删不复制）
        wb_name = wb.Name if wb.Name else ''
        is_summary = '总' in wb_name

        if not is_summary:
            # 查找取消订单Sheet（不自动创建）
            cancel_ws = None
            for i in range(1, wb.Sheets.Count + 1):
                if '取消' in wb.Sheets(i).Name:
                    cancel_ws = wb.Sheets(i)
                    break

            if cancel_ws is None:
                # 无取消订单Sheet → 原行标记：清除填充 + 红色字体，不删除行
                row_rng = ws.Range(ws.Cells(row, 1), ws.Cells(row, mc))
                row_rng.Interior.ColorIndex = 0  # 清除填充（xlNone）
                row_rng.Font.Color = RED_COM
                logging.info(f"[取消] 无取消订单Sheet，原行{row}标记红字无填充")
                return False  # 不删除原行

            # 找取消Sheet下一空行
            try:
                cr = cancel_ws.Cells(cancel_ws.Rows.Count, 1).End(-4162).Row + 1  # xlUp
                if cr < 1:
                    cr = 1
            except:
                cr = 1

            # 复制整行到取消Sheet（保留原格式，超范围时自动缩小）
            try:
                src_rng = ws.Range(ws.Cells(row, 1), ws.Cells(row, mc))
                src_rng.Copy(Destination=cancel_ws.Range(cancel_ws.Cells(cr, 1),
                                                          cancel_ws.Cells(cr, mc)))
            except Exception as e:
                logging.warning(f"[取消] 复制{mc}列失败: {e}, 回退到50列")
                mc = 50
                src_rng = ws.Range(ws.Cells(row, 1), ws.Cells(row, mc))
                src_rng.Copy(Destination=cancel_ws.Range(cancel_ws.Cells(cr, 1),
                                                          cancel_ws.Cells(cr, mc)))

            # 取消Sheet中：红色字体 + 浅蓝底色
            dest_rng = cancel_ws.Range(cancel_ws.Cells(cr, 1), cancel_ws.Cells(cr, mc))
            dest_rng.Font.Color = RED_COM
            dest_rng.Interior.Color = BLUE_COM

            # 清剪贴板
            try:
                ws.Application.CutCopyMode = False
            except:
                pass

        # 删除原行（有取消Sheet时才删，总排期也删）
        ws.Rows(row).Delete()
        return True

    def _insert_pos_com(self, ws, ship_dt, col=13, start_after=0):
        """通过COM查找按出货日期的插入位置
        start_after: 从此行之后开始搜索（同批次多条目时保持插入顺序）
        使用 >= 比较: 新条目插在同日期已有条目之前（最新PO排在最上面）"""
        start_row = max(4, start_after + 1)
        last = start_row - 1
        try:
            used_rows = ws.UsedRange.Row + ws.UsedRange.Rows.Count - 1
        except:
            used_rows = 100

        # 若无出货日期，直接插到末尾
        if ship_dt is None:
            for r in range(start_row, min(used_rows + 1, 5000)):
                v = ws.Cells(r, col).Value
                if v is not None and hasattr(v, 'year'):
                    last = r
            return last + 1

        for r in range(start_row, min(used_rows + 1, 5000)):
            v = ws.Cells(r, col).Value
            if v is None:
                continue
            try:
                if hasattr(v, 'year'):
                    dt = datetime(v.year, v.month, v.day)
                    last = r
                    if dt >= ship_dt:
                        return r
            except:
                continue
        return last + 1

    def _note_col_com(self, ws, row, max_col):
        """通过COM查找备注列"""
        for c in [23, 22, 26]:
            for r in range(max(2, row - 10), row):
                try:
                    v = ws.Cells(r, c).Value
                    if v and ('日期码' in str(v) or 'Remark' in str(v)):
                        return c
                except:
                    pass
        return 23

    # =================== 兼容旧接口（单条操作，也用COM）===================

    def enter_new(self, sched, header, lines):
        local = os.path.join(DESKTOP, 'schedule_temp.xlsx')
        shutil.copy2(sched['file'], local)
        app = self._com_app()
        try:
            wb = app.Workbooks.Open(os.path.abspath(local))
            ws = wb.Sheets(sched['sheet'])
            ref = sched['ref']
            mc = min(sched.get('mcol', 100), 100)
            inserted_positions = []
            last_insert_pos = 0  # 同批次保持插入顺序
            ratio_warnings = []
            for ln in lines:
                # 累计偏移：之前插入的行会使ref下移
                adj_ref = ref
                for p in inserted_positions:
                    if p <= adj_ref:
                        adj_ref += 1
                _ptm = sched.get('po_type_match', True)
                pos, _w, _rw = self._do_new_com(ws, adj_ref, mc, header, ln,
                                       start_after=last_insert_pos,
                                       po_type_match=_ptm)
                inserted_positions.append(pos)
                last_insert_pos = pos
                if _rw:
                    ratio_warnings.append({
                        'po': ln.get('po', '') or header.get('po', ''),
                        'sku': ln.get('sku', ''),
                        'pattern': _rw,
                        'file': os.path.basename(sched['file']),
                        'sheet': sched['sheet'],
                        'row': pos
                    })
            wb.Save()
            wb.Close(False)
        finally:
            self._com_quit(app)
        msg = f'已录入{len(lines)}行'
        result = {'ok': True, 'local': local, 'z': sched['file'], 'msg': msg}
        if ratio_warnings:
            result['ratio_warnings'] = ratio_warnings
        return result

    def modify(self, record, changes):
        local = os.path.join(DESKTOP, 'schedule_temp.xlsx')
        shutil.copy2(record['file'], local)
        app = self._com_app()
        try:
            wb = app.Workbooks.Open(os.path.abspath(local))
            ws = wb.Sheets(record['sheet'])
            mc = min(ws.UsedRange.Columns.Count + ws.UsedRange.Column, 100)
            self._do_modify_com(ws, record['row'], mc, changes)
            wb.Save()
            wb.Close(False)
        finally:
            self._com_quit(app)
        return {'ok': True, 'local': local, 'z': record['file'],
                'msg': f"第{record['row']}行已修改，修改字段已标红"}

    def cancel(self, record):
        local = os.path.join(DESKTOP, 'schedule_temp.xlsx')
        shutil.copy2(record['file'], local)
        app = self._com_app()
        try:
            wb = app.Workbooks.Open(os.path.abspath(local))
            ws = wb.Sheets(record['sheet'])
            mc = min(ws.UsedRange.Columns.Count + ws.UsedRange.Column, 100)
            _deleted = self._do_cancel_com(wb, ws, record['row'], mc)
            wb.Save()
            wb.Close(False)
        finally:
            self._com_quit(app)
        if _deleted:
            _msg = f"第{record['row']}行已移至取消订单Sheet"
        else:
            _msg = f"第{record['row']}行已标记取消（红字无填充）"
        return {'ok': True, 'local': local, 'z': record['file'], 'msg': _msg}

    def save_z(self, local, z):
        self._try_save_z(local, z)
        return {'ok': True, 'msg': '已保存到Z盘'}

    @staticmethod
    def _classify_save_error(err_str):
        """统一分类保存错误：占用/权限 vs 其他异常"""
        if not err_str:
            return '保存Z盘失败（未知原因）'
        _lock_kw = ('占用', '编辑', '只读', 'Permission', 'denied')
        if any(kw in err_str for kw in _lock_kw):
            return f'文件被占用: {err_str[:100]}'
        return f'保存失败: {err_str[:100]}'

    def _try_save_z(self, local, z):
        # 第1关：检查~$锁文件（WPS/Excel应用级软锁，比OS权限更可靠）
        is_locked, user = self._check_lock_file(z)
        if is_locked:
            raise PermissionError(f'文件正在被{user or "其他人"}编辑')
        # 第2关：OS级写入权限测试
        try:
            with open(z, 'r+b'):
                pass
        except:
            raise PermissionError('Z盘文件被占用或只读')
        shutil.copy2(local, z)

    # =================== 重试保存 ===================

    def retry_save(self, items):
        ok = []
        still_failed = []
        for item in items:
            try:
                self._try_save_z(item['local'], item['z'])
                ok.append(item)
            except:
                still_failed.append(item)
        return {'ok': ok, 'failed': still_failed}

    # =================== 备份系统 ===================

    def create_backup(self, modified_files=None):
        today = date.today()
        monday = today - timedelta(days=today.weekday())
        saturday = monday + timedelta(days=5)
        folder_name = f"排期备份 {monday.month}.{monday.day}-{saturday.month}.{saturday.day}"
        backup_dir = os.path.join(DESKTOP, folder_name)
        os.makedirs(backup_dir, exist_ok=True)

        date_suffix = f"{today.month}.{today.day}"
        backed_up = []
        files_to_backup = modified_files or self._list_xlsx()

        for fp in files_to_backup:
            fname = os.path.basename(fp)
            base, ext = os.path.splitext(fname)
            new_name = f"{base} {date_suffix}{ext}"
            dest = os.path.join(backup_dir, new_name)
            counter = 0
            while os.path.exists(dest):
                counter += 1
                new_name = f"{base} {date_suffix}-{counter}{ext}"
                dest = os.path.join(backup_dir, new_name)
            try:
                shutil.copy2(fp, dest)
                backed_up.append(new_name)
            except Exception as e:
                backed_up.append(f"{fname} (备份失败: {e})")

        return {'ok': True, 'folder': backup_dir, 'files': backed_up,
                'msg': f'已备份{len(backed_up)}个文件到 {folder_name}'}

    # =================== 历史记录 ===================

    @staticmethod
    def add_history(po, action, detail, files=''):
        os.makedirs(DATA_DIR, exist_ok=True)
        records = []
        if os.path.exists(HISTORY_FILE):
            try:
                with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                    records = json.load(f)
                if not isinstance(records, list):
                    records = []
            except:
                records = []
        records.append({
            'time': datetime.now().strftime('%Y-%m-%d %H:%M'),
            'po': str(po), 'action': action,
            'detail': detail, 'files': files
        })
        with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
            json.dump(records, f, ensure_ascii=False, indent=1)

    @staticmethod
    def get_history():
        if not os.path.exists(HISTORY_FILE):
            return []
        try:
            with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []

    @staticmethod
    def export_history_excel():
        """将历史记录导出为Excel文件，返回文件路径"""
        records = ExcelHandler.get_history()
        if not records:
            return None
        export_dir = os.path.join(DATA_DIR, 'exports')
        os.makedirs(export_dir, exist_ok=True)
        fname = f"操作历史_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fpath = os.path.join(export_dir, fname)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = '操作历史'
        # 表头
        headers = ['时间', '操作类型', 'PO号', '详情', '文件']
        for i, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=i, value=h)
            c.font = openpyxl.styles.Font(bold=True)
        # 数据
        for idx, rec in enumerate(reversed(records), 2):
            ws.cell(row=idx, column=1, value=rec.get('time', ''))
            ws.cell(row=idx, column=2, value=rec.get('action', ''))
            ws.cell(row=idx, column=3, value=rec.get('po', ''))
            ws.cell(row=idx, column=4, value=rec.get('detail', ''))
            ws.cell(row=idx, column=5, value=rec.get('files', ''))
        # 自动列宽
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
        wb.save(fpath)
        return fpath

    # =================== 文件占用检测 ===================

    def _check_lock_file(self, filepath):
        """检查Excel文件是否被占用（~$锁文件 + OS级写锁 双重验证）
        返回 (is_locked, user): is_locked=True表示被占用，user为编辑者名"""
        fname = os.path.basename(filepath)
        lock_path = None
        lock_user = ''
        # 第1关：~$锁文件（WPS/Excel应用级软锁，能识别编辑者）
        # WPS风格：~$ + 完整文件名；Excel风格：~$ + 去掉前2字符
        _dir = os.path.dirname(filepath)
        for _candidate in ('~$' + fname, '~$' + fname[2:] if len(fname) > 2 else None):
            if _candidate and os.path.exists(os.path.join(_dir, _candidate)):
                lock_path = os.path.join(_dir, _candidate)
                break
        if lock_path:
                lock_user = self._read_lock_user(lock_path)
                # ~$锁文件存在即视为占用（Z盘SMB共享OS级open()可能误判为可写）
                # 不做OS写入测试，避免把活跃锁误清理
                if lock_user:
                    return True, lock_user
                # 锁文件存在但读不到用户名：做OS级测试判断是否残留
                if os.path.exists(filepath):
                    try:
                        with open(filepath, 'r+b'):
                            pass
                        # 可写且无用户名 → 残留锁，清理放行
                        try:
                            os.remove(lock_path)
                            logging.info(f"[锁检测] 清理残留锁文件: {os.path.basename(lock_path)}")
                        except:
                            pass
                        return False, ''
                    except PermissionError:
                        return True, lock_user
                    except:
                        pass
                return True, lock_user
        # 无锁文件时：纯OS级写入权限测试（捕获SMB/NTFS文件锁）
        if os.path.exists(filepath):
            try:
                with open(filepath, 'r+b'):
                    pass
            except PermissionError:
                return True, ''
            except:
                pass
        return False, ''

    def check_all_file_status(self):
        """扫描所有排期文件的锁定/占用状态"""
        results = []
        if not os.path.isdir(self.z_path):
            return results

        # 扫描排期目录（含子目录FUGGLER河源排期），收集所有xlsx和~$锁文件
        lock_set = {}  # {还原后的文件名: lock_path}
        all_xlsx = []  # [(filepath, fname)] 含.xlsx0.xlsx
        dirs_to_scan = [self.z_path]
        _fg = os.path.join(self.z_path, 'FUGGLER河源排期')
        if os.path.isdir(_fg):
            dirs_to_scan.append(_fg)
        for d in dirs_to_scan:
            try:
                for item in os.listdir(d):
                    fp = os.path.join(d, item)
                    if item.startswith('~$') and '.xlsx' in item:
                        orig_name = item[2:]
                        orig_name = re.sub(r'\.xlsx[0-9]*\.xlsx$', '.xlsx', orig_name)
                        lock_set[orig_name] = fp
                    elif item.endswith('.xlsx') and not item.startswith('~$'):
                        all_xlsx.append((fp, item))
            except OSError:
                pass

        for fp, fname in all_xlsx:
            status = 'available'
            user = ''
            lock_type = ''

            # 先检查~$锁文件或.xlsx0.xlsx命名（初步判断）
            _has_lock_hint = False
            _lock_path = None
            _is_wps_editing = bool(re.search(r'\.xlsx\d*\.xlsx$', fname))
            if _is_wps_editing:
                _has_lock_hint = True
            else:
                _names_to_try = [fname]
                fname_suffix = fname[2:] if len(fname) > 2 else fname
                _names_to_try.append(fname_suffix)
                for _try_name in _names_to_try:
                    if _try_name in lock_set:
                        _lock_path = lock_set[_try_name]
                        _has_lock_hint = True
                        break

            if _has_lock_hint:
                # 有锁标记时，用实际写入测试验证是否真锁（~$可能是残留的）
                try:
                    _fh = open(fp, 'r+b')
                    _fh.close()
                    # 可写入 → 假锁，标记为可用
                    status = 'available'
                except PermissionError:
                    # 真锁 → 确实被占用
                    status = 'locked'
                    if _is_wps_editing:
                        lock_type = '正在编辑(WPS)'
                    else:
                        lock_type = '正在编辑'
                    if _lock_path:
                        user = self._read_lock_user(_lock_path)
                except OSError:
                    status = 'locked'
                    lock_type = '无法访问'
            elif not os.access(fp, os.W_OK):
                status = 'readonly'
                lock_type = '只读'

            results.append({
                'file': fp, 'fname': fname,
                'status': status, 'user': user, 'lock_type': lock_type
            })

        return results

    def _read_lock_user(self, lock_path):
        """从~$锁文件读取使用者用户名"""
        try:
            with open(lock_path, 'rb') as f:
                raw = f.read(200)
            if not raw:
                return '未知用户'

            # 尝试方法1：第一个字节是长度，后面是用户名
            try:
                name_len = raw[0]
                if 1 <= name_len <= 50:
                    # 尝试GBK（中文Windows常用）
                    name = raw[1:1+name_len*2].decode('utf-16-le', errors='ignore')
                    name = name.split('\x00')[0].strip()
                    if name and len(name) >= 1:
                        return name
            except:
                pass

            # 尝试方法2：直接UTF-16LE
            try:
                text = raw[:54].decode('utf-16-le', errors='ignore')
                name = text.split('\x00')[0].strip()
                if name and len(name) >= 2:
                    return name
            except:
                pass

            # 尝试方法3：GBK/ASCII
            try:
                text = raw[:54].decode('gbk', errors='ignore')
                name = ''.join(c for c in text if c.isprintable()).strip()
                if name:
                    return name
            except:
                pass

            return '未知用户'
        except:
            return '未知用户'

    # =================== 待重试队列 ===================

    @staticmethod
    def get_pending_retries():
        if not os.path.exists(RETRY_FILE):
            return []
        try:
            with open(RETRY_FILE, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []

    @staticmethod
    def save_pending_retries(items):
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(RETRY_FILE, 'w', encoding='utf-8') as f:
            json.dump(items, f, ensure_ascii=False, indent=1)

    def auto_retry_pending(self):
        """自动重试所有待保存的文件"""
        items = self.get_pending_retries()
        if not items:
            return {'ok': [], 'failed': [], 'msg': '无待重试项'}

        ok = []
        still_failed = []
        for item in items:
            try:
                self._try_save_z(item['local'], item['z'])
                ok.append(item)
                self.add_history(
                    item.get('po', ''), 'auto_retry',
                    f"自动重试成功: {item['file']}", item['file']
                )
            except:
                item['retries'] = item.get('retries', 0) + 1
                item['last_retry'] = datetime.now().strftime('%H:%M')
                still_failed.append(item)

        self.save_pending_retries(still_failed)
        return {'ok': ok, 'failed': still_failed,
                'msg': f'成功{len(ok)}个，仍有{len(still_failed)}个待重试'}

    def delete_entries_com(self, entries):
        """删除指定的排期行（支持撤销）
        entries: [{file: z盘路径, sheet: 工作表名, row: 行号, sku: 显示用}]
        返回: {ok: True/False, deleted: [...], failed: [...], undo_id: ...}
        """
        if not entries:
            return {'ok': False, 'error': '没有要删除的条目'}

        os.makedirs(UNDO_DIR, exist_ok=True)
        batch_id = datetime.now().strftime('%Y%m%d-%H%M%S') + '_del'

        # 按文件分组，行号从大到小排列（先删大行号避免偏移）
        file_groups = {}
        for e in entries:
            fkey = e['file']
            if fkey not in file_groups:
                file_groups[fkey] = []
            file_groups[fkey].append(e)
        for fk in file_groups:
            file_groups[fk].sort(key=lambda x: x.get('row', 0), reverse=True)

        deleted = []
        failed = []
        undo_files = []
        app = None
        try:
            app = self._com_app()
            for fkey, group in file_groups.items():
                fname = os.path.basename(fkey)
                # 备份
                undo_fp = os.path.join(UNDO_DIR, f"{batch_id}_{fname}")
                try:
                    import shutil
                    shutil.copy2(fkey, undo_fp)
                except Exception as e:
                    for g in group:
                        failed.append({'sku': g.get('sku', ''), 'reason': f'备份失败: {e}'})
                    continue

                try:
                    wb = app.Workbooks.Open(os.path.abspath(fkey))
                    for e in group:
                        sn = e.get('sheet', '')
                        rn = e.get('row', 0)
                        try:
                            ws = wb.Sheets(sn)
                            # 记录被删行的内容（用于反馈）
                            row_info = {}
                            for ci in range(1, min(20, ws.UsedRange.Columns.Count + 1)):
                                v = ws.Cells(rn, ci).Value
                                if v is not None:
                                    row_info[f'col{ci}'] = str(v)[:50]
                            ws.Rows(rn).Delete()
                            deleted.append({
                                'sku': e.get('sku', ''),
                                'file': fname,
                                'sheet': sn,
                                'row': rn,
                                'row_info': row_info
                            })
                        except Exception as ex:
                            failed.append({'sku': e.get('sku', ''), 'reason': str(ex)[:100]})
                    wb.Save()
                    wb.Close(False)
                    undo_files.append({'name': fname, 'backup': undo_fp, 'z_path': fkey})
                except Exception as ex:
                    for g in group:
                        if not any(f.get('sku') == g.get('sku') for f in failed):
                            failed.append({'sku': g.get('sku', ''), 'reason': str(ex)[:100]})
        finally:
            if app:
                try:
                    app.Quit()
                except:
                    pass

        # 保存撤销记录
        if deleted:
            ops = [{'type': 'delete', 'sku': d['sku'],
                    'detail': f"删除 {d['sheet']} 行{d['row']}"} for d in deleted]
            self._save_undo_entry({
                'id': batch_id,
                'time': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                'operations': ops,
                'files': undo_files,
                'label': f"删除 {len(deleted)} 行",
            })

        return {
            'ok': len(deleted) > 0,
            'deleted': deleted,
            'failed': failed,
            'undo_id': batch_id if deleted else None,
            'msg': f'已删除 {len(deleted)} 行' + (f'，{len(failed)} 行失败' if failed else '')
        }

    def reentry_batch(self, orders):
        """重新入单（删除旧行+重新写入），返回与batch_process相同格式的详细结果"""
        return self.batch_process(orders)

    # =================== 定时重试 ===================

    SCHEDULED_FILE = os.path.join(DATA_DIR, 'scheduled_retries.json')

    @staticmethod
    def get_scheduled_retries():
        fp = os.path.join(DATA_DIR, 'scheduled_retries.json')
        if os.path.exists(fp):
            try:
                with open(fp, 'r', encoding='utf-8') as f:
                    return json.load(f)
            except:
                pass
        return []

    @staticmethod
    def save_scheduled_retries(items):
        os.makedirs(DATA_DIR, exist_ok=True)
        fp = os.path.join(DATA_DIR, 'scheduled_retries.json')
        with open(fp, 'w', encoding='utf-8') as f:
            json.dump(items, f, ensure_ascii=False, indent=1)

    # =================== 撤回操作 ===================

    def undo_selected(self, batch_ids):
        """撤回指定批次的操作：根据batch_id恢复对应备份文件"""
        history = self._load_undo_history()
        if not history:
            return {'error': '没有可撤回的操作'}

        # 找到要撤回的批次
        to_undo = [h for h in history if h['id'] in batch_ids]
        if not to_undo:
            return {'error': '未找到指定的操作记录'}

        restored = []
        failed = []
        undone_ids = []

        for entry in to_undo:
            for finfo in entry.get('files', []):
                backup = finfo.get('backup', '')
                z_path = finfo.get('z_path', '')
                fname = finfo.get('name', '')

                if not backup or not os.path.exists(backup):
                    # 兼容旧格式：尝试不带前缀的备份
                    old_backup = os.path.join(UNDO_DIR, fname)
                    if os.path.exists(old_backup):
                        backup = old_backup
                    else:
                        failed.append({'file': fname, 'reason': '备份文件不存在'})
                        continue

                if not z_path or not os.path.exists(z_path):
                    # 尝试从z_path构建
                    z_path = os.path.join(self.z_path, fname)
                    if not os.path.exists(z_path):
                        failed.append({'file': fname, 'reason': '目标文件不存在'})
                        continue

                try:
                    with open(z_path, 'r+b'):
                        pass
                    shutil.copy2(backup, z_path)
                    restored.append(fname)
                except:
                    failed.append({'file': fname, 'reason': '文件被占用'})

            if not any(f['file'] == fi.get('name') for fi in entry.get('files', []) for f in failed):
                undone_ids.append(entry['id'])

        # 从历史中移除已成功撤回的批次，并清理备份文件
        if undone_ids:
            new_history = []
            for h in history:
                if h['id'] in undone_ids:
                    # 清理备份文件
                    for finfo in h.get('files', []):
                        bp = finfo.get('backup', '')
                        if bp and os.path.exists(bp):
                            try:
                                os.remove(bp)
                            except:
                                pass
                else:
                    new_history.append(h)
            self._write_undo_history(new_history)

        return {
            'ok': True, 'restored': restored, 'failed': failed,
            'undone_ids': undone_ids,
            'msg': f'已撤回 {len(restored)} 个文件' +
                   (f'，{len(failed)} 个失败' if failed else '')
        }

    def undo_last_batch(self):
        """兼容旧接口：撤回最近一次操作"""
        history = self._load_undo_history()
        if history:
            return self.undo_selected([history[-1]['id']])
        # 兼容旧格式undo目录
        if not os.path.isdir(UNDO_DIR):
            return {'error': '没有可撤回的操作'}
        undo_files = [f for f in os.listdir(UNDO_DIR)
                      if f.endswith('.xlsx') and not f.startswith('~$')
                      and not re.match(r'\d{8}-\d{6}_', f)]
        if not undo_files:
            return {'error': '没有可撤回的操作'}
        restored = []
        failed = []
        for fname in undo_files:
            undo_fp = os.path.join(UNDO_DIR, fname)
            z_fp = os.path.join(self.z_path, fname)
            if not os.path.exists(z_fp):
                failed.append({'file': fname, 'reason': '目标文件不存在'})
                continue
            try:
                with open(z_fp, 'r+b'):
                    pass
                shutil.copy2(undo_fp, z_fp)
                restored.append(fname)
            except:
                failed.append({'file': fname, 'reason': '文件被占用'})
        if restored and not failed:
            for f in os.listdir(UNDO_DIR):
                try:
                    os.remove(os.path.join(UNDO_DIR, f))
                except:
                    pass
        return {'ok': True, 'restored': restored, 'failed': failed,
                'msg': f'已撤回 {len(restored)} 个文件' + (f'，{len(failed)} 个失败' if failed else '')}

    def delete_undo_records(self, batch_ids=None):
        """删除撤回记录（仅从列表移除+清理备份文件，不恢复排期）
        batch_ids=None时删除全部记录"""
        history = self._load_undo_history()
        if not history:
            return {'ok': True, 'deleted': 0, 'msg': '没有可删除的记录'}
        if batch_ids is None:
            to_delete = history
            new_history = []
        else:
            id_set = set(batch_ids)
            to_delete = [h for h in history if h['id'] in id_set]
            new_history = [h for h in history if h['id'] not in id_set]
        # 清理备份文件
        for entry in to_delete:
            for finfo in entry.get('files', []):
                bp = finfo.get('backup', '')
                if bp and os.path.exists(bp):
                    try:
                        os.remove(bp)
                    except Exception:
                        pass
        self._write_undo_history(new_history)
        return {'ok': True, 'deleted': len(to_delete),
                'msg': f'已删除 {len(to_delete)} 条记录'}

    def _load_undo_history(self):
        if not os.path.exists(UNDO_HISTORY):
            return []
        try:
            with open(UNDO_HISTORY, 'r', encoding='utf-8') as f:
                return json.load(f)
        except:
            return []

    def _write_undo_history(self, history):
        os.makedirs(DATA_DIR, exist_ok=True)
        with open(UNDO_HISTORY, 'w', encoding='utf-8') as f:
            json.dump(history, f, ensure_ascii=False, indent=1)

    @staticmethod
    def get_undo_info():
        """获取可撤回的操作列表（含历史详情）"""
        history = []
        if os.path.exists(UNDO_HISTORY):
            try:
                with open(UNDO_HISTORY, 'r', encoding='utf-8') as f:
                    history = json.load(f)
            except:
                pass

        # 检查每个批次的备份文件是否仍然存在
        valid = []
        for h in history:
            has_backup = False
            for finfo in h.get('files', []):
                bp = finfo.get('backup', '')
                if bp and os.path.exists(bp):
                    has_backup = True
                    break
            if has_backup:
                valid.append(h)

        if not valid:
            # 兼容旧格式
            if os.path.isdir(UNDO_DIR):
                old_files = [f for f in os.listdir(UNDO_DIR)
                             if f.endswith('.xlsx') and not f.startswith('~$')
                             and not re.match(r'\d{8}-\d{6}_', f)]
                if old_files:
                    latest = max(os.path.getmtime(os.path.join(UNDO_DIR, f)) for f in old_files)
                    return {
                        'available': True,
                        'batches': [{
                            'id': 'legacy',
                            'time': datetime.fromtimestamp(latest).strftime('%Y-%m-%d %H:%M:%S'),
                            'label': f'上次操作（{len(old_files)}个文件）',
                            'operations': [],
                            'files': [{'name': f} for f in old_files],
                        }],
                        'count': 1,
                        # 兼容旧UI
                        'files': old_files,
                        'time': datetime.fromtimestamp(latest).strftime('%Y-%m-%d %H:%M'),
                    }
            return {'available': False, 'batches': [], 'count': 0}

        return {
            'available': True,
            'batches': list(reversed(valid)),  # 最新的在前
            'count': len(valid),
            # 兼容旧UI
            'files': [f['name'] for h in valid for f in h.get('files', [])],
            'time': valid[-1]['time'][:16] if valid else '',
        }

    # =================== 总排期操作 ===================

    def find_master_schedule(self):
        """查找总排期文件"""
        for fp in self._list_xlsx():
            fn = os.path.basename(fp)
            if '总' in fn:
                return fp
        return None

    def scan_yellow_rows(self, use_cache=True, progress_callback=None):
        """扫描所有分排期文件中的黄色填充行（带缓存 + 进度回调）"""
        global _yellow_cache
        results = []
        files = [fp for fp in self._list_xlsx() if '总' not in os.path.basename(fp)]
        total = len(files)

        for idx, fp in enumerate(files):
            fn = os.path.basename(fp)
            if progress_callback:
                progress_callback(idx + 1, total, fn)

            # 缓存检查：文件未修改则复用
            try:
                mtime = os.path.getmtime(fp)
            except:
                continue
            if use_cache and fp in _yellow_cache and _yellow_cache[fp]['mtime'] == mtime:
                results.extend(_yellow_cache[fp]['rows'])
                continue

            # 需要扫描此文件
            file_rows = []
            try:
                wb = openpyxl.load_workbook(fp, data_only=True)
                for sn in wb.sheetnames:
                    sn_lower = sn.lower()
                    if '取消' in sn or '明细' in sn or 'ma' in sn_lower or '对应' in sn or '旧' in sn:
                        continue
                    ws = wb[sn]
                    row_count = 0
                    for row in ws.iter_rows(min_row=2, max_col=30):
                        row_count += 1
                        if row_count > 2000:
                            break
                        if not any(c.value for c in row[:6]):
                            continue
                        is_yellow = False
                        for c in row[:6]:
                            if c.value is not None:
                                is_yellow = _is_yellow_fill(c)
                                break
                        if not is_yellow:
                            is_yellow = _is_yellow_fill(row[0])
                        if is_yellow:
                            data = {}
                            for c in row:
                                cl = openpyxl.utils.get_column_letter(c.column)
                                v = c.value
                                if isinstance(v, datetime):
                                    v = v.strftime('%Y-%m-%d')
                                data[cl] = v
                            file_rows.append({
                                'file': fp, 'fname': fn, 'sheet': sn,
                                'row': row[0].row, 'data': data
                            })
                wb.close()
            except:
                continue
            # 更新缓存
            _yellow_cache[fp] = {'mtime': mtime, 'rows': file_rows}
            results.extend(file_rows)
        return results

    def _read_headers(self, fp, sheet_name=None):
        """读取文件表头行，返回 {列字母: 表头名称}"""
        try:
            wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
            sn = sheet_name or wb.sheetnames[0]
            ws = wb[sn]
            headers = {}
            for row in ws.iter_rows(min_row=1, max_row=4, max_col=30):
                for cell in row:
                    if cell.value and str(cell.value).strip():
                        cl = openpyxl.utils.get_column_letter(cell.column)
                        if cl not in headers:
                            headers[cl] = str(cell.value).strip()
                if len(headers) >= 5:
                    break
            wb.close()
            return headers
        except:
            return {}

    def _build_column_mapping(self, src_headers, dst_headers):
        """构建 分排期列→总排期列 的映射"""
        dst_name_col = {v: k for k, v in dst_headers.items()}
        used_dst = set()
        mapping = {}

        # 特殊别名表
        ALIASES = {
            'ZURU PO NO#': 'PO号', 'PO NO.': 'PO号', 'PO NUMBER': 'PO号',
            'SKU': '系统货号', 'ITEM CODE': '系统货号',
            'ITEM#': '货号#', 'ITEM NO.': '货号#',
            '货品名称': '中文名', '品名': '中文名',
            '出货日期': '预计船期', '船期': '预计船期', '走货日期': '预计船期',
            '验货日期': '预计验货日期',
        }

        for src_col, src_name in src_headers.items():
            sn = src_name.strip()
            # 1. 精确匹配
            if sn in dst_name_col and dst_name_col[sn] not in used_dst:
                mapping[src_col] = dst_name_col[sn]
                used_dst.add(dst_name_col[sn])
                continue
            # 2. 别名匹配
            alias_target = ALIASES.get(sn, '')
            if alias_target and alias_target in dst_name_col and dst_name_col[alias_target] not in used_dst:
                mapping[src_col] = dst_name_col[alias_target]
                used_dst.add(dst_name_col[alias_target])
                continue
            # 3. 包含匹配
            for dn, dc in dst_name_col.items():
                if dc in used_dst:
                    continue
                if sn in dn or dn in sn:
                    mapping[src_col] = dc
                    used_dst.add(dc)
                    break

        return mapping

    def copy_to_master(self, yellow_rows=None):
        """将黄色填充行复制到总排期"""
        master_fp = self.find_master_schedule()
        if not master_fp:
            return {'error': '未找到总排期文件（文件名需包含"总"字）'}
        try:
            with open(master_fp, 'r+b'):
                pass
        except:
            return {'error': '总排期文件被占用或只读，无法写入'}

        if yellow_rows is None:
            yellow_rows = self.scan_yellow_rows()
        if not yellow_rows:
            return {'error': '未找到黄色填充的行'}

        master_headers = self._read_headers(master_fp)
        if not master_headers:
            return {'error': '无法读取总排期表头'}

        src_cache = {}
        map_cache = {}

        app = None
        try:
            app = self._com_app()
            wb = app.Workbooks.Open(os.path.abspath(master_fp))
            ws = wb.Sheets(1)

            try:
                last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
                if last_row < 2:
                    last_row = 2
            except:
                last_row = 2

            insert_row = last_row + 1
            copied = 0
            mc = max(_col_num(c) for c in master_headers.keys()) if master_headers else 20

            for yr in yellow_rows:
                src_key = yr['file'] + '|' + yr['sheet']
                if src_key not in src_cache:
                    src_cache[src_key] = self._read_headers(yr['file'], yr['sheet'])
                    map_cache[src_key] = self._build_column_mapping(
                        src_cache[src_key], master_headers)

                col_map = map_cache[src_key]
                has_data = False

                for src_col, value in yr['data'].items():
                    if value is None:
                        continue
                    dst_col = col_map.get(src_col)
                    if not dst_col:
                        continue
                    dst_num = _col_num(dst_col)
                    cell = ws.Cells(insert_row, dst_num)
                    if isinstance(value, str) and re.match(r'\d{4}-\d{2}-\d{2}', value):
                        try:
                            dt = datetime.strptime(value, '%Y-%m-%d')
                            cell.Value = dt
                            cell.NumberFormat = 'yyyy/m/d'
                        except:
                            cell.Value = value
                    else:
                        cell.Value = value
                    has_data = True

                if has_data:
                    rng = ws.Range(ws.Cells(insert_row, 1), ws.Cells(insert_row, mc))
                    rng.Interior.Color = YELLOW_COM
                    insert_row += 1
                    copied += 1

            wb.Save()
            wb.Close(False)
        finally:
            self._com_quit(app)

        return {
            'ok': True, 'copied': copied,
            'master_file': os.path.basename(master_fp),
            'msg': f'已复制 {copied} 行到总排期（{os.path.basename(master_fp)}）'
        }

    def clear_master_yellow(self):
        """清除总排期中的黄色填充"""
        master_fp = self.find_master_schedule()
        if not master_fp:
            return {'error': '未找到总排期文件'}
        try:
            with open(master_fp, 'r+b'):
                pass
        except:
            return {'error': '总排期文件被占用或只读'}

        app = None
        try:
            app = self._com_app()
            wb = app.Workbooks.Open(os.path.abspath(master_fp))
            ws = wb.Sheets(1)

            try:
                last_row = ws.Cells(ws.Rows.Count, 1).End(-4162).Row
            except:
                last_row = 100

            mc = min(ws.UsedRange.Columns.Count + ws.UsedRange.Column, 100)
            cleared = 0

            for r in range(2, last_row + 1):
                try:
                    color = ws.Cells(r, 1).Interior.Color
                    rc = color % 256
                    gc = (color // 256) % 256
                    bc = (color // 65536) % 256
                    if rc > 200 and gc > 180 and bc < 100:
                        rng = ws.Range(ws.Cells(r, 1), ws.Cells(r, mc))
                        rng.Interior.Pattern = -4142  # xlNone
                        cleared += 1
                except:
                    continue

            wb.Save()
            wb.Close(False)
        finally:
            self._com_quit(app)

        return {
            'ok': True, 'cleared': cleared,
            'msg': f'已清除 {cleared} 行的黄色填充'
        }

    # =================== 排期文件列表（手动选择用）===================

    def list_schedule_files(self):
        """列出所有排期文件及其Sheet，供手动选择"""
        result = []
        for fp in self._list_xlsx():
            fn = os.path.basename(fp)
            if _should_skip_file(fn):
                continue
            sheets = []
            try:
                wb = openpyxl.load_workbook(fp, read_only=True)
                for sn in wb.sheetnames:
                    if '取消' not in sn and '明细' not in sn and '旧' not in sn:
                        sheets.append(sn)
                wb.close()
            except:
                sheets = ['Sheet1']
            result.append({'file': fp, 'fname': fn, 'sheets': sheets})
        return result

    def manual_find_ref(self, filepath, sheet_name):
        """在指定文件+Sheet中查找参考行（最后一个有数据的行）"""
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[sheet_name]
            ref = 2
            for row in ws.iter_rows(min_row=2, max_col=10):
                if any(c.value for c in row[:6]):
                    ref = row[0].row
            wb.close()
            return {'file': filepath, 'fname': os.path.basename(filepath),
                    'sheet': sheet_name, 'ref': ref, 'cnt': 0, 'mcol': 30}
        except Exception as e:
            return {'error': str(e)}


# =================== 邮件集成 ===================

class EmailHandler:
    """通过IMAP读取邮箱附件PDF"""

    def __init__(self, config):
        self.server = config.get('email_server', '')
        self.port = int(config.get('email_port', 993))
        self.user = config.get('email_user', '')
        self.password = config.get('email_password', '')
        self.ssl = config.get('email_ssl', True)

    def check_new_emails(self, folder='INBOX', limit=20):
        """检查邮箱中的新邮件，返回含PDF附件的邮件列表"""
        if not self.server or not self.user:
            return {'error': '邮箱未配置，请在设置页面配置IMAP信息'}
        import imaplib
        import email
        from email.header import decode_header

        results = []
        try:
            if self.ssl:
                mail = imaplib.IMAP4_SSL(self.server, self.port)
            else:
                mail = imaplib.IMAP4(self.server, self.port)
            mail.login(self.user, self.password)
            mail.select(folder)

            # 搜索未读邮件
            status, messages = mail.search(None, 'UNSEEN')
            if status != 'OK':
                mail.logout()
                return {'emails': [], 'msg': '无新邮件'}

            msg_ids = messages[0].split()[-limit:]  # 只取最近的
            for mid in reversed(msg_ids):
                status, msg_data = mail.fetch(mid, '(RFC822)')
                if status != 'OK':
                    continue
                msg = email.message_from_bytes(msg_data[0][1])

                # 解码主题
                subject = ''
                raw_subject = msg.get('Subject', '')
                if raw_subject:
                    decoded = decode_header(raw_subject)
                    subject = ''.join(
                        part.decode(enc or 'utf-8') if isinstance(part, bytes) else part
                        for part, enc in decoded
                    )

                from_addr = msg.get('From', '')
                date_str = msg.get('Date', '')

                # 查找PDF附件
                attachments = []
                for part in msg.walk():
                    if part.get_content_maintype() == 'multipart':
                        continue
                    fname = part.get_filename()
                    if fname:
                        decoded_fname = decode_header(fname)
                        fname = ''.join(
                            p.decode(enc or 'utf-8') if isinstance(p, bytes) else p
                            for p, enc in decoded_fname
                        )
                        if fname.lower().endswith('.pdf'):
                            attachments.append({
                                'filename': fname,
                                'size': len(part.get_payload(decode=True) or b''),
                                'msg_id': mid.decode() if isinstance(mid, bytes) else str(mid),
                            })

                if attachments:
                    results.append({
                        'subject': subject, 'from': from_addr, 'date': date_str,
                        'msg_id': mid.decode() if isinstance(mid, bytes) else str(mid),
                        'attachments': attachments,
                    })

            mail.logout()
            return {'emails': results, 'count': len(results)}

        except Exception as e:
            return {'error': f'邮箱连接失败: {str(e)}'}

    def download_attachment(self, msg_id, filename, save_dir):
        """下载指定邮件的PDF附件"""
        if not self.server or not self.user:
            return None
        import imaplib
        import email
        from email.header import decode_header

        os.makedirs(save_dir, exist_ok=True)
        try:
            if self.ssl:
                mail = imaplib.IMAP4_SSL(self.server, self.port)
            else:
                mail = imaplib.IMAP4(self.server, self.port)
            mail.login(self.user, self.password)
            mail.select('INBOX')

            status, msg_data = mail.fetch(msg_id.encode() if isinstance(msg_id, str) else msg_id,
                                          '(RFC822)')
            if status != 'OK':
                mail.logout()
                return None

            msg = email.message_from_bytes(msg_data[0][1])
            for part in msg.walk():
                fname = part.get_filename()
                if fname:
                    decoded_fname = decode_header(fname)
                    fname = ''.join(
                        p.decode(enc or 'utf-8') if isinstance(p, bytes) else p
                        for p, enc in decoded_fname
                    )
                    if fname == filename:
                        content = part.get_payload(decode=True)
                        save_path = os.path.join(save_dir, fname)
                        with open(save_path, 'wb') as f:
                            f.write(content)
                        mail.logout()
                        return save_path

            mail.logout()
            return None
        except Exception as e:
            logging.error(f"[邮件] 下载附件失败: {e}")
            return None
